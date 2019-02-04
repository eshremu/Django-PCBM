"""
Views used for downloading configuration, baseline, pricing, and search result
data to Excel file.
"""

from django.http import HttpResponse, Http404
from django.core.urlresolvers import reverse
from django.core.mail import EmailMultiAlternatives
from django.contrib.staticfiles.finders import find
from django.contrib.auth.models import User

from BoMConfig.models import Header, ConfigLine, Baseline, Baseline_Revision, \
    DistroList, REF_PROGRAM, PricingObject
from BoMConfig.templatetags.bomconfig_customtemplatetags import searchscramble
from BoMConfig.utils import GenerateRevisionSummary, GrabValue, \
    HeaderComparison, RevisionCompare, TitleShorten, char_to_num
from BoMConfig.views.configuration import BuildDataArray
from BoMConfig.views.pricing import PricingOverviewLists



import datetime
from io import BytesIO
from itertools import chain
import openpyxl
from openpyxl import utils
from openpyxl.styles import Font, Color, colors, Border, Alignment, Side, \
    borders, GradientFill
from openpyxl.comments import Comment
from urllib.parse import unquote
import zipfile
import re


def WriteConfigToFile(oHeader, sHyperlinkURL=''):
    """
    Writes a single Header/Configuration to a file.  Uses an existing template.
    :param oHeader: Header object being written to file
    :param sHyperlinkURL: String containing hyperlink details that will allow
    the file to direct a user back to this record in the tool
    :return: An OpenPyXL file object (acts as data stream)
    """

    # Open the template
    oFile = openpyxl.load_workbook(
        find('BoMConfig/PSM BoM Upload Template PA5 FORMULAS.xlsx')
    )

    # Populate Header data
    oFile.active = oFile.sheetnames.index('1) BOM Config Header')
    oFile.active['B1'] = oHeader.person_responsible
    oFile.active['B2'] = oHeader.react_request
    oFile.active['B3'] = oHeader.bom_request_type.name
    oFile.active['B4'] = oHeader.customer_unit.name
    oFile.active['B5'] = oHeader.customer_name
    oFile.active['B6'] = oHeader.sales_office
    oFile.active['B7'] = oHeader.sales_group
    oFile.active['B8'] = oHeader.sold_to_party
    oFile.active['B9'] = oHeader.ship_to_party
    oFile.active['B10'] = oHeader.bill_to_party
    oFile.active['B11'] = oHeader.ericsson_contract
    oFile.active['B12'] = oHeader.payment_terms

    oFile.active['B14'] = oHeader.projected_cutover
    oFile.active['B15'] = oHeader.program.name if oHeader.program else None
    oFile.active['B16'] = oHeader.configuration_designation
    oFile.active['B17'] = oHeader.customer_designation
    oFile.active['B18'] = oHeader.technology.name if oHeader.technology else \
        None
    oFile.active['B19'] = oHeader.product_area1.name if oHeader.product_area1 \
        else None
    oFile.active['B20'] = oHeader.product_area2.name if oHeader.product_area2 \
        else None
    oFile.active['B21'] = oHeader.radio_frequency.name if \
        oHeader.radio_frequency else None
    oFile.active['B22'] = oHeader.radio_band.name if oHeader.radio_band else \
        None
    oFile.active['B23'] = oHeader.optional_free_text1
    oFile.active['B24'] = oHeader.optional_free_text2
    oFile.active['B25'] = oHeader.optional_free_text3
    oFile.active['B26'] = (oHeader.inquiry_site_template * -1) if \
        oHeader.inquiry_site_template and oHeader.inquiry_site_template < -1 \
        else oHeader.inquiry_site_template
    oFile.active['B27'] = oHeader.readiness_complete / 100 if \
        oHeader.readiness_complete else oHeader.readiness_complete
    oFile.active['B28'] = ('X' if oHeader.complete_delivery else None)
    oFile.active['B29'] = ('X' if oHeader.no_zip_routing else None)
    oFile.active['B30'] = oHeader.valid_to_date
    oFile.active['B31'] = oHeader.valid_from_date
    oFile.active['B32'] = oHeader.shipping_condition

    oFile.active['B34'] = oHeader.baseline_impacted
    oFile.active['B35'] = oHeader.model
    oFile.active['B36'] = oHeader.model_description
    oFile.active['B37'] = oHeader.model_replaced
    oFile.active['B38'] = oHeader.initial_revision

    oFile.active['B40'] = oHeader.configuration_status.name

    oFile.active['B42'] = oHeader.workgroup
    oFile.active['B43'] = oHeader.name

    oFile.active['B45'] = ('X' if oHeader.pick_list else None)
    oFile.active.sheet_view.showGridLines = False

    # Populate Config Entry tab
    oFile.active = oFile.sheetnames.index('2) BOM Config Entry')
    oFile.active.sheet_view.showGridLines = False

    if oHeader.configuration:
        oFile.active['D5'] = ('X' if oHeader.configuration.ready_for_forecast
                              else None)
        oFile.active['D8'] = ('X' if oHeader.configuration.PSM_on_hold else
                              None)
        oFile.active['N4'] = ('X' if
                              oHeader.configuration.internal_external_linkage
                              else None)
        oFile.active['Q4'] = (
            '$ ' + str(oHeader.configuration.zpru_total) if
            oHeader.configuration.zpru_total else '$ xx,xxx.xx'
        )

        iRow = 13
        aConfigLines = ConfigLine.objects.filter(
            config=oHeader.configuration).order_by('line_number')
        aConfigLines = sorted(
            aConfigLines,
            key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')]
        )
        for oConfigLine in aConfigLines:
            oFile.active['A' + str(iRow)] = oConfigLine.line_number
            oFile.active['B' + str(iRow)] = \
                (
                    '..' if oConfigLine.is_grandchild else '.' if
                    oConfigLine.is_child else ''
                ) + oConfigLine.part.base.product_number
            oFile.active['C' + str(iRow)] = oConfigLine.part.product_description
            oFile.active['D' + str(iRow)] = oConfigLine.order_qty
            oFile.active['E' + str(iRow)] = \
                oConfigLine.part.base.unit_of_measure
            oFile.active['F' + str(iRow)] = oConfigLine.plant
            oFile.active['G' + str(iRow)] = oConfigLine.sloc
            oFile.active['H' + str(iRow)] = oConfigLine.item_category
            oFile.active['I' + str(iRow)] = oConfigLine.pcode
            oFile.active['J' + str(iRow)] = oConfigLine.commodity_type
            oFile.active['K' + str(iRow)] = oConfigLine.package_type
            oFile.active['L' + str(iRow)] = str(oConfigLine.spud)
            oFile.active['M' + str(iRow)] = oConfigLine.REcode
            oFile.active['N' + str(iRow)] = oConfigLine.mu_flag
            oFile.active['O' + str(iRow)] = oConfigLine.x_plant
            oFile.active['P' + str(iRow)] = oConfigLine.internal_notes
            oFile.active['Q' + str(iRow)] = '' if not oHeader.pick_list and str(
                oConfigLine.line_number) != '10' else str(
                GrabValue(oConfigLine,
                          'linepricing.override_price',
                          GrabValue(oConfigLine,
                                    'linepricing.pricing_object.unit_price',
                                    '')
                          )
            )

            oFile.active['R' + str(iRow)] = oConfigLine.higher_level_item
            oFile.active['S' + str(iRow)] = oConfigLine.material_group_5
            oFile.active['T' + str(iRow)] = oConfigLine.purchase_order_item_num
            oFile.active['U' + str(iRow)] = oConfigLine.condition_type
            oFile.active['V' + str(iRow)] = oConfigLine.amount
            oFile.active['W' + str(iRow)] = oConfigLine.traceability_req
            oFile.active['X' + str(iRow)] = oConfigLine.customer_asset
            oFile.active['Y' + str(iRow)] = oConfigLine.customer_asset_tagging
            oFile.active['Z' + str(iRow)] = oConfigLine.customer_number
            oFile.active['AA' + str(iRow)] = oConfigLine.sec_customer_number
            oFile.active['AB' + str(iRow)] = oConfigLine.vendor_article_number
            oFile.active['AC' + str(iRow)] = oConfigLine.comments
            oFile.active['AD' + str(iRow)] = oConfigLine.additional_ref
            iRow += 1
            # end for
    # end if

    # Populate Config ToC tab
    oFile.active = oFile.sheetnames.index('2a) BoM Config ToC')
    oFile.active.sheet_view.showGridLines = False
    if oHeader.configuration:
        oFile.active['D5'] = ('X' if oHeader.configuration.ready_for_forecast
                              else None)
        oFile.active['D8'] = (
            'X' if oHeader.configuration.PSM_on_hold else None)
        oFile.active['A13'] = oHeader.configuration_designation
        oFile.active['B13'] = oHeader.customer_designation or None
        oFile.active['C13'] = oHeader.technology.name if oHeader.technology \
            else None
        oFile.active['D13'] = oHeader.product_area1.name if \
            oHeader.product_area1 else None
        oFile.active['E13'] = oHeader.product_area2.name if \
            oHeader.product_area2 else None
        oFile.active['F13'] = oHeader.model or None
        oFile.active['G13'] = oHeader.model_description or None
        oFile.active['H13'] = oHeader.model_replaced or None
        oFile.active['J13'] = oHeader.bom_request_type.name
        oFile.active['K13'] = oHeader.configuration_status.name or None
        oFile.active['L13'] = oHeader.inquiry_site_template if str(
            oHeader.inquiry_site_template).startswith('1') else \
            oHeader.inquiry_site_template * -1 if \
            oHeader.inquiry_site_template and \
            oHeader.inquiry_site_template < -1 and str(
                    oHeader.inquiry_site_template).startswith('-1') else None
        oFile.active['M13'] = oHeader.inquiry_site_template if str(
            oHeader.inquiry_site_template).startswith('4') else \
            oHeader.inquiry_site_template * -1 if \
            oHeader.inquiry_site_template and \
            oHeader.inquiry_site_template < -1 and str(
                    oHeader.inquiry_site_template).startswith('-4') else None
        oFile.active['N13'] = oHeader.internal_notes
        oFile.active['O13'] = oHeader.external_notes
        if sHyperlinkURL:
            oFile.active['I13'] = '=HYPERLINK("' + sHyperlinkURL + '","' + \
                                  sHyperlinkURL + '")'
    # end if

    # Populate Config Revision tab
    oFile.active = oFile.sheetnames.index('2b) BoM Config Revision')
    oFile.active.sheet_view.showGridLines = False
    if oHeader.configuration:
        iRow = 13
        aRevData = BuildDataArray(oHeader, revision=True)
        for oRev in aRevData:
            oFile.active['A' + str(iRow)] = oRev['0']
            oFile.active['B' + str(iRow)] = oRev['1']
            oFile.active['C' + str(iRow)] = oRev['2']
            oFile.active['D' + str(iRow)] = oRev['3']
            oFile.active['E' + str(iRow)] = oRev['4']
            oFile.active['F' + str(iRow)] = oRev['5']
            oFile.active['G' + str(iRow)] = oRev['6']
            # end for
    # end if

    # Populate SAP inquiry tab
    oFile.active = oFile.sheetnames.index('3a) SAP Inquiry Creation')
    oFile.active.sheet_view.showGridLines = False
    if oHeader.configuration:
        oFile.active['I4'] = (
            'X' if oHeader.configuration.internal_external_linkage else None)
        oFile.active['M4'] = (
            '$ ' + str(oHeader.configuration.zpru_total) if
            oHeader.configuration.zpru_total else '$ xx,xxx.xx'
        )

        iRow = 13
        aConfigLines = ConfigLine.objects.filter(
            config=oHeader.configuration).order_by('line_number')
        aConfigLines = sorted(
            aConfigLines,
            key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')]
        )
        for oConfigLine in aConfigLines:
            if '.' in oConfigLine.line_number:
                continue
            oFile.active['A' + str(iRow)] = oConfigLine.line_number
            oFile.active['B' + str(iRow)] = oConfigLine.part.base.product_number
            oFile.active['C' + str(iRow)] = oConfigLine.part.product_description
            oFile.active['D' + str(iRow)] = oConfigLine.order_qty
            oFile.active['E' + str(iRow)] = oConfigLine.plant
            oFile.active['F' + str(iRow)] = oConfigLine.sloc
            oFile.active['G' + str(iRow)] = oConfigLine.item_category
            oFile.active['H' + str(iRow)] = str(
                oConfigLine.linepricing.override_price) if \
                str(oConfigLine.line_number) == '10' and hasattr(
                    oConfigLine, 'linepricing') and \
                oConfigLine.linepricing.override_price else \
                oConfigLine.linepricing.pricing_object.unit_price if GrabValue(
                    oConfigLine, 'linepricing.pricing_object') else ''
            oFile.active['J' + str(iRow)] = oConfigLine.material_group_5
            oFile.active['K' + str(iRow)] = oConfigLine.purchase_order_item_num
            oFile.active['L' + str(iRow)] = oConfigLine.condition_type
            oFile.active['M' + str(iRow)] = oConfigLine.amount
            iRow += 1
            # end for
    # end if

    # Populate SAP Site Template tab
    oFile.active = oFile.sheetnames.index('3b) SAP ST Creation')
    oFile.active.sheet_view.showGridLines = False
    if oHeader.configuration:
        iRow = 13
        aConfigLines = ConfigLine.objects.filter(
            config=oHeader.configuration).order_by('line_number')
        aConfigLines = sorted(
            aConfigLines,
            key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')]
        )
        for oConfigLine in aConfigLines:
            if '.' in oConfigLine.line_number:
                continue

            oFile.active['A' + str(iRow)] = oConfigLine.line_number
            oFile.active['B' + str(iRow)] = oConfigLine.part.base.product_number
            oFile.active['C' + str(iRow)] = oConfigLine.part.product_description
            oFile.active['D' + str(iRow)] = oConfigLine.order_qty
            oFile.active['E' + str(iRow)] = oConfigLine.plant
            oFile.active['F' + str(iRow)] = oConfigLine.sloc
            oFile.active['G' + str(iRow)] = oConfigLine.item_category
            oFile.active['H' + str(iRow)] = oConfigLine.higher_level_item
            iRow += 1
            # end for
    # end if

    return oFile


def Download(oRequest, iHeaderId=None):
    """
    View used by user to download a single configuration as an MS Excel file.
    :param oRequest: Django HTTP request object
    :param iHeaderId: Integer storing pk/id of Header object to download
    :return: HttpResponse containing downloaded file
    """

    # Attempt to retrieve the desired Header pk from parameters,
    # request query parameters, or session variables.
    if iHeaderId:
        oHeader = iHeaderId
    else:
        download = oRequest.GET.get('download', None)
        if download:
            oHeader = download[5:-10]
        else:
            oHeader = oRequest.session.get('existing', None)
        # end if

    # If a Header pk was found, download the file
    if oHeader:
        # Get Header object
        oHeader = Header.objects.get(pk=oHeader)

        # generate filename of downlaoded file
        sFileName = oHeader.configuration_designation + (
            '_' + oHeader.program.name if oHeader.program else '') + '.xlsx'

        # Generate file stream
        oFile = WriteConfigToFile(
            oHeader,
            oRequest.build_absolute_uri(reverse('bomconfig:search')) +
            '?config=' + searchscramble(oHeader.pk)
        )

        # Set response headers and save file stream to response
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
            sFileName)
        oFile.save(response)

        return response
    else:
        return HttpResponse()
# end def


def DownloadMultiple(oRequest):
    """
    View to download multiple Headers/Configurations in a compressed archive.
    :param oRequest: Django HTTP request
    :return: HttpResponse containing compressed archive
    """

    # Get list of record to download from URL query string, and remove any
    # duplicates
    records = unquote(oRequest.GET.get('list')).split(',')
    records = list(set(records))

    # If only one record is desired, then perform the download of a single
    # record
    if len(records) == 1:
        return Download(oRequest, records[0])

    # Generate archive name
    # S-10576: Change the header of the tool to ACC :- Changed the tool name from pcbm to acc
    sFilename = 'ACC_{}_{}.zip'.format(
        oRequest.user.username,
        datetime.datetime.now().strftime('%d%m%Y%H%M%S')
    )

    # Create archive object and stream
    zipStream = BytesIO()
    zf = zipfile.ZipFile(zipStream, "w")

    for headID in records:
        # Get Header object, generate file name, and generate filestream
        oHeader = Header.objects.get(id=headID)
        sHeaderName = oHeader.configuration_designation + (
            '_' + oHeader.program.name if oHeader.program else '') + '.xlsx'
        oFile = WriteConfigToFile(
            oHeader,
            oRequest.build_absolute_uri(reverse('bomconfig:search')) +
            '?config=' + searchscramble(oHeader.pk)
        )
        fileStream = BytesIO()
        oFile.save(fileStream)

        # Save filestream to zip archive
        zf.writestr(sHeaderName, fileStream.getvalue())
        fileStream.close()

    # Close archive and write to response
    zf.close()
    response = HttpResponse(
        zipStream.getvalue(), content_type='application/x-zip-compressed')
    response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
        sFilename)
    return response


def DownloadBaseline(oRequest):
    """
    View for downloading a single baseline revision to a file
    :param oRequest: Django HTTP request object
    :return: HttpResponse containing baseline file data
    """
    # Return a 404 error for any request that is NOT a POST request or does not
    # actually contain any POSTed data
    if oRequest.method != 'POST' or not oRequest.POST:
        return Http404()
    # end if

    oBaseline = oRequest.POST['baseline']
    sVersion = oRequest.POST['version']
    sCookie = oRequest.POST['file-cookie']
    sCust = oRequest.POST['customer']

    if oBaseline:
        # Retrieve Baseline object and generate filename
        oBaseline = Baseline.objects.get(title__iexact=oBaseline)
        sFileName = str(Baseline_Revision.objects.get(
            baseline=oBaseline, version=sVersion)) + ".xlsx"

        # Write Baseline to file stream
        oFile = WriteBaselineToFile(oBaseline, sVersion, sCust)
        # Attach filestream to HTTP response
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
            sFileName)
        response.set_cookie('fileMark', sCookie, max_age=60)
        oFile.save(response)

        return response
    else:
        return HttpResponse()
# end def


def DownloadBaselineMaster(oRequest):
    """
    View for downloading an overview of all baselines to a file.  This download
    only contains information about the currently active and in-process revision
    of each baseline.
    :param oRequest: Django HTTP request object
    :return: HttpResponse containing baseline file data
    """
    if oRequest.method != 'POST' or not oRequest.POST:
        return Http404()
    # end if

    sCustomer = oRequest.POST['customer']
    sCookie = oRequest.POST['file-cookie']
    sFileName = "BOM Master File - {}.xlsx".format(
        str(datetime.datetime.now().strftime('%d%b%Y')))

    aTable = []

    # If no customer was specified, we want to get all actual baselines first,
    # so we exclude the pseudo-baseline "No Associated Baseline". When customer
    # is provided, the filtering will ensure the pseudo-baseline is excluded
    if not sCustomer:
        aBaselines = Baseline.objects.exclude(title='No Associated Baseline').exclude(isdeleted=1)
    else:
        aBaselines = Baseline.objects.filter(customer__name=sCustomer).exclude(isdeleted=1)

    # Add the pseudo-baseline to the end of the list to ensure it is displayed
    # last
    oBaseline = Baseline.objects.filter(title='No Associated Baseline').first()
    if oBaseline:
        aBaselines = chain(aBaselines, [oBaseline])

    for oBaseline in aBaselines:
        # Limit to only currently active and in-process revision
        aRevisions = [oBaseline.current_inprocess_version or None,
                      oBaseline.current_active_version or None]
        dTableData = {'baseline': oBaseline, 'revisions': []}

        for sRev in aRevisions:
            if not sRev:
                continue
            # Retrieve all Header objects associated with the baseline and
            # revision
            aHeads = Header.objects.filter(baseline__baseline=oBaseline).filter(
                baseline_version=sRev)

            # Further filter by customer.  This will only affect the Headers
            # retrieved from the pseudo-baseline
            if sCustomer:
                aHeads = aHeads.filter(customer_unit__name=sCustomer)
            aHeads = aHeads.order_by('configuration_status', 'pick_list',
                                     'configuration_designation')
            if aHeads:
                dTableData['revisions'].append(
                    {'revision': Baseline_Revision.objects.get(
                        baseline=oBaseline, version=sRev), 'configs': aHeads})
                # end if
        # end for

        aTable.append(dTableData)
    # end for

    headerFont = Font(name='Arial', size=10, bold=True)
    ipFont = Font(name='Arial', size=10, color='009900')
    activeTitleFont = Font(name='Arial', size=10, color='990000')
    activeFont = Font(name='Arial', size=10)

    centerAlign = Alignment(horizontal='center')
    headerAlign = Alignment(horizontal='center', wrap_text=True)

    # Open file stream, and write header row
    oFile = openpyxl.Workbook()
    oSheet = oFile.active
    oSheet.title = "BOM Master File"

    # oSheet['A1'] = 'Network Info'
    # oSheet['A1'].font = headerFont
    # oSheet['A1'].alignment = headerAlign
    # oSheet.column_dimensions['A'].width = 10
    # S-05748-Remove columns in downloaded all Baselines Master file for MTW Customer( added Scustomer=MTW , to have a new baseline download format)(line 528-626)
    if sCustomer == 'MTW':
        oSheet['B1'] = 'Baseline'
        oSheet['B1'].font = headerFont
        oSheet.column_dimensions['B'].width = 25

        oSheet['C1'] = 'Product Area 2'
        oSheet['C1'].font = headerFont
        oSheet['C1'].alignment = headerAlign
        oSheet.column_dimensions['C'].width = 10

        oSheet['D1'] = 'Status'
        oSheet['D1'].font = headerFont
        oSheet['D1'].alignment = headerAlign
        oSheet.column_dimensions['D'].width = 10

        oSheet['E1'] = 'Article code FAA'
        oSheet['E1'].font = headerFont
        oSheet['E1'].alignment = headerAlign
        oSheet.column_dimensions['E'].width = 20

        oSheet['F1'] = 'Article code ZENG'
        oSheet['F1'].font = headerFont
        oSheet['F1'].alignment = headerAlign
        oSheet.column_dimensions['F'].width = 20

        oSheet['G1'] = 'Configuration'
        oSheet['G1'].font = headerFont
        oSheet['G1'].alignment = headerAlign
        oSheet.column_dimensions['G'].width = 30

        oSheet['H1'] = 'Model/VAN'
        oSheet['H1'].font = headerFont
        oSheet.column_dimensions['H'].width = 30

        oSheet['I1'] = 'Description'
        oSheet['I1'].font = headerFont
        oSheet.column_dimensions['I'].width = 30

        oSheet['J1'] = 'Current Customer Price'
        oSheet['J1'].font = headerFont
        oSheet['J1'].alignment = headerAlign
        oSheet.column_dimensions['J'].width = 15

        oSheet['K1'] = 'Price Year'
        oSheet['K1'].font = headerFont
        oSheet['K1'].alignment = headerAlign
        oSheet.column_dimensions['K'].width = 10

        oSheet['L1'] = 'Site Template/Inquiry'
        oSheet['L1'].font = headerFont
        oSheet['L1'].alignment = headerAlign
        oSheet.column_dimensions['L'].width = 20

        oSheet['M1'] = 'Model Replacing'
        oSheet['M1'].font = headerFont
        oSheet.column_dimensions['M'].width = 30

        oSheet['N1'] = 'Replaced By Model'
        oSheet['N1'].font = headerFont
        oSheet.column_dimensions['N'].width = 30

        oSheet['O1'] = 'Comments'
        oSheet['O1'].font = headerFont
        oSheet.column_dimensions['O'].width = 30
    else:
        oSheet['B1'] = 'Baseline'
        oSheet['B1'].font = headerFont
        oSheet.column_dimensions['B'].width = 25

        oSheet['C1'] = 'Product Area 2'
        oSheet['C1'].font = headerFont
        oSheet['C1'].alignment = headerAlign
        oSheet.column_dimensions['C'].width = 10

        oSheet['D1'] = 'Status'
        oSheet['D1'].font = headerFont
        oSheet['D1'].alignment = headerAlign
        oSheet.column_dimensions['D'].width = 10

        oSheet['E1'] = 'Oracle Item No.'
        oSheet['E1'].font = headerFont
        oSheet['E1'].alignment = headerAlign
        oSheet.column_dimensions['E'].width = 20

        oSheet['F1'] = 'Model'
        oSheet['F1'].font = headerFont
        oSheet.column_dimensions['F'].width = 30

        oSheet['G1'] = 'Description'
        oSheet['G1'].font = headerFont
        oSheet.column_dimensions['G'].width = 40

        oSheet['H1'] = 'Current Customer Price'
        oSheet['H1'].font = headerFont
        oSheet['H1'].alignment = headerAlign
        oSheet.column_dimensions['H'].width = 15

        oSheet['I1'] = 'Price Year'
        oSheet['I1'].font = headerFont
        oSheet['I1'].alignment = headerAlign
        oSheet.column_dimensions['I'].width = 10

        oSheet['J1'] = 'AT&T (or Affiliate) Quotes'
        oSheet['J1'].font = headerFont
        oSheet['J1'].alignment = headerAlign
        oSheet.column_dimensions['J'].width = 15

        oSheet['K1'] = 'KGP Quotes'
        oSheet['K1'].font = headerFont
        oSheet['K1'].alignment = headerAlign
        oSheet.column_dimensions['K'].width = 15

        oSheet['L1'] = 'DTS STs'
        oSheet['L1'].font = headerFont
        oSheet['L1'].alignment = headerAlign
        oSheet.column_dimensions['L'].width = 15

        oSheet['M1'] = 'Model Replacing'
        oSheet['M1'].font = headerFont
        oSheet.column_dimensions['M'].width = 30

        oSheet['N1'] = 'Replaced By Model'
        oSheet['N1'].font = headerFont
        oSheet.column_dimensions['N'].width = 30

        oSheet['O1'] = 'Comments'
        oSheet['O1'].font = headerFont
        oSheet.column_dimensions['O'].width = 50

    iRow = 2
    for dBaseline in aTable:
        if not dBaseline['revisions']:
            continue

        aConfigs = []
        aTitles = []

        for dRevision in dBaseline['revisions']:
            aTitles.insert(
                0,
                str(dRevision['revision'] or dBaseline['baseline']).replace(
                    'Rev ', '').replace('_', ' ')
            )
            aConfigs.extend(dRevision['configs'])
        # end for

        # Write baseline title info
        for idx, sName in enumerate(aTitles):
            if 'No Associated Baseline' in sName and idx > 0:
                continue
            elif 'No Associated Baseline' in sName:
                sName = 'No Associated Baseline'

            oSheet['B' + str(iRow)] = sName
            if not re.match('.+\d{6}C', sName) and 'No Associated Baseline' \
                    not in sName:
                oSheet['B' + str(iRow)].font = ipFont
            else:
                oSheet['B' + str(iRow)].font = activeTitleFont
            iRow += 1
        # end for

        # Write data for each Header in the baseline
        for oHead in aConfigs:
            # added oHead.customer_unit_id!=9 and updated blocks for  S-05748- Remove columns in downloaded all Baselines Master file for MTW Customer(line 693-911)
            if oHead.customer_unit_id!=9:
                oSheet['C' + str(iRow)] = oHead.product_area2.name if \
                    oHead.product_area2 else ''
                oSheet['C' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['C' + str(iRow)].font = ipFont
                else:
                    oSheet['C' + str(iRow)].font = activeFont

                oSheet['D' + str(iRow)] = oHead.configuration_status.name.replace(
                    '/Pending', '').replace('In Process', 'IP')
                oSheet['D' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['D' + str(iRow)].font = ipFont
                else:
                    oSheet['D' + str(iRow)].font = activeFont

                oSheet['E' + str(iRow)] = getattr(
                    oHead.configuration.first_line, 'customer_number', '')
                oSheet['E' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['E' + str(iRow)].font = ipFont
                else:
                    oSheet['E' + str(iRow)].font = activeFont

                oSheet['F' + str(iRow)] = oHead.configuration_designation
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['F' + str(iRow)].font = ipFont
                else:
                    oSheet['F' + str(iRow)].font = activeFont

                oSheet['G' + str(iRow)] = oHead.model_description
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['G' + str(iRow)].font = ipFont
                else:
                    oSheet['G' + str(iRow)].font = activeFont

                oSheet['H' + str(iRow)] = \
                    oHead.configuration.override_net_value or \
                    oHead.configuration.net_value
                oSheet['H' + str(iRow)].number_format = \
                    '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['H' + str(iRow)].font = ipFont
                else:
                    oSheet['H' + str(iRow)].font = activeFont

                oSheet['I' + str(iRow)] = \
                    oHead.headertimetracker_set.first().created_on.strftime('%Y')
                oSheet['I' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['I' + str(iRow)].font = ipFont
                else:
                    oSheet['I' + str(iRow)].font = activeFont

                oSheet['J' + str(iRow)] = (
                    oHead.inquiry_site_template if str(
                        oHead.inquiry_site_template).startswith('1') else
                    oHead.inquiry_site_template * -1 if
                    oHead.inquiry_site_template and
                    oHead.inquiry_site_template < -1 and
                    str(oHead.inquiry_site_template).startswith('-1') else ''
                ) if not (oHead.sold_to_party == 626136 or 'KGP' in
                          str(oHead.customer_name).upper()) else ''
                oSheet['J' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['J' + str(iRow)].font = ipFont
                else:
                    oSheet['J' + str(iRow)].font = activeFont

                oSheet['K' + str(iRow)] = (
                    oHead.inquiry_site_template if
                    str(oHead.inquiry_site_template).startswith('1') else
                    oHead.inquiry_site_template * -1 if
                    oHead.inquiry_site_template and
                    oHead.inquiry_site_template < -1 and
                    str(oHead.inquiry_site_template).startswith('-1') else ''
                ) if (oHead.sold_to_party == 626136 or 'KGP' in
                      str(oHead.customer_name).upper()) else ''
                oSheet['K' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['K' + str(iRow)].font = ipFont
                else:
                    oSheet['K' + str(iRow)].font = activeFont

                oSheet['L' + str(iRow)] = oHead.inquiry_site_template if \
                    str(oHead.inquiry_site_template).startswith('4') else \
                    oHead.inquiry_site_template * -1 if \
                    oHead.inquiry_site_template and \
                    oHead.inquiry_site_template < -1 and \
                    str(oHead.inquiry_site_template).startswith('-4') else ''
                oSheet['L' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['L' + str(iRow)].font = ipFont
                else:
                    oSheet['L' + str(iRow)].font = activeFont

                oSheet['M' + str(iRow)] = getattr(
                    oHead.model_replaced_link, 'configuration_designation', None
                ) or oHead.model_replaced or 'N/A'
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['M' + str(iRow)].font = ipFont
                else:
                    oSheet['M' + str(iRow)].font = activeFont

                oSheet['N' + str(iRow)] = getattr(
                    oHead.replaced_by_model.exclude(
                        bom_request_type__name='Discontinue').first(),
                    'configuration_designation', 'N/A')
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['N' + str(iRow)].font = ipFont
                else:
                    oSheet['N' + str(iRow)].font = activeFont

                oSheet['O' + str(iRow)] = oHead.change_comments
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['O' + str(iRow)].font = ipFont
                else:
                    oSheet['O' + str(iRow)].font = activeFont
            else:
                oSheet['C' + str(iRow)] = oHead.product_area2.name if \
                    oHead.product_area2 else ''
                oSheet['C' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['C' + str(iRow)].font = ipFont
                else:
                    oSheet['C' + str(iRow)].font = activeFont

                oSheet['D' + str(iRow)] = oHead.configuration_status.name.replace(
                    '/Pending', '').replace('In Process', 'IP')
                oSheet['D' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['D' + str(iRow)].font = ipFont
                else:
                    oSheet['D' + str(iRow)].font = activeFont

                oSheet['E' + str(iRow)] = getattr(
                    oHead.configuration.first_line, 'customer_number', '')
                oSheet['E' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['E' + str(iRow)].font = ipFont
                else:
                    oSheet['E' + str(iRow)].font = activeFont

                oSheet['F' + str(iRow)] = getattr(
                    oHead.configuration.first_line, 'sec_customer_number', '')
                oSheet['F' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['F' + str(iRow)].font = ipFont
                else:
                    oSheet['F' + str(iRow)].font = activeFont

                oSheet['G' + str(iRow)] = oHead.configuration_designation
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['G' + str(iRow)].font = ipFont
                else:
                    oSheet['G' + str(iRow)].font = activeFont

                oSheet['H' + str(iRow)] = oHead.inquiry_site_template if \
                    oHead.inquiry_site_template else oHead.configuration_designation
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['H' + str(iRow)].font = ipFont
                else:
                    oSheet['H' + str(iRow)].font = activeFont

                oSheet['I' + str(iRow)] = oHead.model_description
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['I' + str(iRow)].font = ipFont
                else:
                    oSheet['I' + str(iRow)].font = activeFont

                oSheet['J' + str(iRow)] = \
                    oHead.configuration.override_net_value or \
                    oHead.configuration.net_value
                oSheet['J' + str(iRow)].number_format = \
                    '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['J' + str(iRow)].font = ipFont
                else:
                    oSheet['J' + str(iRow)].font = activeFont

                oSheet['K' + str(iRow)] = \
                    oHead.headertimetracker_set.first().created_on.strftime('%Y')
                oSheet['K' + str(iRow)].alignment = centerAlign
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['K' + str(iRow)].font = ipFont
                else:
                    oSheet['K' + str(iRow)].font = activeFont

                oSheet['L' + str(iRow)] = oHead.inquiry_site_template if \
                    oHead.inquiry_site_template else ''
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['L' + str(iRow)].font = ipFont
                else:
                    oSheet['L' + str(iRow)].font = activeFont

                oSheet['M' + str(iRow)] = getattr(
                    oHead.model_replaced_link, 'configuration_designation', None
                ) or oHead.model_replaced or 'N/A'
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['M' + str(iRow)].font = ipFont
                else:
                    oSheet['M' + str(iRow)].font = activeFont

                oSheet['N' + str(iRow)] = getattr(
                    oHead.replaced_by_model.exclude(
                        bom_request_type__name='Discontinue').first(),
                    'configuration_designation', 'N/A')
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['N' + str(iRow)].font = ipFont
                else:
                    oSheet['N' + str(iRow)].font = activeFont

                oSheet['O' + str(iRow)] = oHead.change_comments
                if 'In Process' in oHead.configuration_status.name:
                    oSheet['O' + str(iRow)].font = ipFont
                else:
                    oSheet['O' + str(iRow)].font = activeFont

            iRow += 1
        # end for

        iRow += 1
    # end for
    oSheet.sheet_view.zoomScale = 80

    # Save file data to HTTP response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
        sFileName)
    response.set_cookie('fileMark', sCookie, max_age=60)
    oFile.save(response)

    return response
# end def


def WriteBaselineToFile(oBaseline, sVersion, sCustomer):
    """
    Function for writing a single baseline revision to a file
    :param oBaseline: Baseline object being written
    :param sVersion: String indicating which revision of baseline object to
    download
    :param sCustomer: String containing name of REF_CUSTOMER object to limit
    records downloaded (if this does not match the Baseline customer, user will
    get an empty file)
    :return: OpenPyXL file object containing data (file stream)
    """

    aColumnTitles = ['Line #', 'Product Number', 'Product Description',
                     'Order Qty', 'UoM', 'HW/SW Ind','Net Price',
                     'Traceability Req. (Serialization)', 'Customer Asset?',
                     'Customer Asset Tagging Requirement', 'Customer number',
                     'Second Customer Number', 'Vendor Article Number',
                     'Comments', 'Additional Reference\n(if required)']

    # Open workbook object
    oFile = openpyxl.Workbook()

    oCentered = Alignment(horizontal='center')
    oRevFont = Font(name='Arial', size=10, bold=True)

    # Fill Revision tab
    oSheet = oFile.active
    oSheet.title = 'Revisions'
    oSheet.sheet_properties.tabColor = '0F243E'

    oSheet['A1'] = 'Revision'
    oSheet['A1'].font = oRevFont
    oSheet['A1'].alignment = oCentered
    oSheet['B1'] = ''
    oSheet['C1'] = 'Date'
    oSheet['C1'].font = oRevFont
    oSheet['C1'].alignment = oCentered

    oSheet.column_dimensions['A'].width = 90
    oSheet.column_dimensions['B'].width = 17
    oSheet.column_dimensions['C'].width = 17

    # Ensure latest revision summary is up to date
    GenerateRevisionSummary(oBaseline, oBaseline.current_active_version,
                            oBaseline.current_inprocess_version)
    aRevisions = [oRev.revision for oRev in oBaseline.revisionhistory_set.all()]
    aRevisions = sorted(list(set(aRevisions)),
                        key=RevisionCompare)

    # Limit revision histories to only those up to, and including, the revision
    # being downloaded
    aRevisions = aRevisions[:aRevisions.index(sVersion) + 1]

    # Update revision history of revision being downloaded
    try:
        GenerateRevisionSummary(oBaseline, aRevisions[-2], sVersion)
    except:
        pass

    iRow = 2
    iColorIndex = 0
    aColors = ['C0C0C0', '9999FF', 'FFFFCC', 'CCFFFF', 'FF8080', 'CCCCFF',
               '00CCFF', 'CCFFCC', 'FFFF99', '99CCFF', 'FF99CC', 'CC99FF',
               'FFCC99', '33CCCC', '99CC00', 'FFCC00', 'FF9900', 'FF6600',
               '339966']

    # Write each revision history to the file
    for sRev in aRevisions:
        # Skip empty revisions
        if not oBaseline.baseline_revision_set.get(
                version=sRev).header_set.all():
            continue

        # Skip empty revision histories
        sHistory = oBaseline.revisionhistory_set.get(revision=sRev).history
        if not sHistory or not sHistory.strip():
            continue

        sRevHistory = 'Revision ' + sRev + ':\n' + \
                      oBaseline.revisionhistory_set.get(revision=sRev).history

        aLines = sRevHistory.split('\n')
        for i, sLine in enumerate(aLines):
            oSheet['A' + str(iRow)] = sLine
            if sLine in ('Added:', 'Discontinued:',
                         'Removed:', "Updated:") or i == 0:
                oSheet['A' + str(iRow)].font = Font(bold=True)
            oSheet['A' + str(iRow)].alignment = Alignment(wrap_text=True)
            oSheet['A' + str(iRow)].fill = GradientFill(
                type='linear',
                stop=[Color(aColors[iColorIndex]), Color(aColors[iColorIndex])])
            oSheet['B' + str(iRow)].fill = GradientFill(
                type='linear',
                stop=[Color(aColors[iColorIndex]), Color(aColors[iColorIndex])])
            oSheet['C' + str(iRow)].fill = GradientFill(
                type='linear',
                stop=[Color(aColors[iColorIndex]), Color(aColors[iColorIndex])])
            if i != len(aLines) - 1:
                iRow += 1

        if Baseline_Revision.objects.get(baseline=oBaseline,
                                         version=sRev).completed_date:
            oSheet['C' + str(iRow)] = Baseline_Revision.objects.get(
                baseline=oBaseline,
                version=sRev).completed_date.strftime('%m/%d/%Y')
            oSheet['C' + str(iRow)].alignment = oCentered

        iRow += 2
        iColorIndex += 1
        if iColorIndex >= len(aColors):
            iColorIndex = 0
    # end for

    # If downloading current in-process revision, only show in-process records
    # that are pending approval, and don't show records from the active revision
    # that are already discontinued or inactive.  This download is intended to
    # show how the baseline would appear if the in-process revision were
    # released
    if sVersion == oBaseline.current_inprocess_version:
        if oBaseline.latest_revision:
            aHeaders = oBaseline.latest_revision.header_set.exclude(
                configuration_status__name__in=(
                    'Discontinued', 'Inactive', 'On Hold')
            ).exclude(program__name__in=('DTS',) if oBaseline.title !=
                      'No Associated Baseline' else []
                      )
            # .order_by('configuration_status', 'pick_list',
            #           'configuration_designation')
        else:
            aHeaders = []

        aHeaders = list(chain(Baseline_Revision.objects.get(
            baseline=oBaseline,
            version=oBaseline.current_inprocess_version
        ).header_set.exclude(
            configuration_status__name__in=('On Hold', 'In Process')),
            aHeaders
        ))
    else:
        aHeaders = oBaseline.baseline_revision_set.get(
            version=sVersion).header_set.exclude(
            configuration_status__name__in=('On Hold',)).exclude(
            program__name__in=('DTS',) if oBaseline.title !=
                'No Associated Baseline' else [])
        aHeaders = list(aHeaders)

    # Sort Headers by baseline version and Product Area 2 values
    # aHeaders.sort(key=lambda inst: inst.baseline_version, reverse=True)
    aHeaders.sort(
        key=lambda inst: (
            int(inst.pick_list),
            str(inst.product_area2.name).upper() if inst.product_area2 else
            'ZZZZ',
            str(inst.configuration_designation),
            -(char_to_num(inst.baseline_version))
        )
    )
    # Added below sorting block for D-04368, D-04484: PA2 incorrect on baseline download.
    aHeaders.sort(
        key=lambda inst: (
            -int(inst.pick_list),
            str(inst.configuration_status)
        ), reverse=True
    )

    for oHeader in list(aHeaders):
        if sCustomer:
            # Skip records for the wrong customers
            if oHeader.customer_unit.name != sCustomer:
                continue

        # Remove Headers that have already been written (this occurs when
        # writing the in-process version of a record that will be replaced)
        if str(oHeader.configuration_designation) in oFile.sheetnames:
            aHeaders.remove(oHeader)
            continue

        # Records are separated by Product Area 2 value, so each new Product
        # area 2 value encountered needs to be written as a new separation tab
        # S-08413: Adjust category tab name based on Product Area 2 name for picklists:added not oHeader.pick_list
        if not oHeader.pick_list and oHeader.product_area2 and oHeader.product_area2.name not in \
                oFile.sheetnames:
            oSheet = oFile.create_sheet(title=re.sub(r'[\*\\\[\]:\'\?/]', '_',
                                                     oHeader.product_area2.name)
                                        )
            oSheet.sheet_properties.tabColor = '0062FF'
        # S-08413: Adjust category tab name based on Product Area 2 name for picklists:added not oHeader.pick_list
        elif not oHeader.pick_list and not oHeader.product_area2 and 'None' not in \
                oFile.sheetnames:
            oSheet = oFile.create_sheet(title="None")
            oSheet.sheet_properties.tabColor = '0062FF'
        # S-08413: Adjust category tab name based on Product Area 2 name for picklists: Commented line 1114 to 1117 and
        #  Added below lines from 1118 to 1132
        # elif oHeader.pick_list and 'Pick Lists' not in oFile.sheetnames \
        #         and 'Optional Hardware' not in oFile.sheetnames:
        #     oSheet = oFile.create_sheet(title="Pick Lists")
        #     oSheet.sheet_properties.tabColor = '0062FF'
        elif oHeader.pick_list:
            if not oHeader.product_area2 and 'None (Opt HW)' not in oFile.sheetnames:
                oSheet = oFile.create_sheet(title="None (Opt HW)")
                oSheet.sheet_properties.tabColor = '0062FF'
            elif oHeader.product_area2 and oHeader.product_area2.name == 'Optional Hardware' \
                    and 'Optional Hardware' not in oFile.sheetnames:
                oSheet = oFile.create_sheet(title="Optional Hardware")
                oSheet.sheet_properties.tabColor = '0062FF'
            elif oHeader.product_area2 and oHeader.product_area2.name != 'Optional Hardware' \
                    and (oHeader.product_area2.name + '(Opt HW)') not in oFile.sheetnames: # D-04626: Picklist tab names for Opt HW are generating duplicate tabs added 2nd and part
                oSheet = oFile.create_sheet(title=re.sub(r'[\*\\\[\]:\'\?/]', '_', # D-04403: Unable to download Baseline file
                                                     oHeader.product_area2.name)+'(Opt HW)')
                oSheet.sheet_properties.tabColor = '0062FF'

        iCurrentRow = 2

    # D-03942:Column width error on baseline download:- Changed the 2nd,8th,10th,13th width value in the array aColumnWidths below from (23,22,18,22) to (30,30,40,30) to increase the size of column(B,H,J,M)
    # Product Number(B),Traceability Req. (Serialization)(H),Customer Asset Tagging Requirement(J),VAN (M) in the downloaded excel
    # Also for UoM(E),USCC Article Code FAA(K) ,USCC Article code ZENG(L) from (6,18,18) to (10,25,25) for MTW

        if(oHeader.customer_unit_id==9):
            aColumnWidths = [9, 30, 58, 12, 10, 15, 18, 30, 30, 25, 25, 25, 30, 22, 23,
                             67, 83]
            aDynamicWidths = [0]*17
        else:
            aColumnWidths = [9, 30, 58, 12, 10, 15, 18, 30, 18, 40, 16, 22, 30, 23,
                             67, 83]
            aDynamicWidths = [0] * 16

        # Write record to sheet
        sTitle = str(oHeader.configuration_designation)
        if len(sTitle) > 31:
            sTitle = TitleShorten(sTitle)
            if len(sTitle) > 31:
                sTitle = "(Record title too long)"

        # Collect revision history data, which will be used to highlight changes
        if oHeader.bom_request_type.name == 'Update':
            try:
                oPrev = oHeader.model_replaced_link or Header.objects.get(
                    configuration_designation=oHeader.configuration_designation,
                    program=oHeader.program,
                    baseline_version=aRevisions[-2] if len(aRevisions) > 1 else
                    None
                )
                aHistory = HeaderComparison(oHeader, oPrev).split('\n')
            except (Header.DoesNotExist, ValueError):
                aHistory = []
            # end try
        else:
            aHistory = []

        dHistory = {}
        for sLine in aHistory:
            if not sLine:
                continue
            key, value = re.match(
                '\W*(\d+(?:\.\d+){0,2}) - (.+)', sLine).groups()
            if key not in dHistory:
                dHistory[key] = []
            dHistory[key].append(value)

        oSheet = oFile.create_sheet(
            title=re.sub(r'[\*\\\[\]:\'\?/]', '_', sTitle))

        if 'In Process' in oHeader.configuration_status.name and \
                oHeader.bom_request_type.name in ('New', 'Update'):
            oSheet.sheet_properties.tabColor = '80FF00'
        elif oHeader.bom_request_type.name == "Discontinue" or \
                oHeader.configuration_status.name == 'Discontinued' or (
                    hasattr(oHeader, 'replaced_by_model') and
                    oHeader.replaced_by_model.first() in aHeaders):
            oSheet.sheet_properties.tabColor = 'FF0000'

        # Build Header row
        if(oHeader.customer_unit_id == 9):
#  S-05743: Renaming Baseline columns for USCC Customer added if clause
# S-08087: Change to Customer Asset? and Customer Asset tagging rename "Customer Asset?":- Removed Customer Asset Tagging Requirement column & renamed
# Customer Asset column to "in the USCC BOM"
            aColumnTitles = ['Line #', 'Product Number', 'Product Description',
                             'Order Qty', 'UoM', 'HW/SW Ind', 'Unit Price', 'Net Price',
                             'Traceability Req. (Serialization)',
                             'In the USCC BOM', 'USCC Article Code FAA',
                         	 'USCC Article code ZENG', 'Vendor Article Number',
                             'Comments', 'Additional Reference\n(if required)']
            for iIndex in range(len(aColumnTitles)):
                oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'] = \
                    aColumnTitles[iIndex]
                oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].font = Font(
                    bold=True)
                oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].border = \
                    Border(
                        left=Side(color=colors.BLACK,
                                  border_style=borders.BORDER_THIN),
                        right=Side(color=colors.BLACK,
                                   border_style=borders.BORDER_THIN),
                        top=Side(color=colors.BLACK,
                                 border_style=borders.BORDER_THIN),
                        bottom=Side(color=colors.BLACK,
                                    border_style=borders.BORDER_MEDIUM)
                    )
                oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].alignment = \
                    Alignment(horizontal='center', vertical='center', wrapText=True)

                # oSheet.column_dimensions[
                #     str(utils.get_column_letter(iIndex + 1))
                # ].width = aColumnWidths[iIndex]

    #S-08088: Highlight Unit price Header:- Added column no. 6 with 7 to show the unit price column highlighted
                if iIndex in (6,7):                  #S-05744: change from 6 to 7, since index of Net Price is now 7
                    oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].fill = \
                        GradientFill(
                            type='linear',
                            stop=[Color(colors.GREEN), Color(colors.GREEN)])
                else:
                    oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].fill = \
                        GradientFill(
                            type='linear',
                            stop=[Color('fcd5b4'), Color('fcd5b4')])
            # end for

            oSheet.row_dimensions[1].height = 72.5
            oSheet.sheet_view.showGridLines = False
            oSheet.sheet_view.zoomScale = 60

            oBorder = Border(
                left=Side(color=colors.BLACK, border_style=borders.BORDER_THIN),
                right=Side(color=colors.BLACK, border_style=borders.BORDER_THIN),
                top=Side(color=colors.BLACK, border_style=borders.BORDER_THIN),
                bottom=Side(color=colors.BLACK, border_style=borders.BORDER_THIN)
            )

            oFirstRowColor = GradientFill(
                type='linear', stop=[Color('77b6e3'), Color('77b6e3')])
            oOffRowColor = GradientFill(
                type='linear', stop=[Color('e6e6e6'), Color('e6e6e6')])

            # Add line items (ordered by line number)
            aLineItems = oHeader.configuration.configline_set.all().order_by(
                'line_number')
            aLineItems = sorted(
                aLineItems,
                key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')]
            )
            oFirstItem = aLineItems[0] if len(aLineItems) > 0 else None
            for oLineItem in aLineItems:
                oSheet['A' + str(iCurrentRow)] = oLineItem.line_number
                oSheet['A' + str(iCurrentRow)].alignment = oCentered
                oSheet['A' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['A' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['A' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['A' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['B' + str(iCurrentRow)] = \
                    ('..' if oLineItem.is_grandchild else '.' if
                     oLineItem.is_child else '') + \
                    oLineItem.part.base.product_number
                oSheet['B' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['B' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['B' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['B' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['C' + str(iCurrentRow)] = oLineItem.part.product_description
                oSheet['C' + str(iCurrentRow)].border = oBorder
    # S-05816:Fix column within downloaded baseline - Added for wrapping text in product_description column in the downloaded file
                oSheet['C' + str(iCurrentRow)].alignment = Alignment(wrap_text=True)
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['C' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['C' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['C' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['D' + str(iCurrentRow)] = oLineItem.order_qty
                oSheet['D' + str(iCurrentRow)].alignment = oCentered
                oSheet['D' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['D' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['D' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['D' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['E' + str(iCurrentRow)] = oLineItem.part.base.unit_of_measure
                oSheet['E' + str(iCurrentRow)].alignment = oCentered
                oSheet['E' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['E' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['E' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['E' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['F' + str(iCurrentRow)] = oLineItem.commodity_type
                oSheet['F' + str(iCurrentRow)].alignment = oCentered
                oSheet['F' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['F' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['F' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['F' + str(iCurrentRow)].fill = oOffRowColor
                #
                if not oHeader.pick_list:
                    if  oLineItem == oFirstItem :
                        oSheet['G' + str(iCurrentRow)] = ''
                    else:
                        if GrabValue(oLineItem, 'linepricing.override_price'):
                            oSheet['G' + str(iCurrentRow)] = GrabValue(oLineItem, 'linepricing.override_price')
                        elif GrabValue(oLineItem,'linepricing.pricing_object.unit_price'):
                            oSheet['G' + str(iCurrentRow)] = GrabValue(oLineItem, 'linepricing.pricing_object.unit_price')
                        else:
                            oSheet['G' + str(iCurrentRow)] = ''
                            # end if
                            # end if
                            # end if
                else:
                    if GrabValue(oLineItem, 'linepricing.override_price'):
                        oSheet['G' + str(iCurrentRow)] = GrabValue(oLineItem, 'linepricing.override_price')
                    elif GrabValue(oLineItem, 'linepricing.pricing_object.unit_price'):
                        oSheet['G' + str(iCurrentRow)] = GrabValue(oLineItem, 'linepricing.pricing_object.unit_price')
                    else:
                        oSheet['G' + str(iCurrentRow)] = ''

                # oSheet['G' + str(iCurrentRow)] = ''
                oSheet['G' + str(iCurrentRow)].number_format = \
                    '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                oSheet['G' + str(iCurrentRow)].border = oBorder

                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['G' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['G' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['G' + str(iCurrentRow)].fill = oOffRowColor



                if oHeader.pick_list:
                    if GrabValue(oLineItem,
                                 'linepricing.override_price') is not None:
                        oSheet['H' + str(iCurrentRow)] = \
                            GrabValue(oLineItem, 'linepricing.override_price')
                    else:
                        oSheet['H' + str(iCurrentRow)] = \
                            GrabValue(oLineItem,
                                      'linepricing.pricing_object.unit_price')
                else:
                    if oLineItem == oFirstItem:
                        if GrabValue(oLineItem,
                                     'linepricing.override_price') is not None:
                            oSheet['H' + str(iCurrentRow)] = \
                                GrabValue(oLineItem, 'linepricing.override_price')
                        else:
                            oSheet['H' + str(iCurrentRow)] = \
                                GrabValue(oLineItem,
                                          'linepricing.pricing_object.unit_price')
                    else:
                        oSheet['H' + str(iCurrentRow)] = ''

                oSheet['H' + str(iCurrentRow)].number_format = \
                    '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                oSheet['H' + str(iCurrentRow)].border = oBorder

                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['H' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['H' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['H' + str(iCurrentRow)].fill = oOffRowColor


                oSheet['I' + str(iCurrentRow)] = oLineItem.traceability_req
                oSheet['I' + str(iCurrentRow)].alignment = oCentered
                oSheet['I' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['I' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['I' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['I' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['J' + str(iCurrentRow)] = oLineItem.customer_asset
                oSheet['J' + str(iCurrentRow)].alignment = oCentered
                oSheet['J' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['J' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['J' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['J' + str(iCurrentRow)].fill = oOffRowColor

# S-08087: Change to Customer Asset? and Customer Asset tagging rename "Customer Asset?":- Removed Customer Asset Tagging Requirement column(K)
# & rearranged the Excel sheet columns accordingly
                oSheet['K' + str(iCurrentRow)] = oLineItem.customer_number
                oSheet['K' + str(iCurrentRow)].alignment = oCentered
                oSheet['K' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['K' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['K' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['K' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['L' + str(iCurrentRow)] = oLineItem.sec_customer_number
                oSheet['L' + str(iCurrentRow)].alignment = oCentered
                oSheet['L' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['L' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['L' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['L' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['M' + str(iCurrentRow)] = oLineItem.vendor_article_number
                oSheet['M' + str(iCurrentRow)].border = oBorder
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['M' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['M' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['M' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['N' + str(iCurrentRow)] = oLineItem.comments
                oSheet['N' + str(iCurrentRow)].border = oBorder
        # S-05816:Fix column within downloaded baseline - Added for wrapping text in comments column in the downloaded file
                oSheet['N' + str(iCurrentRow)].alignment = Alignment(wrap_text=True)
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['N' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['N' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['N' + str(iCurrentRow)].fill = oOffRowColor

                oSheet['O' + str(iCurrentRow)] = oLineItem.additional_ref
                oSheet['O' + str(iCurrentRow)].border = oBorder
        # S-05816:Fix column within downloaded baseline - Added for wrapping text in Additional reference column in the downloaded file
                oSheet['O' + str(iCurrentRow)].alignment = Alignment(wrap_text=True)
                if oLineItem == oFirstItem and not oHeader.pick_list:
                    oSheet['O' + str(iCurrentRow)].fill = oFirstRowColor
                    oSheet['O' + str(iCurrentRow)].font = Font(bold=True)
                elif iCurrentRow % 2 == 0:
                    oSheet['O' + str(iCurrentRow)].fill = oOffRowColor


                # Guesstimate the width of the data entered, so that we can try to
                # make the cell wide enough to show the entire value
                for idx in range(17):
                    if len(str(oSheet[utils.get_column_letter(idx+1) +
                                      str(iCurrentRow)
                                      ].value)) > aDynamicWidths[idx]:
                        aDynamicWidths[idx] = \
                            len(str(oSheet[utils.get_column_letter(idx+1) +
                                           str(iCurrentRow)].value))

                # Determine and mark each changed value for the row
                #S-07841 Highlighting for fields in baseline download(added for description, add_ref, comments)
                aChangeCols = []
                if oLineItem.line_number in dHistory:
                    for sChange in dHistory[oLineItem.line_number]:
                        if 'description' in sChange:
                            aChangeCols.append(3)

                        if 'quantity' in sChange:
                            aChangeCols.append(4)

                        if 'line price' in sChange:
                            aChangeCols.append(8)

                        if 'comments' in sChange:
                            aChangeCols.append(14)

                        if 'Additional Reference' in sChange:
                            aChangeCols.append(15)

                        if 'added' in sChange:
                            aChangeCols.extend(list(range(1, 16)))

                        if 'replaced' in sChange:
                            aChangeCols.extend(list(range(2, 16)))

                    aChangeCols = list(set(aChangeCols))

                for iCol in aChangeCols:
                    oSheet[
                        str(utils.get_column_letter(iCol)) + str(iCurrentRow)
                    ].fill = GradientFill(
                        type='linear', stop=[Color('e0fa2e'), Color('e0fa2e')])

                iCurrentRow += 1
            # end for
            if not oHeader.pick_list:
                oSheet['H2'] = oHeader.configuration.override_net_value or \
                               oHeader.configuration.net_value or ''
        else:
              for iIndex in range(len(aColumnTitles)):
                  oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'] = \
                      aColumnTitles[iIndex]
                  oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].font = Font(
                      bold=True)
                  oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].border = \
                      Border(
                          left=Side(color=colors.BLACK,
                                    border_style=borders.BORDER_THIN),
                          right=Side(color=colors.BLACK,
                                     border_style=borders.BORDER_THIN),
                          top=Side(color=colors.BLACK,
                                   border_style=borders.BORDER_THIN),
                          bottom=Side(color=colors.BLACK,
                                      border_style=borders.BORDER_MEDIUM)
                      )
                  oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].alignment = \
                      Alignment(horizontal='center', vertical='center', wrapText=True)

                  # oSheet.column_dimensions[
                  #     str(utils.get_column_letter(iIndex + 1))
                  # ].width = aColumnWidths[iIndex]

                  if iIndex in (6,):
                      oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].fill = \
                          GradientFill(
                              type='linear',
                              stop=[Color(colors.GREEN), Color(colors.GREEN)])
                  else:
                      oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'].fill = \
                          GradientFill(
                              type='linear',
                              stop=[Color('fcd5b4'), Color('fcd5b4')])
              # end for

              oSheet.row_dimensions[1].height = 72.5
              oSheet.sheet_view.showGridLines = False
              oSheet.sheet_view.zoomScale = 60

              oBorder = Border(
                  left=Side(color=colors.BLACK, border_style=borders.BORDER_THIN),
                  right=Side(color=colors.BLACK, border_style=borders.BORDER_THIN),
                  top=Side(color=colors.BLACK, border_style=borders.BORDER_THIN),
                  bottom=Side(color=colors.BLACK, border_style=borders.BORDER_THIN)
              )

              oFirstRowColor = GradientFill(
                  type='linear', stop=[Color('77b6e3'), Color('77b6e3')])
              oOffRowColor = GradientFill(
                  type='linear', stop=[Color('e6e6e6'), Color('e6e6e6')])

              # Add line items (ordered by line number)
              aLineItems = oHeader.configuration.configline_set.all().order_by(
                  'line_number')
              aLineItems = sorted(
                  aLineItems,
                  key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')]
              )
              oFirstItem = aLineItems[0] if len(aLineItems) > 0 else None
              for oLineItem in aLineItems:
                  oSheet['A' + str(iCurrentRow)] = oLineItem.line_number
                  oSheet['A' + str(iCurrentRow)].alignment = oCentered
                  oSheet['A' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['A' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['A' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['A' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['B' + str(iCurrentRow)] = \
                      ('..' if oLineItem.is_grandchild else '.' if
                      oLineItem.is_child else '') + \
                      oLineItem.part.base.product_number
                  oSheet['B' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['B' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['B' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['B' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['C' + str(iCurrentRow)] = oLineItem.part.product_description
                  oSheet['C' + str(iCurrentRow)].border = oBorder
                  # S-05816:Fix column within downloaded baseline - Added for wrapping text in product_description column in the downloaded file
                  oSheet['C' + str(iCurrentRow)].alignment = Alignment(wrap_text=True)
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['C' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['C' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['C' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['D' + str(iCurrentRow)] = oLineItem.order_qty
                  oSheet['D' + str(iCurrentRow)].alignment = oCentered
                  oSheet['D' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['D' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['D' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['D' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['E' + str(iCurrentRow)] = oLineItem.part.base.unit_of_measure
                  oSheet['E' + str(iCurrentRow)].alignment = oCentered
                  oSheet['E' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['E' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['E' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['E' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['F' + str(iCurrentRow)] = oLineItem.commodity_type
                  oSheet['F' + str(iCurrentRow)].alignment = oCentered
                  oSheet['F' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['F' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['F' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['F' + str(iCurrentRow)].fill = oOffRowColor

                  if oHeader.pick_list:
                      if GrabValue(oLineItem,
                                   'linepricing.override_price') is not None:
                          oSheet['G' + str(iCurrentRow)] = \
                              GrabValue(oLineItem, 'linepricing.override_price')
                      else:
                          oSheet['G' + str(iCurrentRow)] = \
                              GrabValue(oLineItem,
                                        'linepricing.pricing_object.unit_price')
                  else:
                      oSheet['G' + str(iCurrentRow)] = ''

                  oSheet['G' + str(iCurrentRow)].number_format = \
                      '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                  oSheet['G' + str(iCurrentRow)].border = oBorder

                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['G' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['G' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['G' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['H' + str(iCurrentRow)] = oLineItem.traceability_req
                  oSheet['H' + str(iCurrentRow)].alignment = oCentered
                  oSheet['H' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['H' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['H' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['H' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['I' + str(iCurrentRow)] = oLineItem.customer_asset
                  oSheet['I' + str(iCurrentRow)].alignment = oCentered
                  oSheet['I' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['I' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['I' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['I' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['J' + str(iCurrentRow)] = oLineItem.customer_asset_tagging
                  oSheet['J' + str(iCurrentRow)].alignment = oCentered
                  oSheet['J' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['J' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['J' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['J' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['K' + str(iCurrentRow)] = oLineItem.customer_number
                  oSheet['K' + str(iCurrentRow)].alignment = oCentered
                  oSheet['K' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['K' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['K' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['K' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['L' + str(iCurrentRow)] = oLineItem.sec_customer_number
                  oSheet['L' + str(iCurrentRow)].alignment = oCentered
                  oSheet['L' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['L' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['L' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['L' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['M' + str(iCurrentRow)] = oLineItem.vendor_article_number
                  oSheet['M' + str(iCurrentRow)].border = oBorder
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['M' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['M' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['M' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['N' + str(iCurrentRow)] = oLineItem.comments
                  oSheet['N' + str(iCurrentRow)].border = oBorder
                  # S-05816:Fix column within downloaded baseline - Added for wrapping text in comments column in the downloaded file
                  oSheet['N' + str(iCurrentRow)].alignment = Alignment(wrap_text=True)
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['N' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['N' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['N' + str(iCurrentRow)].fill = oOffRowColor

                  oSheet['O' + str(iCurrentRow)] = oLineItem.additional_ref
                  oSheet['O' + str(iCurrentRow)].border = oBorder
                  # S-05816:Fix column within downloaded baseline - Added for wrapping text in Additional reference column in the downloaded file
                  oSheet['O' + str(iCurrentRow)].alignment = Alignment(wrap_text=True)
                  if oLineItem == oFirstItem and not oHeader.pick_list:
                      oSheet['O' + str(iCurrentRow)].fill = oFirstRowColor
                      oSheet['O' + str(iCurrentRow)].font = Font(bold=True)
                  elif iCurrentRow % 2 == 0:
                      oSheet['O' + str(iCurrentRow)].fill = oOffRowColor

                  # Guesstimate the width of the data entered, so that we can try to
                  # make the cell wide enough to show the entire value
                  for idx in range(16):
                      if len(str(oSheet[utils.get_column_letter(idx + 1) +
                              str(iCurrentRow)
                                 ].value)) > aDynamicWidths[idx]:
                          aDynamicWidths[idx] = \
                              len(str(oSheet[utils.get_column_letter(idx + 1) +
                                             str(iCurrentRow)].value))

                  # Determine and mark each changed value for the row
                  # S-07841 Highlighting for fields in baseline download(added for description, add_ref, comments)
                  aChangeCols = []
                  if oLineItem.line_number in dHistory:
                      for sChange in dHistory[oLineItem.line_number]:
                          if 'description' in sChange:
                              aChangeCols.append(3)

                          if 'quantity' in sChange:
                              aChangeCols.append(4)

                          if 'line price' in sChange:
                              aChangeCols.append(7)

                          if 'comments' in sChange:
                              aChangeCols.append(14)

                          if 'Additional Reference' in sChange:
                              aChangeCols.append(15)

                          if 'added' in sChange:
                              aChangeCols.extend(list(range(1, 16)))

                          if 'replaced' in sChange:
                              aChangeCols.extend(list(range(2, 16)))

                      aChangeCols = list(set(aChangeCols))

                  for iCol in aChangeCols:
                      oSheet[
                              str(utils.get_column_letter(iCol)) + str(iCurrentRow)
                              ].fill = GradientFill(
                              type='linear', stop=[Color('e0fa2e'), Color('e0fa2e')])

                  iCurrentRow += 1
              # end for
              if not oHeader.pick_list:
                   oSheet['G2'] = oHeader.configuration.override_net_value or \
                                  oHeader.configuration.net_value or ''

        # Update cell widths
        for iIndex in range(len(aColumnTitles)):
            oSheet.column_dimensions[
                str(utils.get_column_letter(iIndex + 1))].width = \
                max(aColumnWidths[iIndex], aDynamicWidths[iIndex])
    # end for

    # Write Table of Contents tab.  This is last so that we only write ToC data
    # for Header objects still in the array after the previous section
    oSheet = oFile.create_sheet('ToC', 0)
    #  S-05743: Renaming Baseline columns for USCC Customer added if clause
    if oBaseline.customer_id == 9 or sCustomer == 'MTW':
        dTOCData = {
            3: ['Configuration', 'configuration_designation', 25],
            15: ['Customer Designation', 'customer_designation', 15],
            10: ['Technology', 'technology', 15],
            1: ['Product Area 1', 'product_area1', 25],
            2: ['Product Area 2', 'product_area2', 10],
            4: ['Model', 'model', 25],
            7: ['Model Description', 'model_description', 50],
            11: ['What Model is this replacing?', 'model_replaced', 25],
            9: ['BoM & Inquiry Details', None, 25],
            12: ['BoM Request Type', 'bom_request_type', 10],
            8: ['Configuration / Ordering Status', 'configuration_status', 15],
            13: ['Inquiry', 'inquiry_site_template', 10],
            14: ['Site Template', 'inquiry_site_template', 10],
            16: ['Ext Notes', 'external_notes', 50],
            # D- 03265 - Missing columns in downloaded baseline files -added below two columns
            5: ['USCC Article Code FAA', 'configuration.first_line.customer_number', 20],
            6: ['USCC Article code ZENG', 'configuration.first_line.sec_customer_number', 20]
    }
    else:
        dTOCData = {
            3: ['Configuration', 'configuration_designation', 25],
            15: ['Customer Designation', 'customer_designation', 15],
            10: ['Technology', 'technology', 15],
            1: ['Product Area 1', 'product_area1', 25],
            2: ['Product Area 2', 'product_area2', 10],
            4: ['Model', 'model', 25],
            7: ['Model Description', 'model_description', 50],
            11: ['What Model is this replacing?', 'model_replaced', 25],
            9: ['BoM & Inquiry Details', None, 25],
            12: ['BoM Request Type', 'bom_request_type', 10],
            8: ['Configuration / Ordering Status', 'configuration_status', 15],
            13: ['Inquiry', 'inquiry_site_template', 10],
            14: ['Site Template', 'inquiry_site_template', 10],
            16: ['Ext Notes', 'external_notes', 50],
            # D- 03265 - Missing columns in downloaded baseline files -added below two columns
            5: ['Customer Number', 'configuration.first_line.customer_number', 20],
            6: ['Second Customer Number', 'configuration.first_line.sec_customer_number', 20]
        }

    for iIndex in sorted(dTOCData.keys()):
        oSheet[utils.get_column_letter(iIndex) + '1'].value = \
            dTOCData[iIndex][0]
        oSheet[utils.get_column_letter(iIndex) + '1'].alignment = Alignment(
            wrap_text=True)
        oSheet[utils.get_column_letter(iIndex) + '1'].font = Font(bold=True)
        oSheet[utils.get_column_letter(iIndex) + '1'].fill = GradientFill(
            type='linear', stop=[Color('C0C0C0'), Color('C0C0C0')])
        oSheet[utils.get_column_letter(iIndex) + '1'].border = Border(
            left=Side(color=colors.BLACK, border_style=borders.BORDER_THIN),
            right=Side(color=colors.BLACK, border_style=borders.BORDER_THIN),
            top=Side(color=colors.BLACK, border_style=borders.BORDER_THIN),
            bottom=Side(color=colors.BLACK, border_style=borders.BORDER_THIN))
        oSheet.column_dimensions[str(utils.get_column_letter(iIndex))].width = \
            dTOCData[iIndex][2]

    iCurrentRow = 2

    # Write ToC data for each Header object
    # Added below sorting block for D-04368, D-04484: PA2 incorrect on baseline download.
    aHeaders.sort(
        key=lambda inst: (
            int(inst.pick_list),
            str(inst.product_area2.name).upper() if inst.product_area2 else
            'ZZZZ')
    )
    for oHeader in aHeaders:
        if sCustomer:
            if oHeader.customer_unit.name != sCustomer:
                continue

        for iIndex in sorted(dTOCData.keys()):
            if dTOCData[iIndex][1] and dTOCData[iIndex][1] != \
                    'inquiry_site_template':
                if 'Customer Number' in dTOCData[iIndex][0] and \
                        oHeader.pick_list:
                    oSheet[utils.get_column_letter(iIndex) +
                           str(iCurrentRow)].value = 'Multiple'
                else:
                    oSheet[utils.get_column_letter(iIndex) +
                           str(iCurrentRow)].value = str(
                        GrabValue(oHeader, dTOCData[iIndex][1], '')
                    ).replace('/Pending', '')
            elif dTOCData[iIndex][1] == 'inquiry_site_template':
                if dTOCData[iIndex][0] == 'Inquiry':
                    if str(GrabValue(oHeader, dTOCData[iIndex][1], None)
                           ).startswith('1'):
                        oSheet[utils.get_column_letter(iIndex) +
                               str(iCurrentRow)].value = GrabValue(
                            oHeader, dTOCData[iIndex][1], '')
                    elif GrabValue(oHeader, dTOCData[iIndex][1], 0) < -1 and \
                            str(GrabValue(
                                oHeader, dTOCData[iIndex][1], 0) * -1
                                ).startswith('1'):
                        oSheet[utils.get_column_letter(iIndex) +
                               str(iCurrentRow)].value = GrabValue(
                            oHeader, dTOCData[iIndex][1], 0) * -1
                elif dTOCData[iIndex][0] == 'Site Template':
                    if str(GrabValue(
                            oHeader, dTOCData[iIndex][1], None
                    )).startswith('4'):
                        oSheet[utils.get_column_letter(iIndex) +
                               str(iCurrentRow)].value = GrabValue(
                            oHeader, dTOCData[iIndex][1], '')
                    elif GrabValue(oHeader, dTOCData[iIndex][1], 0) < -1 and \
                            str(GrabValue(oHeader, dTOCData[iIndex][1], 0) * -1
                                ).startswith('4'):
                        oSheet[utils.get_column_letter(iIndex) +
                               str(iCurrentRow)].value = GrabValue(
                            oHeader, dTOCData[iIndex][1], 0) * -1
                else:
                    oSheet[utils.get_column_letter(iIndex) +
                           str(iCurrentRow)].value = None
            elif not dTOCData[iIndex][1]:
                oSheet[utils.get_column_letter(iIndex) +
                       str(iCurrentRow)].value = str(
                    oHeader.configuration_designation)
                oSheet[utils.get_column_letter(iIndex) +
                       str(iCurrentRow)].hyperlink = "#'" + (
                    TitleShorten(str(oHeader.configuration_designation)) if
                    len(str(oHeader.configuration_designation)) > 31 else
                    str(oHeader.configuration_designation)) + "'!A1"
                oSheet[utils.get_column_letter(iIndex) +
                       str(iCurrentRow)].font = Font(color=colors.BLUE,
                                                     underline='single')
        iCurrentRow += 1

    return oFile
# end def


def EmailDownload(oRequest,oBaseline):
    """
    Function that emails the latest active version of a baseline to all users
    in a DistroList object or all users in the BOM_PSM_Baseline_Manager
    user group
    :param oBaseline: Baseline object to send in email
    :return: None
    """

    # Determine latest active version and filename
    sVersion = oBaseline.current_active_version
    sFileName = str(Baseline_Revision.objects.get(
        baseline=oBaseline, version=sVersion)) + ".xlsx"

    # Retrieve DistroList object, if it exists
    # commented out line 1377-1379 for fix D-03265- Review & Approval & Baseline Release mail not sent
    # try:
    #     oDistroList = DistroList.objects.get(customer_unit=oBaseline.customer)
    # except DistroList.DoesNotExist:
    oDistroList = None
    # D-03452: Some emails are not being tagged as test system, added below lines
    if oRequest.POST.get('windowurl') == 'local':
        envName = 'Local System:'
    elif oRequest.POST.get('windowurl') == 'test':
        envName = 'Test System:'
    else:
        envName = ''


    # Build email message
    sSubject = envName + 'New revision released: ' + oBaseline.title  # D-03452: Some emails are not being tagged as test system, added envName
    sMessage = ('Revision {} of {} has been released as of {}.  A copy of the '
                'baseline has been attached.\nIssues may be addressed with '
                'Katya Pridgen at Katya.Pridgen@Ericsson.com.\n\n'
                '***This is an automated message. Do not reply to this '
                'message.***').format(
        oBaseline.current_active_version, oBaseline.title,
        oBaseline.latest_revision.completed_date.strftime('%m/%d/%Y'))

    sMessageHtml = ('Revision {} of {} has been released as of {}.  '
                    'A copy of the baseline has been attached.<br/>'
                    'Issues may be addressed with Katya Pridgen at '
                    '<a href="mailto:Katya.Pridgen@Ericsson.com">'
                    'Katya.Pridgen@ericsson.com</a>.<br/><br/>'
                    '<div style="color: red">***This is an automated message. '
                    'Do not reply to this message.***</div>').format(
        oBaseline.current_active_version, oBaseline.title,
        oBaseline.latest_revision.completed_date.strftime('%m/%d/%Y'))
    # S-10576: Change the header of the tool to ACC :- Changed the tool name from pcbm to acc
    oNewMessage = EmailMultiAlternatives(
        sSubject,
        sMessage,
        'acc.admin@ericsson.com',
        [obj.email for obj in oDistroList.users_included.all()] if oDistroList
        else [user.email for user in User.objects.filter(
            groups__name="BOM_PSM_Baseline_Manager")],
        cc=oDistroList.additional_addresses.split() if oDistroList else None
    )
    oNewMessage.attach_alternative(sMessageHtml, 'text/html')

    # Write baseline to data stream and attach data stream to email
    oStream = BytesIO()
    oFile = WriteBaselineToFile(
        oBaseline,
        sVersion,
        oBaseline.customer.name if oBaseline.customer else None
    )
    oFile.save(oStream)

    oNewMessage.attach(sFileName, oStream.getvalue(), 'application/ms-excel')
    oNewMessage.send()
# end def


def ConfigPriceDownload(oRequest):
    """
    View to download pricing data for a specific Header/Configuration
    :param oRequest: Django HTTP request object
    :return: HttpResponse containing downloaded data file
    """

    if oRequest.POST:
        # Retrieve desired Header object and determine filename
        oHeader = Header.objects.get(
            configuration_designation=oRequest.POST['config'],
            baseline=Baseline_Revision.objects.get(
                id=oRequest.POST['baseline']) or None)
        # S-05923: Pricing - Restrict View commented out below line to download configs which don't have program.
            # program = REF_PROGRAM.objects.get(
            # id=oRequest.POST['program']) or None

        sFileName = oHeader.configuration_designation + (
            '_' + oHeader.program.name if oHeader.program else ''
        ) + ' Pricing.xlsx'

        # Write Header pricing data to file stream
        oFile = WriteConfigPriceToFile(oHeader)

        # Save file stream to HTTP response
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
            sFileName)
        response.set_cookie('fileMark', oRequest.POST['file_cookie'],
                            max_age=60)
        oFile.save(response)

        return response
    else:
        return HttpResponse()
# end def


def WriteConfigPriceToFile(oHeader):
    """
    Function to write Header pricing data to file stream
    :param oHeader: Header object to use to create file
    :return: OpenPyXL file object
    """
    # Retrieve ConfigLines and sort by line number
    aLine = oHeader.configuration.configline_set.all()
    aLine = sorted(aLine,
                   key=lambda x: ([int(y) for y in x.line_number.split('.')]))

    # Build data rows for .xls file
    aConfigLines = [
        {
            '0': oLine.line_number,
            '1': ('..' if oLine.is_grandchild else '.' if
                  oLine.is_child else '') + oLine.part.base.product_number,
            '2': str(oLine.part.base.product_number) +
            str('_' + oLine.spud.name if oLine.spud else ''),
            '3': oLine.part.product_description,
            '4': float(oLine.order_qty if oLine.order_qty else 0.0),
            '5': float(GrabValue(oLine,
                                 'linepricing.pricing_object.unit_price', 0.0)),
            '6': float(oLine.order_qty or 0) * float(GrabValue(
                oLine, 'linepricing.pricing_object.unit_price', 0.0)),
            '7': GrabValue(oLine, 'linepricing.override_price', ''),
            '8': oLine.higher_level_item or '',
            '9': oLine.material_group_5 or '',
            '10': oLine.commodity_type or '',
            '11': oLine.comments or '',
            '12': oLine.additional_ref or ''
        } for oLine in aLine
        ]

    if not oHeader.pick_list:
        config_total = sum([float(line['6']) for line in aConfigLines])
        aConfigLines[0]['5'] = aConfigLines[0]['6'] = config_total

    headers = ['Line #', 'Product Number', 'Internal Product Number',
               'Product Description', 'Order Qty', 'Unit Price', 'Total Price',
               'Manual Override for Total NET Price', 'Linkage',
               'Material Group 5', 'HW/SW Indicator',
               'Comments (viewable by customer)',
               'Additional Reference (viewable by customer)']

    # Open new workbook
    oFile = openpyxl.Workbook()
    oSheet = oFile.active

    # Write column titles
    for i in range(len(headers)):
        oSheet[utils.get_column_letter(i + 1) + '1'] = headers[i]

    # Write data to table
    for i in range(len(aConfigLines)):
        for j in range(len(headers)):
            if j in (5, 6, 7):
                oSheet[
                    utils.get_column_letter(j + 1) + str(2 + i)
                ].number_format = '_($* #,##0.00_);_($* (#,##0.00);_(@_)'
            oSheet[utils.get_column_letter(j + 1) + str(2 + i)] = \
                aConfigLines[i][str(j)]

    return oFile


def PriceOverviewDownload(oRequest):
    """
    View to download all part number unit pricing information
    :param oRequest: Django HTTP request object
    :return: HTTPResponse containing data file download
    """
    if oRequest.POST:
        sFileName = 'Pricing Overview.xlsx'

        # Write pricing data to data stream
        oFile = WritePriceOverviewToFile(*PricingOverviewLists(oRequest)) # S-05923: Pricing - Restrict View added oRequest

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
            sFileName)
        response.set_cookie('fileMark', oRequest.POST['file_cookie'],
                            max_age=60)
        oFile.save(response)

        return response
    else:
        return HttpResponse()


def WritePriceOverviewToFile(aPricingLines, aComments):
    """
    Function to write Part unit price information
    :param aPricingLines: List of lists containing each row of information
    :param aComments: List of lists containing comment info for each row
    :return: OpenPyXL file object
    """
    # Create column headers
    aHeaders = ['Part Number', 'Customer', 'Sold-To', 'SPUD', 'Technology',
                'Latest Unit Price ($)']
    for j in range(1, 5):
        aHeaders.append(str(datetime.datetime.now().year - j) + ' Price ($)')

    # Create new workbook
    oFile = openpyxl.Workbook()
    oSheet = oFile.active

    # Write column headers
    for i in range(len(aHeaders)):
        oSheet[utils.get_column_letter(i + 1) + '1'] = aHeaders[i]

    # Write each cell of data and add a cell comment containing any comment data
    # provided
    for i in range(len(aPricingLines)):
        for j in range(len(aHeaders)):
            oSheet[utils.get_column_letter(j + 1) + str(2 + i)] = \
                aPricingLines[i][j]
            if j >= 5 and aComments[i][j - 5]:
                oSheet[utils.get_column_letter(j + 1) + str(2 + i)].comment = \
                    Comment(aComments[i][j - 5], 'System')

    return oFile


def PartPriceDownload(oRequest):
    """
    View to download pricing information for a specific part
    :param oRequest: Django HTTP request object
    :return: HTTPResponse containing data file download
    """
    if oRequest.POST:
        aPriceList = PricingObject.objects.filter(
            part__product_number=oRequest.POST['initial'],
            is_current_active=True
        ).order_by('customer', 'sold_to', 'spud')

        sFileName = oRequest.POST['initial'] + ' Pricing.xlsx'

        aPriceList = [[
                          oPriceObj.part.product_number,
                          oPriceObj.customer.name,
                          oPriceObj.sold_to or '(None)',
                          getattr(oPriceObj.spud, 'name', '(None)'),
                          getattr(oPriceObj.technology, 'name', '(None)'),
                          oPriceObj.unit_price or '',
                          oPriceObj.valid_to_date.strftime('%m/%d/%Y') if
                          oPriceObj.valid_to_date else '',
                          oPriceObj.valid_from_date.strftime('%m/%d/%Y') if
                          oPriceObj.valid_from_date else '',
                          oPriceObj.cutover_date.strftime('%m/%d/%Y') if
                          oPriceObj.cutover_date else '',
                          str(oPriceObj.price_erosion),
                          oPriceObj.erosion_rate or '',
                          oPriceObj.comments or '',
                       ] for oPriceObj in aPriceList]

        oFile = WritePartPriceToFile(aPriceList)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
            sFileName)
        response.set_cookie('fileMark', oRequest.POST['file_cookie'],
                            max_age=60)
        oFile.save(response)

        return response
    else:
        return HttpResponse()
# end def


def WritePartPriceToFile(aPriceList):
    """
    Function to write individual part pricing information to xls file
    :param aPriceList: List of list containing pricing row information
    :return: OpenPyXL file object
    """
    aHeaders = ['Part Number', 'Customer', 'Sold-To', 'SPUD', 'Technology',
                'Latest Unit Price ($)', 'Valid To', 'Valid From',
                'Cut-over Date', 'Price Erosion', 'Erosion Rate (%)', 'Comments'
                ]
    oFile = openpyxl.Workbook()
    oSheet = oFile.active

    for i in range(len(aHeaders)):
        oSheet[utils.get_column_letter(i + 1) + '1'] = aHeaders[i]

    for i in range(len(aPriceList)):
        for j in range(len(aHeaders)):
            oSheet[utils.get_column_letter(j + 1) + str(2 + i)] = \
                aPriceList[i][j]

    return oFile
# end def


def ErosionDownload(oRequest):
    """
    View to download list of erodible part prices to file
    :param oRequest: Django HTTP request object
    :return: HTTPResponse containing data file download
    """
    if oRequest.POST:
        dArgs = {'price_erosion': True,
                 'is_current_active': True}

        # Extract filtering parameters from request
        if oRequest.POST['cu'] != 'all':
            dArgs['customer__name'] = oRequest.POST['cu']

        if oRequest.POST['soldto'] != 'all':
            if oRequest.POST['soldto'] != 'None':
                dArgs['sold_to'] = oRequest.POST['soldto']
            else:
                dArgs['sold_to'] = None

        if oRequest.POST['spud'] != 'all':
            if oRequest.POST['spud'] != 'None':
                dArgs['spud__name'] = oRequest.POST['spud']
            else:
                dArgs['spud'] = None

        if oRequest.POST['technology'] != 'all':
            if oRequest.POST['technology'] != 'None':
                dArgs['technology__name'] = oRequest.POST['technology']
            else:
                dArgs['technology'] = None

        # Retrieve all PricingObject objects that match above filters
        aRecords = PricingObject.objects.filter(**dArgs).order_by(
            'part__product_number', 'customer', 'sold_to', 'spud')

        # Create 2x2 list representing datat as it will appear in file
        aPriceList = [[oPO.part.product_number,
                       oPO.customer.name,
                       oPO.sold_to or "(None)",
                       oPO.spud.name if oPO.spud else '(None)',
                       oPO.technology.name if oPO.technology else "(None)",
                       oPO.unit_price,
                       oPO.erosion_rate,
                       '{:.2f}'.format(
                           round(
                               round(
                                   (float(oPO.unit_price) - (
                                       float(oPO.unit_price) * (
                                           float(oPO.erosion_rate)/100)))*100
                               )/100, 2)
                       ),
                       '', '', ] for oPO in aRecords] if aRecords else [[]]

        sFileName = 'Price Erosion Pricing.xlsx'

        # Create file stream from data and return response
        oFile = WritePriceErosionToFile(aPriceList)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
            sFileName)
        response.set_cookie('fileMark', oRequest.POST['file_cookie'],
                            max_age=60)
        oFile.save(response)

        return response
    else:
        return HttpResponse()
# end def


def WritePriceErosionToFile(aPriceList):
    """
    Function to write list of erodible part prices to file
    :param aPriceList: List of lists containing row data to write to file
    :return: OpenPyXL file object
    """
    aHeaders = ['Part Number', 'Customer', 'Sold-To', 'SPUD', 'Technology',
                'Latest Unit Price ($)', 'Erosion Rate (%)',
                'Projected Unit Price ($)', 'Valid From', 'Valid To']
    oFile = openpyxl.Workbook()
    oSheet = oFile.active

    for i in range(len(aHeaders)):
        oSheet[utils.get_column_letter(i + 1) + '1'] = aHeaders[i]

    for i in range(len(aPriceList)):
        for j in range(len(aPriceList[i])):
            oSheet[utils.get_column_letter(j + 1) + str(2 + i)] = \
                aPriceList[i][j]

    return oFile
# end def


def DownloadSearchResults(oRequest):
    """
    View to download search result overview to allow user to generate a
    non-detailed list of Headers matching search criteria.
    :param oRequest: Django HTTP request object
    :return: HTTPResponse containing data file download
    """
    # Retrieve data to download from URL query parameters
    getDict = dict(oRequest.GET)
    keys = list(getDict.keys())
    keys.remove('header')
    keys.sort(key=lambda x: int(x.replace('row', '')))

    aWidths = [0] * len(getDict['header'])

    oHeadingFont = Font(name='Arial', size=10, bold=True)

    sFileName = "Search Results " + \
                datetime.datetime.now().strftime('%Y%m%d_%H%M%S') + ".xlsx"

    # Create new workbook
    oFile = openpyxl.Workbook()
    oSheet = oFile.active

    # Write column title information
    for col in range(len(getDict['header'])):
        oSheet[utils.get_column_letter(col + 1) + '1'] = getDict['header'][col]
        oSheet[utils.get_column_letter(col + 1) + '1'].font = oHeadingFont
        aWidths[col] = max(aWidths[col], len(getDict['header'][col]))

    # Write cell data to table
    row = 2
    for rowKey in keys:
        for col in range(len(getDict[rowKey])):
            oSheet[utils.get_column_letter(col + 1) + str(row)] = \
                getDict[rowKey][col]
            aWidths[col] = max(aWidths[col], len(getDict[rowKey][col]))
        row += 1

    # Set column widths
    for col in range(len(aWidths)):
        oSheet.column_dimensions[utils.get_column_letter(col + 1)].width = \
            aWidths[col] + 3

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment;filename="{0}"'.format(
        sFileName)
    oFile.save(response)

    return response
