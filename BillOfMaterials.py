# Process:     Configuration Management
# Script:      BillOfMaterials
# Author:      eanhamp
# Reviewer:    ekurhag
# ReqNum:      #
# DevStart:    2015-07-17
# QCStart:     2015-07-24
# ProdPush:    YYYY-MM-DD
# Purpose:     Container for bill of materials Excel file information.
# Modified:    ProdPush YYYY-MM-DD, ReqNum, Author, Change Description
#-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
'''Container for bill of materials Excel file information.

BillOfMaterials header- and item-level dictionaries are structures as:

    dHeaderData = {
        [Field name]: [Field value], [Field name]: ...
    }
    dItemData = {
        [Item number]: {
            [Field name]: [Field value], [Field name]: ...
        }
    }
'''

from os import path
from re import match
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


class BillOfMaterialsError(Exception):
    '''Generic exception for BillOfMaterials class errors.'''
# end class

class BillOfMaterials(object):
    '''Container for bill of materials Excel file information.'''
    sHeaderSheet = "1) BOM Config Header"
    sItemSheet = sConfigSheet = "2) BOM Config Entry"
    sTOCSheet = "2a) BoM Config ToC"
    sRevisionSheet = "2b) BoM Config Revision"
    sItemHeaderRow = "12"
    sItemStartRow = str(int(sItemHeaderRow) + 1)
    # Header sheet field names and unique key. Header field dictionary includes
    # unique key in case the unique identifier changes later.
    sHeaderId = "Configuration"
    aHeaderFields = [
        "Person Responsible",
        "REACT Request",
        "BoM Request Type",
        "Customer Unit",
        "Customer Name",
        "Sales Office",
        "Sales Group",
        "Sold-to Party",
        "Ship-to Party",
        "Bill-to Party",
        "Ericsson Contract #",
        "Payment Terms",
        "Projected Cut-over Date",
        "Program",
        "Configuration",
        "Customer Designation",
        "Technology",
        "Product Area 1",
        "Product Area 2",
        "Radio Frequency",
        "Radio Band",
        "Inquiry / Site Template Number",
        "Readiness Complete",
        "Complete Delivery",
        "No ZipRouting",
        "Valid-to Date",
        "Valid-from Date",
        "Shipping Condition",
        "Baseline Impacted",
        "Model",
        "Model Description",
        "What Model is this replacing?",
        "Initial Revision",
        "Configuration/Ordering Status",
        "Workgroup",
        "Name",
        "Is this a pick list?"
    ]
    # Item sheet headers and unique identifier. Item field list includes unique
    # identifier in case the unique identifier changes later.
    sItemId = "Line #"
    aItemFields = [
        "Line #",
        "Product Number",
        "Product Description",
        "Order Qty",
        "UoM",
        "Plant",
        "SLOC",
        "Item Cat",
        "P-Code - Fire Code, Desc",
        "HW/SW Ind",
        "Prod Pkg Type",
        "SPUD",
        "RE-Code",
        "MU-Flag",
        "X-plant matl status",
        "Int Notes",
        "Unit Price",
        "Higher Level Item",
        "Material Group 5",
        "Purch Order Item No",
        "Condition Type",
        "Amount",
        "Traceability Req (Serialization)",
        "Customer Asset?",
        "Customer Asset Tagging Requirement",
        "Customer Number",
        "Second Customer Number",
        "Vendor Article Number",
        "Comments",
        "Additional Reference (if required)"
    ]

    def __init__(self, sBOMDirectory, sFileName):
        self.sFilePathName = path.join(sBOMDirectory, sFileName)

        # Determine if this is a bill of materials file. The pattern is:
        #     "PSM BoM" 0 or more non-newlines ".xlsx"
        if match(r"~.*\.xlsx", sFileName):
            self.bIsBOMFile = False
        else:
            self.bIsBOMFile = True
        # end if

        self.oFile = None
        self.bFileOpen = False
        self.dHeaderData = {}
        self.dItemData = {}
        self.dConfigData = {}
    # end def

    def _GetCustomHeaderKeys(self):
        '''Add user-defined fields to the header dictionary.'''
        if self.oFile.active.title != self.sHeaderSheet:
            try:
                self.oFile.active = self.oFile.sheetnames.index(self.sHeaderSheet)
            except:
                raise BillOfMaterialsError('No sheet found with title: ' + self.sHeaderSheet)
        # end if

        sLastRow = str(self.oFile.active.max_row)
        sLastColumn = get_column_letter(self.oFile.active.max_column)
        # Search all cells within the boundaries for:
        #     (A) Non-blank cells, the value of which is not in aHeaderFields.
        #     (B) Cells with non-blank cells to the right of them, indicating
        #         the cell contains a field and not a value.
        #     (*) Raise an error if a cell is found that contains two
        #         consecutive, non-blank cells to its right. This indicates
        #         improper sheet formatting by the user.
        oRng = self.oFile.active.get_squared_range(1, 1, column_index_from_string(sLastColumn), int(sLastRow))
        aHeaderTest = [sField.upper() for sField in self.aHeaderFields]
        aErrors = []
        oRng = list(oRng)
        for oRow in oRng:
            for oCell in oRow:
                sRow = str(oCell.row)
                sValue = oCell.value
                sTestNonBlank = get_column_letter(column_index_from_string(oCell.column) + 1) + sRow
                sTestBlank = get_column_letter(column_index_from_string(oCell.column) + 2) + sRow
                # Continue on failure conditions.
                if not str(sValue or '').strip():
                    continue
                elif str(sValue or '').upper() in aHeaderTest:
                    continue
                elif not str(self.oFile.active[sTestNonBlank].value or '').strip():
                    continue
                elif str(self.oFile.active[sTestBlank].value or '').strip():
                    aErrors.append(get_column_letter(oCell.column) + sRow)
                    continue
                else:
                    self.aHeaderFields.append(sValue)
                # end if
            # end for
        # end for

        if aErrors:
            sCells = ", ".join(sCell for sCell in aErrors)
            sErr = """
                Improper formatting detected on the the header sheet: data
                found in three consecutive cells (horizontally) starting with
                the following cells: {0}. There must be at least one empty
                column between the pairs of columns that contain fields and
                values on the {1} worksheet.
            """
            raise BillOfMaterialsError(sErr.format(sCells, self.sHeaderSheet))
        # end if
    # end def

    def _PullHeaderData(self):
        '''Pull header-level data from the header sheet of the Excel file.'''
        if self.oFile.active.title != self.sHeaderSheet:
            try:
                self.oFile.active = self.oFile.sheetnames.index(self.sHeaderSheet)
            except:
                raise BillOfMaterialsError('No sheet found with title: ' + self.sHeaderSheet)
        # end if

        # Get data from each field in header dictionary. Use GetCellText, which
        # pulls cell data using the Range.Text property, to ensure the data is
        # pulled as it appears on the worksheet (slower method of pulling data,
        # but the volume of data is low).
        for sField in self.aHeaderFields:
            sRange = self._FindCellByValue(self.oFile.active, sField, 1)
            if sRange:
                sValue = self.oFile.active[sRange].value
                self.dHeaderData[sField] = bool(sValue) if sField in ("Complete Delivery", "No ZipRouting") else (True if sValue and sValue.lower() in ('y', 'yes', 'x') else
                    False) if sField == "Is this a pick list?" else sValue*100 if sField == 'Readiness Complete' and sValue else sValue
            # end if
        # end for
    # end def

    @staticmethod
    def _FindCellByValue(oSheet, sValue, iColOffset=0, iRowOffset=0, bByCol=True):
        for line in (oSheet.columns if bByCol else oSheet.rows):
            for cell in line:
                if cell.value == sValue:
                    return get_column_letter(column_index_from_string(cell.column) +\
                        iColOffset) + str(cell.row + iRowOffset)
            # end for
        # end for
        return
    # end def

    @staticmethod
    def _FindColumnByValue(oSheet, sValue, iRow):
        oRow = oSheet.iter_rows(get_column_letter(oSheet.min_col) + str(iRow) + ":" +\
            get_column_letter(oSheet.max_column) + str(iRow))
        # print(get_column_letter(oSheet.min_col) + str(iRow) + ":" + get_column_letter(oSheet.max_column) + str(iRow))
        oRow = list(oRow)
        # print(oRow)
        for index in range(len(oRow[0])):
            if oRow[0][index].value and oRow[0][index].value.replace('\n',' ').replace('.','') == sValue:
                return get_column_letter(oSheet.min_col + index)
            # end if
        # end for
        return
    # end def

    def _PullItemData(self):
        '''Pull item-level data from the item sheet of the Excel file.'''
        if self.oFile.active.title != self.sItemSheet:
            try:
                self.oFile.active = self.oFile.sheetnames.index(self.sItemSheet)
            except:
                raise BillOfMaterialsError('No sheet found with title: ' + self.sItemSheet)
        # end if

        # Get the last row. It may need to be corrected later, though.
        sLastRow = str(self.oFile.active.max_row)
        sColumn = self._FindColumnByValue(self.oFile.active, self.sItemId, self.sItemHeaderRow)
        sRng = sColumn + self.sItemStartRow + ":" + sColumn + sLastRow
        oRng = self.oFile.active[sRng[:sRng.find(':')]:sRng[sRng.find(':') + 1:]]

        # Iterate across the cells in the Line # column. If a blank cell is
        # found in a row that is less than sLastRow, change the value of
        # sLastRow (Line # is unique identifier; no blanks allowed).
        for oCell in oRng:
            if not oCell[0].value:
                sLastRow = str(oCell[0].row - 1)
                break
            # end if
        # end for

        # Pull the data lists from each column and place them in a temporary
        # dictionary for storing values by field.
        dTemp = {}
        for sField in self.aItemFields:
            sColumn = self._FindColumnByValue(self.oFile.active, sField, self.sItemHeaderRow)
            if sColumn:
                dTemp[sField] = list(self.oFile.active[sColumn + str(self.sItemStartRow): sColumn + sLastRow])\
                    if int(sLastRow) >= int(self.sItemStartRow) else None
            # end if
        # end for

        # Move the data from the temporary dictionary to the class's item-level
        # data dictionary, which is structured as:
        #     {[Item]: {[Field]: [Value], [Field]: Value], ... }, [Item]: ... }
        if dTemp[self.sItemId]:
            for iIndex, sItem in enumerate(dTemp[self.sItemId]):
                self.dItemData[sItem[0].value] = {
                    sField: aData[iIndex][0].value for sField, aData in dTemp.items()
                }
        # end for
    # end def

    def _PullConfigData(self):
        '''Pull header-level data from the header sheet of the Excel file.'''
        if self.oFile.active.title != self.sConfigSheet:
            try:
                self.oFile.active = self.oFile.sheetnames.index(self.sConfigSheet)
            except:
                raise BillOfMaterialsError('No sheet found with title: ' + self.sConfigSheet)
        # end if

        dTemp = {'Reassign': bool(self.oFile.active['D5'].value), 'PSM on Hold': bool(self.oFile.active['D8'].value),
                 'Internal/External Linkage': bool(self.oFile.active['N4'].value),
                 'NET Value': float(self.oFile.active['Q2'].value) if self.oFile.active['Q2'].value else 0.0,
                 'ZPRU Total': float(self.oFile.active['Q4'].value.strip('$ ').replace(',', '')) \
                     if 'x' not in self.oFile.active['Q4'].value else None}

        self.dConfigData.update(dTemp)
    # end def

    def _PullTOCData(self):
        '''Pull header-level data from the header sheet of the Excel file.'''
        if self.oFile.active.title != self.sTOCSheet:
            try:
                self.oFile.active = self.oFile.sheetnames.index(self.sTOCSheet)
            except:
                raise BillOfMaterialsError('No sheet found with title: ' + self.sTOCSheet)
        # end if

        dTemp = {'Internal Notes': self.oFile.active['N13'].value, 'External Notes': self.oFile.active['O13'].value}

        self.dHeaderData.update(dTemp)
    # end def

    def _PullRevisionData(self):
        '''Pull header-level data from the header sheet of the Excel file.'''
        if self.oFile.active.title != self.sRevisionSheet:
            try:
                self.oFile.active = self.oFile.sheetnames.index(self.sRevisionSheet)
            except:
                raise BillOfMaterialsError('No sheet found with title: ' + self.sRevisionSheet)
        # end if

        dTemp = {'System Revision': self.oFile.active['A13'].value, 'Baseline Revision': self.oFile.active['B13'].value,
                 'Release Date': self.oFile.active['C13'].value, 'Changes/Comments': self.oFile.active['F13'].value}

        self.dHeaderData.update(dTemp)
    # end def

    def CalculateTimeSavings(self):
        '''Return the time saved by automatically processing the BoM file.'''
        # Add 2 seconds for each header-level value retrieved ...
        iHeaderSavings = len(self.dHeaderData) * 2
        iItemSavings = 0
        for dInfo in self.dItemData.values():
            # ... and 2 seconds for each item-level value retrieved.
            iItemSavings += len(dInfo) * 2
        # end for
        return iHeaderSavings + iItemSavings
    # end def

    def CloseBOMFile(self):
        '''Close the BOM File.'''
        if self.bFileOpen:
            self.bFileOpen = False
        # end if
    # end def

    def GetBOMIdentifier(self):
        '''Return the BoM identifier.

        This is the BoM-equivalent of a document number. It identifies one BoM
        from another.
        '''
        # If the BOM identifier cannot be pulled from the header-level data
        # dictionary, then get it directly from the Excel file.
        if not self.dHeaderData:
            # Open the file if necessary.
            if not self.bFileOpen:
                self.oFile = openpyxl.load_workbook(self.sFilePathName)
                self.bFileOpen = True
            # end if

            # Select the header-level data worksheet, get the cell address
            # of the BOM identifier field (offset by 1 column to get the
            # value, and pull the text from that cell.
            if self.oFile.active.title != self.sHeaderSheet:
                self.oFile.active = self.oFile.sheetnames.index(self.sHeaderSheet)
            # end if
            sRange = self._FindCellByValue(self.oFile.active, self.sHeaderId, iColOffset=1)
            sReturn = self.oFile.active[sRange].value
        else:
            sReturn = self.dHeaderData[self.sHeaderId]
        # end if
        return sReturn
    # end def

    def PullBOMFileData(self):
        '''Pull header- and item-level data from the Excel file.'''
        # Check if the file is open. Open it if necessary.
        if not self.bFileOpen:
            self.oFile = openpyxl.load_workbook(self.sFilePathName, data_only=True)
            self.bFileOpen = True
        # end if

        self._GetCustomHeaderKeys()
        self._PullHeaderData()
        self._PullItemData()
        self._PullConfigData()
        self._PullTOCData()
        self._PullRevisionData()
    # end def
# end class
#-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=
