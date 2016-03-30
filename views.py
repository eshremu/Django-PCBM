from django.shortcuts import render, redirect
from django.utils import timezone
from django.utils.cache import patch_cache_control
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import logout as auth_logout
from django.views.decorators.csrf import ensure_csrf_cookie
from django.http import HttpResponseRedirect, HttpResponse, QueryDict, JsonResponse, Http404
from django.core.urlresolvers import reverse, resolve
from django.contrib import messages
from django.contrib.sessions.models import Session
from django.db.models import Q
from django.db import connections, IntegrityError
from django import forms
from django.contrib.auth.models import Group

from .models import NewsItem, Alert, Header, Part, Configuration, ConfigLine,\
    PartBase, Baseline, Baseline_Revision, LinePricing, REF_CUSTOMER, REF_REQUEST, HeaderLock, SecurityPermission, ParseException,\
    REF_TECHNOLOGY, REF_PRODUCT_AREA_1, REF_PRODUCT_AREA_2, REF_CUSTOMER_NAME, REF_PROGRAM, REF_CONDITION, REF_MATERIAL_GROUP,\
    REF_PRODUCT_PKG, REF_SPUD, HeaderTimeTracker, REF_RADIO_BAND, REF_RADIO_FREQUENCY
from . import menulisting, headerFile, footerFile, pagetitle, supportcontact
from .forms import HeaderForm, ConfigForm, DateForm, FileUploadForm, SubmitForm
from .templatetags.customtemplatetags import searchscramble
from .BillOfMaterials import BillOfMaterials
from .utils import UpRev, MassUploaderUpdate, GenerateRevisionSummary

import copy
from itertools import chain
import json
import openpyxl
from openpyxl import utils
import os
import re
from functools import cmp_to_key
import traceback

# Create your views here.

aHeaderList = None

class LockException(BaseException):
    pass
# end def


def SetSession(oRequest):
    oRequest.session.clear_expired()
    if oRequest.session.get_expiry_age():
        oRequest.session.set_expiry(0)
    # end if
# end def


def Lock(oRequest, iHeaderPK):
    # print('Calling lock')
    if Header.objects.filter(pk=iHeaderPK):
        if Header.objects.get(pk=iHeaderPK).configuration_status != 'In Process':
            return
        # end if

        if not HeaderLock.objects.filter(header=Header.objects.get(pk=iHeaderPK)):
            HeaderLock.objects.create(**{'header':Header.objects.get(pk=iHeaderPK)})
        # end if

        if HeaderLock.objects.filter(header=Header.objects.get(pk=iHeaderPK)).filter(session_key=None):
            HeaderLock.objects.filter(header=Header.objects.get(pk=iHeaderPK)).filter(session_key=None).update(
                **{'session_key': Session.objects.get(session_key=oRequest.session.session_key)}
            )
        elif HeaderLock.objects.filter(header=Header.objects.get(pk=iHeaderPK)).filter(
                session_key=Session.objects.get(session_key=oRequest.session.session_key)):
            pass
        else:
            raise LockException('File is locked for editing')
    else:
        raise Header.DoesNotExist('No such Header found')
    # end if
# end def


def Unlock(oRequest, iHeaderPK):
    # print('Calling unlock')
    if Header.objects.filter(pk=iHeaderPK):
        if HeaderLock.objects.filter(header=Header.objects.get(pk=iHeaderPK)).filter(session_key=Session.objects.get(session_key=oRequest.session.session_key)):
            HeaderLock.objects.filter(header=Header.objects.get(pk=iHeaderPK)).filter(
                session_key=Session.objects.get(session_key=oRequest.session.session_key)
            ).update(**{'session_key': None})
        else:
            pass #raise LockException('Already unlocked')
    else:
        raise Header.DoesNotExist('No existing header')
# end def


def Logout(oRequest):
    try:
        Unlock(oRequest, iHeaderPK=oRequest.session.get('existing', None))
    except Header.DoesNotExist:
        pass
    # end try
    oRequest.session.set_expiry(1)
    return auth_logout(oRequest,next_page= '/pcbm/')
# end def


def FinalUnlock(oRequest):
    # print('Calling final')
    if oRequest.method == 'POST':
        try:
            Unlock(oRequest, iHeaderPK=oRequest.session.get('existing', None))
        except (LockException, Header.DoesNotExist):
            pass
        except Exception as ex:
            print(traceback.format_exc())
        # end try

        if oRequest.POST.get('close',None) == 'true':
            # oRequest.session.set_expiry(3600)
            oRequest.session.set_expiry(0)

        response = HttpResponse()
        response['Content-Length'] = 0
        # response['Connection'] = 'close'
        return response
    else:
        raise Http404()
# end def


def InitialLock(oRequest):
    # print('Calling init lock')
    if oRequest.method == 'POST':
        try:
            SetSession(oRequest)
            Lock(oRequest, iHeaderPK=oRequest.session.get('existing', None))
        except (LockException, Header.DoesNotExist):
            pass
        except Exception as ex:
            print(traceback.format_exc())
        # end try

        response = HttpResponse()
        response['Content-Length'] = 0
        # response['Connection'] = 'close'
        return response
    else:
        raise Http404()
# end def


def Index(oRequest):
    SetSession(oRequest)
    if 'existing' in oRequest.session:
        try:
            Unlock(oRequest, oRequest.session['existing'])
        except:
            pass
        # end try

        del oRequest.session['existing']
    # end if

    aLatestAlerts = Alert.objects.filter(PublishDate__lte=timezone.now()).order_by('-PublishDate')
    aLatestNews = NewsItem.objects.filter(PublishDate__lte=timezone.now()).order_by('-PublishDate')

    dContext = {
        'latest_alerts': aLatestAlerts,
        'latest_news': aLatestNews,
    }
    return Default(oRequest, 'BoMConfig/index.html', dContext)
# end def


def AddHeader(oRequest, sTemplate='BoMConfig/entrylanding.html'):
    # existing_instance is the existing header. Store the pk in the form to return for saving
    # Status message allows another view to redirect to here with an error message explaining the redirect
    bCanReadHeader = bool(SecurityPermission.objects.filter(title='Config_Header_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteHeader = bool(SecurityPermission.objects.filter(title='Config_Header_Write').filter(user__in=oRequest.user.groups.all()))
    bSuccess = False

    # Determine which pages to which the user is able to move forward
    bCanReadConfig = bool(SecurityPermission.objects.filter(title='Config_Entry_BOM_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadTOC = bool(SecurityPermission.objects.filter(title='Config_ToC_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadRevision = bool(SecurityPermission.objects.filter(title='Config_Revision_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadInquiry = bool(SecurityPermission.objects.filter(title='SAP_Inquiry_Creation_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadSiteTemplate = bool(SecurityPermission.objects.filter(title='SAP_ST_Creation_Read').filter(user__in=oRequest.user.groups.all()))

    bCanMoveForward = bCanReadConfig or bCanReadTOC or bCanReadRevision or bCanReadInquiry or bCanReadSiteTemplate

    if sTemplate == 'BoMConfig/entrylanding.html':
        return redirect(reverse('bomconfig:configheader'))
    else:
        oExisting = oRequest.session.get('existing', None)
        status_message = oRequest.session.get('status', None)

        if status_message:
            del oRequest.session['status']
        # end if

        try:
            if oExisting:
                Lock(oRequest, oExisting)
                oExisting = Header.objects.get(pk=oRequest.session['existing'])
            # end if

            if oRequest.method == 'POST' and oRequest.POST:
                oModPost = None
                if not oExisting and 'configuration_status' not in oRequest.POST:
                    oModPost = QueryDict(None, mutable=True)
                    oModPost.update(oRequest.POST)
                    oModPost.update({'configuration_status': 'In Process'})
                    headerForm = HeaderForm(oModPost, instance=oExisting, readonly=not bCanWriteHeader)
                else:
                    headerForm = HeaderForm(oRequest.POST, instance=oExisting, readonly=not bCanWriteHeader)
                # end if

                if oRequest.POST['baseline_impacted'] and oRequest.POST['baseline_impacted'] == 'New' and oRequest.POST.get('new_baseline', None):
                    Baseline.objects.get_or_create(title__iexact=oRequest.POST['new_baseline'],
                                                   defaults={'title':oRequest.POST['new_baseline'], 'customer': REF_CUSTOMER.objects.get(id=oRequest.POST['customer_unit'])})
                    headerForm.data._mutable = True
                    headerForm.data['baseline_impacted'] = oRequest.POST['new_baseline']
                    headerForm.data._mutable = False
                # end if

                if headerForm.is_valid():
                    if headerForm.cleaned_data['configuration_status'] == 'In Process':
                        try:
                            if bCanWriteHeader:
                                oHeader = headerForm.save(commit=False)
                                oHeader.shipping_condition = '71'
                                oHeader.save()

                                if not hasattr(oHeader, 'configuration'):
                                    oConfig = Configuration.objects.create(**{'header': oHeader})

                                    (oPartBase,_) = PartBase.objects.get_or_create(**{'product_number':oHeader.configuration_designation})
                                    oPartBase.unit_of_measure = 'PC'
                                    oPartBase.save()
                                    (oPart,_) = Part.objects.get_or_create(**{'base': oPartBase,'product_description':oHeader.model_description})

                                    oConfigLine = ConfigLine.objects.create(**{
                                        'config': oConfig,
                                        'part': oPart,
                                        'line_number': '10',
                                        'order_qty': 1,
                                        'vendor_article_number': oHeader.configuration_designation,
                                    })
                                # end if

                                if not hasattr(oHeader, 'headerlock'):
                                    oHeadLock = HeaderLock.objects.create(**{
                                        'header': oHeader,
                                        'session_key': Session.objects.get(session_key=oRequest.session.session_key)
                                    })
                                # end if

                                status_message = oRequest.session['status'] = 'Form data saved'
                            else:
                                oHeader = oExisting
                                status_message = None
                            #end if


                            bSuccess = True
                        except IntegrityError as ex:
                            # status_message = 'Configuration already exists in Baseline'
                            status_message = str(ex)
                            print(str(ex))
                        # end try
                    else:
                        oHeader = oExisting
                        bSuccess = True
                    # end if

                    if bSuccess:
                        if oRequest.POST['formaction'] == 'save':
                            oRequest.session['existing'] = oHeader.pk
                            oExisting = oHeader
                        elif oRequest.POST['formaction'] == 'saveexit':
                            oRequest.session['existing'] = oHeader.pk
                            return redirect('bomconfig:index')
                        elif oRequest.POST['formaction'] == 'next':
                            oRequest.session['existing'] = oHeader.pk
                            sDestination = 'bomconfig:configheader'
                            if bCanReadConfig:
                                sDestination = 'bomconfig:config'
                            elif bCanReadTOC:
                                sDestination = 'bomconfig:configtoc'
                            elif bCanReadRevision:
                                sDestination = 'bomconfig:configrevision'
                            elif bCanReadInquiry:
                                sDestination = 'bomconfig:configinquiry'
                            elif bCanReadSiteTemplate:
                                sDestination = 'bomconfig:configsite'
                            # end if

                            return redirect(sDestination)
                        # end if
                    # end if
                else:
                    status_message = 'Error(s) occurred'
                # end if
                headerForm.fields['product_area2'].queryset = REF_PRODUCT_AREA_2.objects.filter(parent=(headerForm.cleaned_data['product_area1'] if headerForm.cleaned_data.get('product_area1', None) else None))
                headerForm.fields['program'].queryset = REF_PROGRAM.objects.filter(parent=(headerForm.cleaned_data['customer_unit'] if headerForm.cleaned_data.get('customer_unit', None) else None))
            else:
                headerForm = HeaderForm(instance=oExisting, readonly=not bCanWriteHeader)
                headerForm.fields['product_area2'].queryset = REF_PRODUCT_AREA_2.objects.filter(parent=(oExisting.product_area1 if oExisting else None))
                headerForm.fields['program'].queryset = REF_PROGRAM.objects.filter(parent=(oExisting.customer_unit if oExisting else None))
            # end if
        except LockException:
            headerForm = HeaderForm(readonly=not bCanWriteHeader)
            status_message = 'File is locked for editing'
            if 'existing' in oRequest.session:
                del oRequest.session['existing']
            #end if
        # end try

        # headerForm.fields['person_responsible'] = fields.ChoiceField(choices=[('1','Paul'),('2','You')]) # This will set the 'Person Responsible' field to a drop-down list of PSM users

        if not oExisting or type(oExisting) != str and oExisting.configuration_status == 'In Process':
            headerForm.fields['baseline_impacted'].widget = forms.widgets.Select(choices=(('','---------'),('New','Create New baseline')) + tuple((obj.title,obj.title) for obj in Baseline_Revision.objects.filter(baseline__customer=oExisting.customer_unit if oExisting else None).filter(completed_date=None)))

        if not oExisting or (oExisting and type(oExisting) != str and oExisting.configuration_status == 'In Process'):
            oCursor = connections['REACT'].cursor()
            oCursor.execute('SELECT DISTINCT [Customer] FROM ps_fas_contracts WHERE [CustomerUnit]=%s',[oExisting.customer_unit.name if oExisting else None])
            tResults = oCursor.fetchall()
            headerForm.fields['customer_name'].widget = forms.widgets.Select(choices=(('','---------'),) + tuple((obj,obj) for obj in chain.from_iterable(tResults)))

        dContext={
            'headerForm': headerForm,
            'break_list': ('Payment Terms', 'Shipping Condition', 'Initial Version', 'Configuration/Ordering Status', 'Name'),
            'status_message': status_message,
            'header_write_authorized': bCanWriteHeader,
            'header_read_authorized': bCanReadHeader,
            'can_continue': bCanMoveForward
        }

        return Default(oRequest, sTemplate, dContext)
# end def


def AddConfig(oRequest):
    oHeader = oRequest.session.get('existing', None)
    status_message = oRequest.session.get('status', None)

    bCanReadConfigBOM = bool(SecurityPermission.objects.filter(title='Config_Entry_BOM_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteConfigBOM = bool(SecurityPermission.objects.filter(title='Config_Entry_BOM_Write').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigSAP = bool(SecurityPermission.objects.filter(title='Config_Entry_SAPDoc_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteConfigSAP = bool(SecurityPermission.objects.filter(title='Config_Entry_SAPDoc_Write').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigAttr = bool(SecurityPermission.objects.filter(title='Config_Entry_Attributes_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteConfigAttr = bool(SecurityPermission.objects.filter(title='Config_Entry_Attributes_Write').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigPrice = bool(SecurityPermission.objects.filter(title='Config_Entry_PriceLinks_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteConfigPrice = bool(SecurityPermission.objects.filter(title='Config_Entry_PriceLinks_Write').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigCust = bool(SecurityPermission.objects.filter(title='Config_Entry_CustomerData_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteConfigCust = bool(SecurityPermission.objects.filter(title='Config_Entry_CustomerData_Write').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigBaseline = bool(SecurityPermission.objects.filter(title='Config_Entry_Baseline_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteConfigBaseline = bool(SecurityPermission.objects.filter(title='Config_Entry_Baseline_Write').filter(user__in=oRequest.user.groups.all()))

    # Determine which pages to which the user is able to move forward
    bCanReadHeader = bool(SecurityPermission.objects.filter(title='Config_Header_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadTOC = bool(SecurityPermission.objects.filter(title='Config_ToC_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadRevision = bool(SecurityPermission.objects.filter(title='Config_Revision_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadInquiry = bool(SecurityPermission.objects.filter(title='SAP_Inquiry_Creation_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadSiteTemplate = bool(SecurityPermission.objects.filter(title='SAP_ST_Creation_Read').filter(user__in=oRequest.user.groups.all()))

    bCanMoveForward = bCanReadTOC or bCanReadRevision or bCanReadInquiry or bCanReadSiteTemplate
    bCanMoveBack = bCanReadHeader

    bCanReadConfig = bCanReadConfigBOM or bCanReadConfigSAP or bCanReadConfigAttr or bCanReadConfigPrice or bCanReadConfigCust or bCanReadConfigBaseline
    bCanWriteConfig = bCanWriteConfigBOM or bCanWriteConfigSAP or bCanWriteConfigAttr or bCanWriteConfigPrice or bCanWriteConfigCust or bCanWriteConfigBaseline

    try:
        if oHeader:
            Lock(oRequest, oHeader)
            oHeader = Header.objects.get(pk=oHeader)
        else:
            oRequest.session['status'] = 'Define header first'
            return redirect(reverse('bomconfig:configheader'))
        # end if
    except LockException:
        oRequest.session['status'] = 'File locked for editing'
        return redirect(reverse('bomconfig:configheader'))
    # end try

    if status_message:
        del oRequest.session['status']
    # end if

    if oRequest.method == 'POST' and oRequest.POST:
        configForm = ConfigForm(oRequest.POST, instance=oHeader.configuration if oHeader and hasattr(oHeader, 'configuration') else None)
        if configForm.is_valid():
            if bCanWriteConfig:
                oConfig = configForm.save(commit=False)
                if not hasattr(oHeader, 'configuration'):
                    oConfig.header = oHeader
                # end if
                oConfig.save()

                if oHeader.configuration_status == 'In Process':
                    oForm = json.loads(oRequest.POST['data_form'])
                    # Clean empty rows out and remove previous row statuses
                    for index in range(len(oForm) - 1, -1, -1):
                        # Convert to dict for uniformity
                        if type(oForm[index]) == list:
                            oForm[index] = {str(x):y for (x,y) in enumerate(oForm[index])}

                        if '0' in oForm[index]:
                            del oForm[index]['0']

                        if all(x in (None, '') for x in list(oForm[index].values())) or len(oForm[index]) < 1:
                            del oForm[index]
                            continue
                        else:
                            for x in range(31):
                                if str(x) not in oForm[index]:
                                    oForm[index][str(x)] = None
                            # end for
                        # end if
                    # end for

                    ConfigLine.objects.filter(config=oConfig).delete()

                    for dConfigLine in oForm:
                        dBaseData = {'product_number': dConfigLine['2'].strip('. '), 'unit_of_measure': dConfigLine['5']}

                        (oBase, _) = PartBase.objects.get_or_create(product_number=dConfigLine['2'].strip('. '))
                        PartBase.objects.filter(pk=oBase.pk).update(**dBaseData)

                        dPartData = {'product_description': dConfigLine['3'].strip().upper() if dConfigLine['3'] else dConfigLine['3'], 'base': oBase}

                        oNullQuery = Q(product_description__isnull=True)
                        oNoneQuery = Q(product_description='None')
                        oBlankQuery = Q(product_description='')
                        if Part.objects.filter(base=oBase).filter(oNoneQuery | oNullQuery | oBlankQuery) and not Part.objects.filter(**dPartData):
                            iPartPk = Part.objects.filter(base=oBase).filter(oNoneQuery | oNullQuery | oBlankQuery)[0].pk
                            Part.objects.filter(pk=iPartPk).update(**dPartData)
                            oPart = Part.objects.get(pk=iPartPk)
                        else:
                            (oPart, _) = Part.objects.get_or_create(**dPartData)

                        dLineData = {'line_number': dConfigLine['1'], 'order_qty': dConfigLine['4'] or None, 'spud': dConfigLine['12'],
                                     'plant': dConfigLine['6'], 'sloc': dConfigLine['7'], 'item_category': dConfigLine['8'],
                                     'internal_notes': dConfigLine['16'],# 'unit_price': dConfigLine['17'],
                                     'higher_level_item': dConfigLine['18'], 'material_group_5': dConfigLine['19'],
                                     'purchase_order_item_num': dConfigLine['20'], 'condition_type': dConfigLine['21'],
                                     'amount': dConfigLine['22'] or None, 'customer_asset': dConfigLine['24'],
                                     'customer_asset_tagging': dConfigLine['25'], 'customer_number': dConfigLine['26'],
                                     'sec_customer_number': dConfigLine['27'], 'vendor_article_number': dConfigLine['28'],
                                     'comments': dConfigLine['29'], 'additional_ref': dConfigLine['30'], 'config': oConfig,
                                     'part': oPart, 'is_child': bool(dConfigLine['2'].startswith('.') and not dConfigLine['2'].startswith('..')),
                                     'is_grandchild': bool(dConfigLine['2'].startswith('..')),
                                     'pcode': dConfigLine['9'], 'commodity_type': dConfigLine['10'], 'REcode': dConfigLine['13'],
                                     'package_type': dConfigLine['11'], 'mu_flag': dConfigLine['14'], 'x_plant': dConfigLine['15'],
                                     'traceability_req': dConfigLine['23']}

                        if ConfigLine.objects.filter(config=oConfig).filter(line_number=dConfigLine['1']):
                            oConfigLine = ConfigLine.objects.get(**{'config': oConfig, 'line_number': dConfigLine['1']})
                            # oConfigLine = ConfigLine.objects.filter(config=oConfig).filter(line_number=dConfigLine['1'])[0].pk
                            # ConfigLine.objects.filter(pk=oConfigLine).update(**dLineData)
                            ConfigLine.objects.filter(pk=oConfigLine.pk).update(**dLineData)
                        else:
                            oConfigLine = ConfigLine(**dLineData)
                            oConfigLine.save()
                        # end if

                        dPriceData = {'unit_price': dConfigLine['17'] or None, 'config_line': oConfigLine}
                        (oPrice, _) = LinePricing.objects.get_or_create(**dPriceData)
                    # end for

                    status_message = oRequest.session['status'] = 'Form data saved'
                # end if
            #end if
            if oRequest.POST['formaction'] == 'prev':
                oRequest.session['existing'] = oHeader.pk
                sDestination = 'bomconfig:config'
                if bCanReadHeader:
                    sDestination = 'bomconfig:configheader'
                # end if

                return redirect(sDestination)
            elif oRequest.POST['formaction'] == 'saveexit':
                oRequest.session['existing'] = oHeader.pk
                return redirect('bomconfig:index')
            elif oRequest.POST['formaction'] == 'save':
                oRequest.session['existing'] = oHeader.pk
                if 'status' in oRequest.session:
                    del oRequest.session['status']
                # end if
            elif oRequest.POST['formaction'] == 'next':
                oRequest.session['existing'] = oHeader.pk
                sDestination = 'bomconfig:config'
                if bCanReadTOC:
                    sDestination = 'bomconfig:configtoc'
                elif bCanReadRevision:
                    sDestination = 'bomconfig:configrevision'
                elif bCanReadInquiry:
                    sDestination = 'bomconfig:configinquiry'
                elif bCanReadSiteTemplate:
                    sDestination = 'bomconfig:configsite'
                # end if

                return redirect(sDestination)
            # end if
        else:
            status_message = configForm.errors['__all__'].as_text()
        # end if
    else:
        configForm = ConfigForm(instance=oHeader.configuration if oHeader and hasattr(oHeader, 'configuration') else None,)
    # end if

    data = BuildDataArray(oHeader, config=True)

    dContext = {
        'data_array': data,
        'form': configForm,
        'header': oHeader,
        'status_message': status_message,
        'config_bom_read_authorized': bCanReadConfigBOM,
        'config_sap_read_authorized': bCanReadConfigSAP,
        'config_attr_read_authorized': bCanReadConfigAttr,
        'config_price_read_authorized': bCanReadConfigPrice,
        'config_cust_read_authorized': bCanReadConfigCust,
        'config_baseline_read_authorized': bCanReadConfigBaseline,
        'config_read_authorized': bCanReadConfig,
        'config_bom_write_authorized': bCanWriteConfigBOM,
        'config_sap_write_authorized': bCanWriteConfigSAP,
        'config_attr_write_authorized': bCanWriteConfigAttr,
        'config_price_write_authorized': bCanWriteConfigPrice,
        'config_cust_write_authorized': bCanWriteConfigCust,
        'config_baseline_write_authorized': bCanWriteConfigBaseline,
        'config_write_authorized': bCanWriteConfig,
        'can_continue': bCanMoveForward,
        'can_previous': bCanMoveBack,
        'condition_list': [obj.name for obj in REF_CONDITION.objects.all()],
        'material_group_list': [obj.name for obj in REF_MATERIAL_GROUP.objects.all()],
        'product_pkg_list': [obj.name for obj in REF_PRODUCT_PKG.objects.all()],
        'spud_list': [obj.name for obj in REF_SPUD.objects.all()],
    }
    return Default(oRequest, sTemplate='BoMConfig/configuration.html', dContext=dContext)
# end def


def AddTOC(oRequest):
    oHeader = oRequest.session.get('existing', None)
    status_message = oRequest.session.get('status', None)

    bCanReadTOC = bool(SecurityPermission.objects.filter(title='Config_ToC_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteTOC = bool(SecurityPermission.objects.filter(title='Config_ToC_Write').filter(user__in=oRequest.user.groups.all()))

    # Determine which pages to which the user is able to move forward
    bCanReadHeader = bool(SecurityPermission.objects.filter(title='Config_Header_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfig = bool(SecurityPermission.objects.filter(title='Config_Entry_BOM_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadRevision = bool(SecurityPermission.objects.filter(title='Config_Revision_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadInquiry = bool(SecurityPermission.objects.filter(title='SAP_Inquiry_Creation_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadSiteTemplate = bool(SecurityPermission.objects.filter(title='SAP_ST_Creation_Read').filter(user__in=oRequest.user.groups.all()))

    bCanMoveForward = bCanReadRevision or bCanReadInquiry or bCanReadSiteTemplate
    bCanMoveBack = bCanReadHeader or bCanReadConfig

    try:
        if oHeader:
            Lock(oRequest, oHeader)
            oHeader = Header.objects.get(pk=oHeader)
        else:
            oRequest.session['status'] = 'Define header first'
            return redirect(reverse('bomconfig:configheader'))
        # end if
    except LockException:
        oRequest.session['status'] = 'File locked for editing'
        return redirect(reverse('bomconfig:configheader'))
    # end try

    if status_message:
        del oRequest.session['status']

    if oRequest.method == 'POST' and oRequest.POST:
        configForm = ConfigForm(oRequest.POST, instance=oHeader.configuration if oHeader and hasattr(oHeader, 'configuration') else None)
        if configForm.is_valid():
            if bCanWriteTOC:
                oConfig = configForm.save(commit=False)
                if not hasattr(oHeader, 'configuration'):
                    oConfig.header = oHeader
                # end if
                oConfig.save()

                if oHeader.configuration_status == 'In Process':
                    oForm = json.loads(oRequest.POST['data_form'])
                    if type(oForm[0]) == list:
                        oForm[0] = {str(x):y for (x,y) in enumerate(oForm[0])}

                    if '1' in oForm[0] and oForm[0]['1'] not in ('', None):
                        oHeader.customer_designation = oForm[0]['1']
                    if '13' in oForm[0] and oForm[0]['13'] not in ('', None):
                        oHeader.internal_notes = oForm[0]['13']
                    if '14' in oForm[0] and oForm[0]['14'] not in ('', None):
                        oHeader.external_notes = oForm[0]['14']

                    oHeader.save()
                    status_message = oRequest.session['status'] = 'Form data saved'
                # end if
            # end if

            if oRequest.POST['formaction'] == 'prev':
                oRequest.session['existing'] = oHeader.pk
                sDestination = 'bomconfig:configtoc'
                if bCanReadConfig:
                    sDestination = 'bomconfig:config'
                elif bCanReadHeader:
                    sDestination = 'bomconfig:configheader'
                # end if

                return redirect(sDestination)
            elif oRequest.POST['formaction'] == 'saveexit':
                oRequest.session['existing'] = oHeader.pk
                return redirect('bomconfig:index')
            elif oRequest.POST['formaction'] == 'save':
                oRequest.session['existing'] = oHeader.pk
                if 'status' in oRequest.session:
                    del oRequest.session['status']
                # end if
            elif oRequest.POST['formaction'] == 'next':
                oRequest.session['existing'] = oHeader.pk
                sDestination = 'bomconfig:configtoc'
                if bCanReadRevision:
                    sDestination = 'bomconfig:configrevision'
                elif bCanReadInquiry:
                    sDestination = 'bomconfig:configinquiry'
                elif bCanReadSiteTemplate:
                    sDestination = 'bomconfig:configsite'
                # end if

                return redirect(sDestination)
            # end if
        # end if
    else:
        configForm = ConfigForm(instance=oHeader.configuration if oHeader and hasattr(oHeader, 'configuration') else None,
                                initial={'net_value': getattr(oHeader.configuration,'net_value') if hasattr(oHeader, 'configuration') else '0',
                                         'zpru_total': getattr(oHeader.configuration,'zpru_total') if hasattr(oHeader, 'configuration') else '0'})
    # end if

    data = BuildDataArray(oHeader=oHeader, toc=True)

    dContext = {
        'data_array': data,
        'status_message': status_message,
        'form': configForm,
        'header': oHeader,
        'toc_read_authorized': bCanReadTOC,
        'toc_write_authorized': bCanWriteTOC,
        'can_continue': bCanMoveForward,
        'can_previous': bCanMoveBack
    }
    return Default(oRequest, sTemplate='BoMConfig/configtoc.html', dContext=dContext)
# end def


def AddRevision(oRequest):
    error_matrix = []
    valid = True
    oForm = None
    oHeader = oRequest.session.get('existing', None)
    status_message = oRequest.session.get('status', None)

    bCanReadRevision = bool(SecurityPermission.objects.filter(title='Config_Revision_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteRevision = bool(SecurityPermission.objects.filter(title='Config_Revision_Write').filter(user__in=oRequest.user.groups.all()))

    # Determine which pages to which the user is able to move forward
    bCanReadHeader = bool(SecurityPermission.objects.filter(title='Config_Header_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfig = bool(SecurityPermission.objects.filter(title='Config_Entry_BOM_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadTOC = bool(SecurityPermission.objects.filter(title='Config_ToC_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadInquiry = bool(SecurityPermission.objects.filter(title='SAP_Inquiry_Creation_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadSiteTemplate = bool(SecurityPermission.objects.filter(title='SAP_ST_Creation_Read').filter(user__in=oRequest.user.groups.all()))

    bCanMoveForward = bCanReadInquiry or bCanReadSiteTemplate
    bCanMoveBack = bCanReadHeader or bCanReadConfig or bCanReadTOC

    try:
        if oHeader:
            Lock(oRequest, oHeader)
            oHeader = Header.objects.get(pk=oHeader)
        else:
            oRequest.session['status'] = 'Define header first'
            return redirect(reverse('bomconfig:configheader'))
        # end if
    except LockException:
        oRequest.session['status'] = 'File locked for editing'
        return redirect(reverse('bomconfig:configheader'))
    # end try

    if status_message:
        del oRequest.session['status']

    if oRequest.method == 'POST' and oRequest.POST:
        oForm = json.loads(oRequest.POST['data_form'])
        if type(oForm[0]) == list:
            oForm[0] = {str(x):y for (x,y) in enumerate(oForm[0])}

        if '2' in oForm[0] and oForm[0]['2'] not in ('', None):
            form = DateForm({'date': oForm[0]['2']})
            if form.is_valid():
                oHeader.release_date = form.cleaned_data['date']
            else:
                error_matrix.append([None, None, 'X - Not a valid date.'])
                valid = False
            # end if
        if '5' in oForm[0] and oForm[0]['5'] not in ('', None):
            oHeader.change_comments = oForm[0]['5']

        if valid:
            try:
                if oHeader.configuration_status == 'In Process' and bCanWriteRevision:
                    oHeader.save()
                    status_message = oRequest.session['status'] = 'Form data saved'
                # end if

                if oRequest.POST['formaction'] == 'prev':
                    sDestination = 'bomconfig:configrevision'
                    if bCanReadTOC:
                        sDestination = 'bomconfig:configtoc'
                    elif bCanReadConfig:
                        sDestination = 'bomconfig:config'
                    elif bCanReadHeader:
                        sDestination = 'bomconfig:configheader'
                    # end if

                    return redirect(sDestination)
                elif oRequest.POST['formaction'] == 'saveexit':
                    return redirect('bomconfig:index')
                elif oRequest.POST['formaction'] == 'save':
                    if 'status' in oRequest.session:
                        del oRequest.session['status']
                    # end if
                elif oRequest.POST['formaction'] == 'next':
                    sDestination = 'bomconfig:configrevision'
                    if bCanReadInquiry:
                        sDestination = 'bomconfig:configinquiry'
                    elif bCanReadSiteTemplate:
                        sDestination = 'bomconfig:configsite'
                    # end if

                    return redirect(sDestination)
                # end if
            except:
                status_message = 'Configuration already exists in Baseline'
        # end if
    # end if

    data = BuildDataArray(oHeader, revision=True)
    if not valid and oForm:
        data[0].update({'2': oForm[0]['2']})
    dContext = {
        'data_array': data,
        'status_message': status_message,
        'header': oHeader,
        'error_matrix': error_matrix,
        'revision_read_authorized': bCanReadRevision,
        'revision_write_authorized': bCanWriteRevision,
        'can_continue': bCanMoveForward,
        'can_previous': bCanMoveBack
    }
    return Default(oRequest, sTemplate='BoMConfig/configrevision.html', dContext=dContext)
# end def


def AddInquiry(oRequest, inquiry):
    oHeader = oRequest.session.get('existing', None)
    status_message = oRequest.session.get('status', None)

    bCanReadInquiry = bool(SecurityPermission.objects.filter(title='SAP_Inquiry_Creation_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadSiteTemplate = bool(SecurityPermission.objects.filter(title='SAP_ST_Creation_Read').filter(user__in=oRequest.user.groups.all()))

    # Determine which pages to which the user is able to move forward
    bCanReadHeader = bool(SecurityPermission.objects.filter(title='Config_Header_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfig = bool(SecurityPermission.objects.filter(title='Config_Entry_BOM_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadTOC = bool(SecurityPermission.objects.filter(title='Config_ToC_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadRevision = bool(SecurityPermission.objects.filter(title='Config_Revision_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadInquiry = bool(SecurityPermission.objects.filter(title='SAP_Inquiry_Creation_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadSiteTemplate = bool(SecurityPermission.objects.filter(title='SAP_ST_Creation_Read').filter(user__in=oRequest.user.groups.all()))

    bCanMoveForward = bCanReadSiteTemplate and inquiry
    bCanMoveBack = bCanReadHeader or bCanReadConfig or bCanReadTOC or bCanReadRevision or (not inquiry and bCanReadInquiry)

    try:
        if oHeader:
            Lock(oRequest, oHeader)
            oHeader = Header.objects.get(pk=oHeader)
        else:
            oRequest.session['status'] = 'Define header first'
            return redirect(reverse('bomconfig:configheader'))
        # end if
    except LockException:
        oRequest.session['status'] = 'File locked for editing'
        return redirect(reverse('bomconfig:configheader'))
    # end try

    if status_message:
        del oRequest.session['status']

    if oRequest.method == 'POST' and oRequest.POST:
        if oRequest.POST['formaction'] == 'prev':
            sDestination = 'bomconfig:configinquiry' if inquiry else 'bomconfig:configsite'
            if bCanReadInquiry and not inquiry:
                sDestination = 'bomconfig:configinquiry'
            elif bCanReadRevision:
                sDestination = 'bomconfig:configrevision'
            elif bCanReadTOC:
                sDestination = 'bomconfig:configtoc'
            elif bCanReadConfig:
                sDestination = 'bomconfig:config'
            elif bCanReadHeader:
                sDestination = 'bomconfig:configheader'
            # end if

            return redirect(sDestination)
        elif oRequest.POST['formaction'] == 'next':
            sDestination = 'bomconfig:configinquiry' if inquiry else 'bomconfig:configsite'
            if bCanReadSiteTemplate and inquiry:
                sDestination = 'bomconfig:configsite'
            # end if

            return redirect(sDestination)
    # end if

    data = BuildDataArray(oHeader=oHeader, inquiry=inquiry, site=not inquiry)
    configForm = ConfigForm(instance=oHeader.configuration if oHeader and hasattr(oHeader, 'configuration') else None)
    configForm.fields['internal_external_linkage'].widget.attrs['disabled'] = True

    dContext = {
        'data_array': data,
        'status_message': status_message,
        'inquiry': inquiry,
        'form': configForm,
        'header': oHeader,
        'inquiry_read_authorized': bCanReadInquiry,
        'sitetemplate_read_authorized': bCanReadSiteTemplate,
        'can_continue': bCanMoveForward,
        'can_previous': bCanMoveBack
    }

    return Default(oRequest, sTemplate='BoMConfig/inquiry.html', dContext=dContext)
# end def


def Search(oRequest, advanced=False):
    if 'existing' in oRequest.session:
        try:
            Unlock(oRequest, oRequest.session['existing'])
        except:
            pass
        # end try

        del oRequest.session['existing']
    # end if

    if 'status' in oRequest.session:
        del oRequest.session['status']
    # end if

    if advanced:
        sTemplate = 'BoMConfig/searchadv.html'
    else:
        sTemplate = 'BoMConfig/search.html'
    # end if

    link = oRequest.GET.get('link', None)
    config = oRequest.GET.get('config', None)
    if link or config:
        if link:
            oRequest.session['existing'] = link[5:-10]
            return redirect(reverse('bomconfig:configheader'))
        else:
            oRequest.session['existing'] = config[5:-10]
            return redirect(reverse('bomconfig:config'))
        # end if
    # end if

    aHeaders = Header.objects.all()

    if oRequest.method == 'POST' and oRequest.POST:

        if not advanced:

            if 'config_design' in oRequest.POST and oRequest.POST['config_design'] != '':
                aHeaders = aHeaders.filter(configuration_designation__iregex="^" + oRequest.POST['config_design'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
            # end if

            if 'person' in oRequest.POST and oRequest.POST['person'] != '':
                aHeaders = aHeaders.filter(person_responsible__iregex="^" + oRequest.POST['person'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
            # end if

            if 'request' in oRequest.POST and oRequest.POST['request'] != '':
                aHeaders = aHeaders.filter(bom_request_type__name=oRequest.POST['request'])
            # end if

            if 'customer' in oRequest.POST and oRequest.POST['customer'] != '':
                aHeaders = aHeaders.filter(customer_unit__name=oRequest.POST['customer'])
            # end if

            if 'status' in oRequest.POST and oRequest.POST['status'] != '':
                aHeaders = aHeaders.filter(configuration_status=oRequest.POST['status'].replace('_', ' ').title())
            # end if

            results = HttpResponse()
            if aHeaders:
                results.write('<h5 style="color:red">Found ' + str(len(aHeaders)) + ' matching record(s)</h5>')
                results.write('<table><thead><tr><th style="width:175px;">Configuration</th>' +
                              '<th style="width:175px;">Version</th><th style="width:175px;">Person Responsible</th>' +
                              '<th style="width:175px;">BoM Request Type</th><th style="width:175px;">Customer Unit</th>' +
                              '<th style="width:175px;">Status</th></tr></thead><tbody>')
                for header in aHeaders:
                    results.write('<tr><td><a href="?link={0}">{1}</a></td><td>{2}</td><td>{3}</td><td>{4}</td><td>{5}</td><td>{6}</td></tr>'\
                        .format(searchscramble(header.pk), header.configuration_designation, header.baseline_version,
                                header.person_responsible, header.bom_request_type.name, header.customer_unit.name, header.configuration_status))
                # end for
                results.write('</tbody></table>')
            else:
                results.write('NO CONFIGURATIONS MATCHING SEARCH')
            # end if

            return results
        else:
            bRemoveDuplicates = True
            sTableHeader = '<table><thead><tr><th style="width:175px;">Configuration</th><th style="width:175px;">Version</th>'
            """ This will be a list of strings.  Each string will be the dot-operator-separated string of attributes
             that would retrieve the desired value (i.e.: 'config.configline.part.description')
             This will be so that the search results list can be easily repeated"""
            aLineFilter = []
            aResults = []

            aConfigLines = ConfigLine.objects.all()

            if 'config_design' in oRequest.POST and oRequest.POST['config_design'] != '':
                # aHeaders = aHeaders.filter(configuration_designation__iregex="^" + oRequest.POST['config_design']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__configuration_designation__iregex="^" + oRequest.POST['config_design'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")

            if 'request' in oRequest.POST and oRequest.POST['request'] != '':
                # aHeaders = aHeaders.filter(react_request__iregex="^" + oRequest.POST['request']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__react_request__iregex="^" + oRequest.POST['request'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">REACT Request</th>'
                aLineFilter.append('config.header.react_request')

            if 'customer' in oRequest.POST and oRequest.POST['customer'] != '':
                # aHeaders = aHeaders.filter(customer_unit=oRequest.POST['customer'])
                aConfigLines = aConfigLines.filter(config__header__customer_unit__name=oRequest.POST['customer'])
                sTableHeader += '<th style="width:175px;">Customer</th>'
                aLineFilter.append('config.header.customer_unit.name')

            if 'person' in oRequest.POST and oRequest.POST['person'] != '':
                # aHeaders = aHeaders.filter(person_responsible__iregex="^" + oRequest.POST['person']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__person_responsible__iregex="^" + oRequest.POST['person'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">Person Responsible</th>'
                aLineFilter.append('config.header.person_responsible')

            if 'sold_to' in oRequest.POST and oRequest.POST['sold_to'] != '':
                # aHeaders = aHeaders.filter(sold_to_party__iregex="^" + oRequest.POST['sold_to']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__sold_to_party__iregex="^" + oRequest.POST['sold_to'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">Sold-to Party</th>'
                aLineFilter.append('config.header.sold_to_party')

            if 'program' in oRequest.POST and oRequest.POST['program'] != '':
                # aHeaders = aHeaders.filter(program__iregex="^" + oRequest.POST['program']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__program__iregex="^" + oRequest.POST['program'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">Program</th>'
                aLineFilter.append('config.header.program')

            if 'technology' in oRequest.POST and oRequest.POST['technology'] != '':
                # aHeaders = aHeaders.filter(technology__iregex="^" + oRequest.POST['technology']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__technology__name__iregex="^" + oRequest.POST['technology'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">Technology</th>'
                aLineFilter.append('config.header.technology.name')

            if 'base_impact' in oRequest.POST and oRequest.POST['base_impact'] != '':
                # aHeaders = aHeaders.filter(baseline_impacted__iregex="^" + oRequest.POST['base_impact']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__baseline_impacted__iregex="^" + oRequest.POST['base_impact'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">Baseline Impacted</th>'
                aLineFilter.append('config.header.baseline_impacted')

            if 'model' in oRequest.POST and oRequest.POST['model'] != '':
                # aHeaders = aHeaders.filter(model__iregex="^" + oRequest.POST['model']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__model__iregex="^" + oRequest.POST['model'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">Model</th>'
                aLineFilter.append('config.header.model')

            if 'model_desc' in oRequest.POST and oRequest.POST['model_desc'] != '':
                # aHeaders = aHeaders.filter(model_description__iregex="^" + oRequest.POST['model_desc']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__model_description__iregex="^" + oRequest.POST['model_desc'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">Model Description</th>'
                aLineFilter.append('config.header.model_description')

            if 'init_rev' in oRequest.POST and oRequest.POST['init_rev'] != '':
                # aHeaders = aHeaders.filter(initial_revision__iregex="^" + oRequest.POST['init_rev']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__initial_revision__iregex="^" + oRequest.POST['init_rev'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">Initial Revision</th>'
                aLineFilter.append('config.header.initial_revision')

            if 'status' in oRequest.POST and oRequest.POST['status'] != '':
                # aHeaders = aHeaders.filter(configuration_status=oRequest.POST['status'])
                aConfigLines = aConfigLines.filter(config__header__configuration_status__iexact=oRequest.POST['status'].replace('_', ' '))
                sTableHeader += '<th style="width:175px;">Configuration Status</th>'
                aLineFilter.append('config.header.configuration_status')

            if 'inquiry_site' in oRequest.POST and oRequest.POST['inquiry_site'] != '':
                # aHeaders = aHeaders.filter(inquiry_site_template__iregex="^" + oRequest.POST['inquiry_site']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__inquiry_site_template__iregex="^" + oRequest.POST['inquiry_site'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTableHeader += '<th style="width:175px;">Inquiry / Site Template Number</th>'
                aLineFilter.append('config.header.inquiry_site_number')

            sTempHeaderLine = ''
            aTempFilters = ['line_number']

            if 'product_num' in oRequest.POST and oRequest.POST['product_num'] != '':
                # aConfigs = Configuration.objects.filter(configline__part__base__product_number__iregex="^" + oRequest.POST['product_num']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$").distinct()
                # aHeaders = aHeaders.filter(configuration__in=aConfigs)
                aConfigLines = aConfigLines.filter(part__base__product_number__iregex="^" + oRequest.POST['product_num'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTempHeaderLine += '<th style="width:175px;">Product Number</th><th style="width:175px;">SPUD</th>'
                aTempFilters.append('part.base.product_number')
                aTempFilters.append('spud')
                bRemoveDuplicates = False

            if 'description' in oRequest.POST and oRequest.POST['description'] != '':
                # aConfigs = Configuration.objects.filter(configline__part__product_description__iregex="^" + oRequest.POST['description']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$").distinct()
                # aHeaders = aHeaders.filter(configuration__in=aConfigs)
                aConfigLines = aConfigLines.filter(part__product_description__iregex="^" + oRequest.POST['description'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTempHeaderLine += '<th style="width:175px;">Product Description</th>'
                aTempFilters.append('part.product_description')
                bRemoveDuplicates = False

            if 'customer_num' in oRequest.POST and oRequest.POST['customer_num'] != '':
                # aConfigs = Configuration.objects.filter(configline__customer_number__iregex="^" + oRequest.POST['customer_num']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$").distinct()
                # aHeaders = aHeaders.filter(configuration__in=aConfigs)
                aConfigLines = aConfigLines.filter(customer_number__iregex="^" + oRequest.POST['customer_num'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTempHeaderLine += '<th style="width:175px;">Customer Number</th>'
                aTempFilters.append('customer_number')
                bRemoveDuplicates = False

            if 'mu_flag' in oRequest.POST and oRequest.POST['mu_flag'] != '':
                # aConfigs = Configuration.objects.filter(configline__part__base__mu_flag__iregex="^" + oRequest.POST['mu_flag']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$").distinct()
                # aHeaders = aHeaders.filter(configuration__in=aConfigs)
                aConfigLines = aConfigLines.filter(mu_flag__iregex="^" + oRequest.POST['mu_flag'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTempHeaderLine += '<th style="width:175px;">MU-Flag</th>'
                aTempFilters.append('mu_flag')
                bRemoveDuplicates = False

            if 'xplant' in oRequest.POST and oRequest.POST['xplant'] != '':
                # aConfigs = Configuration.objects.filter(configline__part__base__x_plant__iregex="^" + oRequest.POST['xplant']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$").distinct()
                # aHeaders = aHeaders.filter(configuration__in=aConfigs)
                aConfigLines = aConfigLines.filter(x_plant__iregex="^" + oRequest.POST['xplant'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                sTempHeaderLine += '<th style="width:175px;">X-Plant Material Status</th>'
                aTempFilters.append('x_plant')
                bRemoveDuplicates = False

            if sTempHeaderLine:
                sTableHeader += '<th style="width:175px;">Line Number</th>' + sTempHeaderLine
                aLineFilter.extend(aTempFilters)

            if 'base_rev' in oRequest.POST and oRequest.POST['base_rev'] != '':
                # aHeaders = aHeaders.filter(baseline_version__iregex="^" + oRequest.POST['base_rev']
                #                            .replace('.', '\.').replace('?','.').replace('*', '.*') + "$")
                aConfigLines = aConfigLines.filter(config__header__baseline_version__iregex="^" + oRequest.POST['base_rev'].strip()
                                           .replace(' ','\W').replace('.', '\.').replace('?','.').replace('*', '.*') + "$")

            if 'release' in oRequest.POST and oRequest.POST['release'] != '':
                # aHeaders = aHeaders.filter(release_date__lte=oRequest.POST['release'])
                aConfigLines = aConfigLines.filter(**{'config__header__release_date__' + oRequest.POST['release_param']: oRequest.POST['release']})
                sTableHeader += '<th style="width:175px;">Release Date</th>'
                aLineFilter.append('config.header.release_date')

            if bRemoveDuplicates:
                aResults = [obj.config for obj in aConfigLines]
                aResults = set(aResults)
                aLineFilter = [sFilter[7:] if sFilter.startswith('config.') else sFilter for sFilter in aLineFilter]
            else:
                aResults = aConfigLines

            results = HttpResponse()
            if aResults:
                results.write('<h5 style="color:red">Found ' + str(len(aResults)) + ' matching record(s)</h5>')
                results.write(sTableHeader + "</tr></thead><tbody>")
                for oResult in aResults:
                    results.write('<tr><td><a href="?link={0}">{1}</a></td><td>{2}</td>'.format(
                        searchscramble(GrabValue(oResult,'config.header.pk' if isinstance(oResult, ConfigLine) else 'header.pk')),
                        GrabValue(oResult,'config.header.configuration_designation' if isinstance(oResult, ConfigLine)
                            else 'header.configuration_designation'),
                        GrabValue(oResult,'config.header.baseline_version' if isinstance(oResult, ConfigLine) else 'header.baseline_version')
                        )
                    )

                    for sFilter in aLineFilter:
                        results.write('<td>' + GrabValue(oResult, sFilter) + '</td>')
                    # end for

                    results.write('</tr>')
                # end for
                results.write('</tbody></table>')
            else:
                results.write('NO CONFIGURATIONS MATCHING SEARCH')
            # end if

            return results
        # end if
    else:
        aHeaders = list(aHeaders)[-10:]
    # end if

    dContext = {
        'header_list': aHeaders,
    }
    return Default(oRequest, sTemplate=sTemplate, dContext=dContext)
# end def


def GrabValue(oStartObj, sAttrChain):
    import functools

    return str(functools.reduce(lambda x,y:getattr(x, y),sAttrChain.split('.'),oStartObj) or '')
# end def


def Upload(oRequest):
    if 'existing' in oRequest.session:
        try:
            Unlock(oRequest, oRequest.session['existing'])
        except:
            pass
        # end try

        del oRequest.session['existing']
    # end if

    if 'status' in oRequest.session:
        del oRequest.session['status']
    # end if

    if oRequest.method == 'POST' and oRequest.POST and oRequest.FILES:
        form = FileUploadForm(oRequest.POST, oRequest.FILES)
        if form.is_valid():
            try:
                sBOM = ParseUpload(oRequest.FILES['file'])
                dContext = {
                    'title': sBOM,
                }
                return Default(oRequest, sTemplate='BoMConfig/success.html', dContext=dContext)
            except Exception as ex:
                if 'No sheet found with title:' in str(ex):
                    messages.add_message(oRequest, messages.ERROR, 'File specified is not a correctly formatted Configuration file.')
                else:
                    print(traceback.format_exc())

                    if isinstance(ex, IntegrityError):
                        messages.add_message(oRequest, messages.ERROR, 'A record matching this upload already exists in the database.')
                    else:
                        messages.add_message(oRequest, messages.ERROR, 'An error occurred during upload. Please check input and/or notify tool admin.')
                # end if
            # end try
        else:
            for error in form.non_field_errors():
                messages.add_message(oRequest, messages.ERROR, error)
            # end for
        # end if
    else:
        if oRequest.method == 'POST' and not oRequest.FILES:
            messages.add_message(oRequest, messages.ERROR, 'No file specified.')
        # end if

        form = FileUploadForm()
    # end if


    dContext ={
        'form': form
    }
    return Default(oRequest, sTemplate='BoMConfig/upload.html', dContext=dContext)
# end def


def ParseUpload(oSourceFile):
    import platform
    import tempfile
    sOS = platform.system()

    stringisless = lambda x,y:bool(len(x.strip('1234567890')) < len(y.strip('1234567890'))
                                                                 or list(x.strip('1234567890')) < (['']*(len(x.strip('1234567890'))-len(y.strip('1234567890'))) +
                                                                                                   list(y.strip('1234567890')))) \
                              or (x.strip('1234567890') == y.strip('1234567890') and list(x) < list(y))

    if sOS in ('Linux', 'Windows'):
        # Write uploaded file to temporary file
        Temp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        for chunk in oSourceFile.chunks():
            Temp.write(chunk)
        # end for
        Temp.close()

        # Process temporary file through BoM Parser
        oBOM = BillOfMaterials(tempfile.gettempdir(), os.path.abspath(Temp.name))
        oBOM.PullBOMFileData()

        # Delete temporary file
        os.remove(Temp.name)

        if oBOM.dHeaderData['Configuration/Ordering Status'] not in ['In Process','Obsolete','Active','Discontinued','Cancelled']:
            raise ParseException(str(oBOM.dHeaderData['Configuration/Ordering Status']) + " is not a valid Configuration/Ordering Status")

        oBaseline = None
        # Build Baseline, Header, Configuration, PartBases, Parts, and ConfigLines for uploaded file
        if oBOM.dHeaderData['Baseline Impacted'] not in (None,'', 'None'):
            cleanedBaseline = oBOM.dHeaderData['Baseline Impacted'].strip()
            if cleanedBaseline.lower().endswith('rev'):
                cleanedBaseline = cleanedBaseline[:-3].strip()
            oBOM.dHeaderData['Baseline Impacted'] = cleanedBaseline
            (oBaseline, _) = Baseline.objects.get_or_create(title=oBOM.dHeaderData['Baseline Impacted'],customer=REF_CUSTOMER.objects.get(name=oBOM.dHeaderData['Customer Unit']))
        # endif

        # If file is not an 'In Process' Configuration
        if oBOM.dHeaderData['Configuration/Ordering Status'] != 'In Process' and oBaseline:
            # Catch the Baseline up to the revision being uploaded, if the file revision exceeds the current active revision
            while stringisless(oBaseline.current_active_version, oBOM.dHeaderData['Baseline Revision']):
                UpRev(oBaseline, oBOM.dHeaderData['Configuration'],oBOM.dHeaderData['Baseline Revision'])
            # end while


            # Get or create the baseline_revision to which the file should be assigned
            # i.e.: If file is in revision A1, but revision A1 does not exist, create it
            (oBaseRev, oNew) = Baseline_Revision.objects.get_or_create(**{'baseline': oBaseline, 'version':oBOM.dHeaderData['Baseline Revision']})
            if oNew:
                oBaseRev.completed_date = timezone.now()
                oBaseRev.save()
        # Else if file is in-process, change file baseline_version to current inprocess version
        elif oBOM.dHeaderData['Configuration/Ordering Status'] == 'In Process' and oBaseline:
            oBOM.dHeaderData['Baseline Revision'] = oBaseline.current_inprocess_version
        # end if

        # Make Header argument dictionary
        dHeadDict = {
            'person_responsible': oBOM.dHeaderData['Person Responsible'],
            'react_request': oBOM.dHeaderData['REACT Request'],
            'bom_request_type': REF_REQUEST.objects.get(name__iexact=oBOM.dHeaderData['BoM Request Type']),
            'customer_unit': REF_CUSTOMER.objects.get(name__iexact=oBOM.dHeaderData['Customer Unit']),
            'customer_name': oBOM.dHeaderData['Customer Name'],
            'sales_office': oBOM.dHeaderData['Sales Office'],
            'sales_group': oBOM.dHeaderData['Sales Group'],
            'sold_to_party': oBOM.dHeaderData['Sold-to Party'],
            'ship_to_party': oBOM.dHeaderData['Ship-to Party'],
            'bill_to_party': oBOM.dHeaderData['Bill-to Party'],
            'ericsson_contract': oBOM.dHeaderData['Ericsson Contract #'],
            'payment_terms': oBOM.dHeaderData['Payment Terms'],
            'projected_cutover': oBOM.dHeaderData['Projected Cut-over Date'],
            'program': REF_PROGRAM.objects.get(name__iexact=oBOM.dHeaderData['Program'], parent=REF_CUSTOMER.objects.get(name__iexact=oBOM.dHeaderData['Customer Unit']))\
                if REF_PROGRAM.objects.filter(name__iexact=oBOM.dHeaderData['Program'], parent=REF_CUSTOMER.objects.get(name__iexact=oBOM.dHeaderData['Customer Unit'])) else None,
            'configuration_designation': oBOM.dHeaderData['Configuration'],
            'customer_designation': oBOM.dHeaderData['Customer Designation'],
            'technology': REF_TECHNOLOGY.objects.get(name__iexact=oBOM.dHeaderData['Technology']) if REF_TECHNOLOGY.objects.filter(name__iexact=oBOM.dHeaderData['Technology']) else None,
            'product_area1': REF_PRODUCT_AREA_1.objects.get(name__iexact=oBOM.dHeaderData['Product Area 1']) if REF_PRODUCT_AREA_1.objects.filter(name__iexact=oBOM.dHeaderData['Product Area 1']) else None,
            'product_area2': REF_PRODUCT_AREA_2.objects.get(name__iexact=oBOM.dHeaderData['Product Area 2']) if REF_PRODUCT_AREA_2.objects.filter(name__iexact=oBOM.dHeaderData['Product Area 2']) else None,
            'radio_frequency': REF_RADIO_FREQUENCY.objects.get(name__iexact=oBOM.dHeaderData['Radio Frequency']) if REF_RADIO_FREQUENCY.objects.filter(name__iexact=oBOM.dHeaderData['Radio Frequency']) else None,
            'radio_band': REF_RADIO_BAND.objects.get(name__iexact=oBOM.dHeaderData['Radio Band']) if REF_RADIO_BAND.objects.filter(name__iexact=oBOM.dHeaderData['Radio Band']) else None,
            # 'optional_free_text1': oBOM.dHeaderData['Person Responsible'],
            # 'optional_free_text2': oBOM.dHeaderData['Person Responsible'],
            # 'optional_free_text3': oBOM.dHeaderData['Person Responsible'],
            'inquiry_site_template': oBOM.dHeaderData['Inquiry / Site Template Number'],
            'readiness_complete': oBOM.dHeaderData['Readiness Complete'],
            'complete_delivery': oBOM.dHeaderData['Complete Delivery'],
            'no_zip_routing': oBOM.dHeaderData['No ZipRouting'],
            'valid_to_date': oBOM.dHeaderData['Valid-to Date'],
            'valid_from_date': oBOM.dHeaderData['Valid-from Date'],
            'shipping_condition': oBOM.dHeaderData['Shipping Condition'],
            'baseline_impacted': oBOM.dHeaderData['Baseline Impacted'],
            'model': oBOM.dHeaderData['Model'],
            'model_description': oBOM.dHeaderData['Model Description'],
            'model_replaced': oBOM.dHeaderData['What Model is this replacing?'],
            'initial_revision': oBOM.dHeaderData['Initial Revision'],
            'configuration_status': oBOM.dHeaderData['Configuration/Ordering Status'],
            'old_configuration_status': None,
            'workgroup': oBOM.dHeaderData['Workgroup'],
            'name': oBOM.dHeaderData['Name'],
            'pick_list': oBOM.dHeaderData['Is this a pick list?'],
            'internal_notes': oBOM.dHeaderData['Internal Notes'],
            'external_notes': oBOM.dHeaderData['External Notes'],
            'baseline_version': oBOM.dHeaderData['Baseline Revision'],
            'bom_version': oBOM.dHeaderData['System Revision'],
            'release_date': oBOM.dHeaderData['Release Date'],
            'change_comments': oBOM.dHeaderData['Changes/Comments'],
            'baseline': Baseline_Revision.objects.get(baseline=oBaseline,version=oBOM.dHeaderData['Baseline Revision'])\
                if Baseline_Revision.objects.filter(baseline=oBaseline,version=oBOM.dHeaderData['Baseline Revision']) else None,
        }

        # Attempt to find Header with configuration and baseline matching upload
        try:
            try:
                tempBase = Baseline_Revision.objects.get(baseline=oBaseline,version=oBOM.dHeaderData['Baseline Revision'])
            except Baseline_Revision.DoesNotExist:
                tempBase = None
            # end try

            oHeader = Header.objects.get(configuration_designation=oBOM.GetBOMIdentifier(),
                                         baseline_version__iexact=oBOM.dHeaderData['Baseline Revision'], baseline=tempBase,
                                         program=REF_PROGRAM.objects.get(name__iexact=oBOM.dHeaderData['Program'], parent=REF_CUSTOMER.objects.get(name__iexact=oBOM.dHeaderData['Customer Unit'])) if REF_PROGRAM.objects.filter(name__iexact=oBOM.dHeaderData['Program'], parent=REF_CUSTOMER.objects.get(name__iexact=oBOM.dHeaderData['Customer Unit'])) else None)
        except Header.DoesNotExist:
            oHeader = None
        # end try

        # Update / Create new Header
        if getattr(oHeader, 'configuration_status', None) in ('In Process', None):
            oHeader = Header(pk=getattr(oHeader,'pk', None),**dHeadDict)
            oHeader.save()
        else:
            raise ValueError('Active Configuration ' + str(oHeader) + ' already exists.')
        # end if

        """
        At this point, the Header is either newly entered or an update of an existing 'In Process' configuration,
        so it can be overwritten with the data provided.
        """
        try:
            oConfig = Configuration.objects.get(header=oHeader)
        except Configuration.DoesNotExist:
            oConfig = None
        # end try

        dConfigDict = {
            'reassign': oBOM.dConfigData['Reassign'],
            'PSM_on_hold': oBOM.dConfigData['PSM on Hold'],
            'internal_external_linkage': oBOM.dConfigData['Internal/External Linkage'],
            'net_value': "{:.2f}".format(round(float(oBOM.dConfigData['NET Value']),2)),
            'zpru_total': "{:.2f}".format(round(float(oBOM.dConfigData['ZPRU Total']), 2)) if oBOM.dConfigData['ZPRU Total'] else None,
            'needs_zpru': False,
            'header': oHeader
        }

        oConfig = Configuration(pk=getattr(oConfig,'pk', None),**dConfigDict)
        oConfig.save()

        ConfigLine.objects.filter(config=oConfig).delete()

        for sKey in oBOM.dItemData.keys():
            dPartBaseDict = {
                'product_number': oBOM.dItemData[sKey]['Product Number'].strip('.'),
                'unit_of_measure': oBOM.dItemData[sKey]['UoM'],
            }

            (oBase,_) = PartBase.objects.get_or_create(product_number=dPartBaseDict['product_number'])
            PartBase(pk=getattr(oBase, 'pk'),**dPartBaseDict).save()

            dPartDict = {
                'product_description': oBOM.dItemData[sKey]['Product Description'],
                'base': oBase,
            }

            (oPart,_) = Part.objects.get_or_create(**dPartDict)

            dConfigLineDict = {
                'line_number': oBOM.dItemData[sKey]['Line #'],
                'order_qty': oBOM.dItemData[sKey]['Order Qty'],
                'plant': oBOM.dItemData[sKey]['Plant'],
                'sloc': oBOM.dItemData[sKey]['SLOC'],
                'item_category': oBOM.dItemData[sKey]['Item Cat'],
                'spud': oBOM.dItemData[sKey]['SPUD'],
                'internal_notes': oBOM.dItemData[sKey]['Int Notes'],
                'higher_level_item': oBOM.dItemData[sKey]['Higher Level Item'],
                'material_group_5': oBOM.dItemData[sKey]['Material Group 5'],
                'purchase_order_item_num': oBOM.dItemData[sKey]['Purch Order Item No'],
                'condition_type': oBOM.dItemData[sKey]['Condition Type'],
                'amount': oBOM.dItemData[sKey]['Amount'],
                'customer_asset': oBOM.dItemData[sKey]['Customer Asset?'],
                'customer_asset_tagging': oBOM.dItemData[sKey]['Customer Asset Tagging Requirement'],
                'customer_number': oBOM.dItemData[sKey]['Customer Number'],
                'sec_customer_number': oBOM.dItemData[sKey]['Second Customer Number'],
                'vendor_article_number': oBOM.dItemData[sKey]['Vendor Article Number'],
                'comments': oBOM.dItemData[sKey]['Comments'],
                'additional_ref': oBOM.dItemData[sKey]['Additional Reference (if required)'],
                'config': oConfig,
                'part': oPart,
                'is_child': bool(str(sKey).count('.') == 1),
                'is_grandchild': bool(str(sKey).count('.') == 2),
                'pcode': oBOM.dItemData[sKey]['P-Code - Fire Code, Desc'],
                'commodity_type': oBOM.dItemData[sKey]['HW/SW Ind'],
                'package_type': oBOM.dItemData[sKey]['Prod Pkg Type'],
                'REcode': oBOM.dItemData[sKey]['RE-Code'],
                'mu_flag': oBOM.dItemData[sKey]['MU-Flag'],
                'x_plant': oBOM.dItemData[sKey]['X-plant matl status'],
                'traceability_req': oBOM.dItemData[sKey]['Traceability Req (Serialization)'],
            }

            oConfigLine = ConfigLine.objects.create(**dConfigLineDict)

            dLinePrice = {
                'unit_price': "{:.2f}".format(round(float(oBOM.dItemData[sKey]['Unit Price']), 2)) if oBOM.dItemData[sKey]['Unit Price'] else None,
                'override_price': None,
                'config_line': oConfigLine
            }

            LinePricing.objects.create(**dLinePrice)
        # end for

        if oBaseline:
            MassUploaderUpdate(oBaseline)

        return oBOM.GetBOMIdentifier()
    else:
        raise EnvironmentError('The system does not support your operating system')
    # end if
# end def


@login_required
def Login(oRequest):
    return HttpResponseRedirect(reverse('bomconfig:index', current_app=resolve(oRequest.path).app_name))
# end def


@ensure_csrf_cookie
def Default(oRequest, sTemplate='BoMConfig/template.html', dContext=None, show_footer=False):
    sUserId = None

    if oRequest.user.is_authenticated() and oRequest.user.is_active:
        sUserId = oRequest.user.username
    # end if

    bCanReadHeader = bool(SecurityPermission.objects.filter(title='Config_Header_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWriteHeader = bool(SecurityPermission.objects.filter(title='Config_Header_Write').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigBOM = bool(SecurityPermission.objects.filter(title='Config_Entry_BOM_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigSAP = bool(SecurityPermission.objects.filter(title='Config_Entry_SAPDoc_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigAttr = bool(SecurityPermission.objects.filter(title='Config_Entry_Attributes_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigPrice = bool(SecurityPermission.objects.filter(title='Config_Entry_PriceLinks_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigCust = bool(SecurityPermission.objects.filter(title='Config_Entry_CustomerData_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadConfigBaseline = bool(SecurityPermission.objects.filter(title='Config_Entry_Baseline_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadTOC = bool(SecurityPermission.objects.filter(title='Config_ToC_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadRevision = bool(SecurityPermission.objects.filter(title='Config_Revision_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadInquiry = bool(SecurityPermission.objects.filter(title='SAP_Inquiry_Creation_Read').filter(user__in=oRequest.user.groups.all()))
    bCanReadSiteTemplate = bool(SecurityPermission.objects.filter(title='SAP_ST_Creation_Read').filter(user__in=oRequest.user.groups.all()))

    if not dContext:
        dContext = {}

    dContext.update([
        ('menulisting', menulisting,),
        ('header_template', headerFile,),
        ('footer_template', footerFile,),
        ('pagetitle', pagetitle,),
        ('supportcontact', supportcontact,),
        ('userId', sUserId,),
        ('show_footer', show_footer,),
        ('header_write_authorized', bCanWriteHeader),
        ('header_read_authorized', bCanReadHeader),
        ('config_read_authorized', bCanReadConfigBOM),
        ('config_sap_read_authorized', bCanReadConfigSAP),
        ('config_attr_read_authorized', bCanReadConfigAttr),
        ('config_price_read_authorized', bCanReadConfigPrice),
        ('config_cust_read_authorized', bCanReadConfigCust),
        ('config_baseline_read_authorized', bCanReadConfigBaseline),
        ('toc_read_authorized', bCanReadTOC),
        ('revision_read_authorized', bCanReadRevision),
        ('inquiry_read_authorized', bCanReadInquiry),
        ('sitetemplate_read_authorized', bCanReadSiteTemplate),
    ])

    response = render(oRequest, sTemplate, dContext)
    patch_cache_control(response, no_cache=True, no_store=True, must_revalidate=True, pragma='no-cache', max_age=0)
    return response
# end def


def BuildDataArray(oHeader=None, config=False, toc=False, inquiry=False, site=False, revision=False):
    """Builds an array to pass to a HTML instance of Handsontable."""
    if not bool(config) ^ bool(toc) ^ bool(inquiry) ^ bool(site) ^ bool(revision):
        raise AssertionError('Only one argument should be true')
    # end if

    if config or inquiry or site:
        if (not oHeader) or (oHeader and not hasattr(oHeader, 'configuration')):
            if config:
                if not oHeader.pick_list:
                    return [{'1': '10', '2': oHeader.configuration_designation.upper(), '3': oHeader.model_description, '4': '1',
                         '5': 'PC'}, ]
                else:
                    return [{}]
            else:
                return [{'0': '10', '1': oHeader.configuration_designation.upper(), '2': oHeader.model_description, '3': '1'},]
            # end if
        # end if

        oConfig = oHeader.configuration
        aConfigLines = ConfigLine.objects.filter(config=oConfig).order_by('line_number')
        aConfigLines = sorted(aConfigLines, key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')])
        aData = []
        for Line in aConfigLines:
            if config:
                dLine ={'1': Line.line_number,
                        '2': ('..' if Line.is_grandchild else '.' if Line.is_child else '') + Line.part.base.product_number,
                        '3': Line.part.product_description, '4': Line.order_qty, '5': Line.part.base.unit_of_measure,
                        '6': Line.plant, '7': ('0' * (4 - len(str(Line.sloc)))) + Line.sloc if Line.sloc else Line.sloc,
                        '8': Line.item_category, '9': Line.pcode, '10': Line.commodity_type,
                        '11': Line.package_type, '12': Line.spud, '13': Line.REcode,
                        '14': Line.mu_flag,
                        '15': ('0' * (2 - len(Line.x_plant))) + Line.x_plant if Line.x_plant else '',
                        '16': Line.internal_notes,
                        '17': "!" + str(Line.linepricing.override_price) if str(Line.line_number) == '10' and
                                hasattr(Line,'linepricing') and Line.linepricing.override_price else Line.linepricing.unit_price if hasattr(Line,'linepricing') else '',
                        '18': Line.higher_level_item, '19': Line.material_group_5,
                        '20': Line.purchase_order_item_num, '21': Line.condition_type, '22': Line.amount,
                        '23': Line.traceability_req, '24': Line.customer_asset, '25': Line.customer_asset_tagging,
                        '26': Line.customer_number, '27': Line.sec_customer_number, '28': Line.vendor_article_number,
                        '29': Line.comments, '30': Line.additional_ref}
            else:
                LineNumber = None
                if '.' not in Line.line_number:
                    try:
                        LineNumber = int(Line.line_number)
                    except (ValueError, TypeError):
                        pass  # This ensures that we only continue when the stored line number is an integer
                    # end try
                # end if

                if LineNumber:
                    dLine = {'0': Line.line_number, '1': Line.part.base.product_number, '2': Line.part.product_description,
                             '3': Line.order_qty, '4': Line.plant, '5': Line.sloc,
                             '6': Line.item_category}
                    if inquiry:
                        dLine.update({'7': Line.linepricing.override_price if str(Line.line_number) == '10' and hasattr(Line,'linepricing')
                                                                              and Line.linepricing.override_price else Line.linepricing.unit_price if hasattr(Line,'linepricing') else '',
                                      '9': Line.material_group_5, '10': Line.purchase_order_item_num,
                                      '11': Line.condition_type, '12': Line.amount})
                        if LineNumber == 10 and not oHeader.pick_list:
                            dLine.update({'8': oHeader.configuration.net_value})
                    else:
                        dLine.update(({'7': Line.higher_level_item}))
                    # end if
                else:
                    continue
            # end if

            for key in copy.deepcopy(dLine):
                if not dLine[key]:
                    del dLine[key]
                # end if
            # end for
            aData.append(dLine)
        # end for
        return aData
    elif toc:
        if not oHeader:
            return [[]]
        else:
            return [{'0': oHeader.configuration_designation, '1': oHeader.customer_designation or '', '2': oHeader.technology.name if oHeader.technology else '',
                     '3': oHeader.product_area1.name if oHeader.product_area1 else '', '4': oHeader.product_area2.name if oHeader.product_area2 else '', '5': oHeader.model or '', '6': oHeader.model_description or '',
                     '7': oHeader.model_replaced or '', '9': oHeader.bom_request_type.name, '10': oHeader.configuration_status,
                     '11': oHeader.inquiry_site_template if str(oHeader.inquiry_site_template).startswith('1') else '',
                     '12': oHeader.inquiry_site_template if str(oHeader.inquiry_site_template).startswith('4') else '',
                     '13': (oHeader.internal_notes or ''), '14': (oHeader.external_notes or '')}]
    elif revision:
        if not oHeader or not hasattr(oHeader, 'configuration'):
            return [{}]
        else:
            data = []
            aPrev = Header.objects.filter(configuration_designation=oHeader.configuration_designation).filter(program=oHeader.program).order_by('baseline_version')

            for header in aPrev:
                if header == oHeader:
                    break

                data.append({
                    '0': header.bom_version, '1': header.baseline_version, '2': header.release_date.strftime('%b. %d, %Y') if header.release_date else '',
                    '3': header.model if not header.pick_list else 'None', '6': header.person_responsible,
                    '4': header.configuration.configline_set.filter(line_number='10')[0].customer_number if
                        not header.pick_list and oHeader.configuration.configline_set.filter(line_number='10')[0].customer_number else '',
                    '5': header.change_comments or ''
                })
            # end for

            data.append({
                '0': oHeader.bom_version, '1': oHeader.baseline_version, '2': oHeader.release_date.strftime('%b. %d, %Y') if header.release_date else '',
                '3': oHeader.model if not oHeader.pick_list else 'None', '6': oHeader.person_responsible,
                '4': oHeader.configuration.configline_set.filter(line_number='10')[0].customer_number if
                    not oHeader.pick_list and oHeader.configuration.configline_set.filter(line_number='10')[0].customer_number else '',
                '5': oHeader.change_comments or ''
            })

            return list(reversed(data))
        # end if
    # end if

    return [{}]
# end def


def Validator(oRequest):
    if oRequest.method == "POST" and oRequest.POST:
        form_data = json.loads(oRequest.POST['entered_data'])
        net_total = 0
        override_total = 0
        base_total = None
        first_line = None
        zpru_total = 0
        needs_zpru = False
        status = 'GOOD'

        Parent = 0
        Child = 0
        Grandchild = 0

        # Convert list of lists to list of dicts
        if type(form_data[1]) == list:
            for index in range(len(form_data)):
                form_data[index] = {str(key): (value if value != '' else None) for key, value in zip(range(len(form_data[index])), form_data[index])}
            #end for)

        # Clean empty rows out and remove previous row statuses
        for index in range(len(form_data) - 1, -1, -1):
            if '0' in form_data[index]:
                del form_data[index]['0']

            if all(x in (None, '', 'None') for x in list(form_data[index].values())) or len(form_data[index]) < 1:
                del form_data[index]
                continue
            else:
                for x in form_data[index]:
                    form_data[index][str(x)] = str(form_data[index][str(x)]) if form_data[index][str(x)] not in (None, 'None') else ''
            # end if
        # end for

        aLineNumbers = []
        error_matrix = [['']*31 for _ in range(len(form_data))]
        for index in range(len(form_data)):

            # Check entered data formats
            # Product Number
            if '2' not in form_data[index] or form_data[index]['2'].strip('.').strip() in ('None', ''):
                error_matrix[index][2] += 'X - No Product Number provided.\n'
                form_data[index]['2'] = ''
            else:
                form_data[index]['2'] = form_data[index]['2'].upper()#.replace(' ','')
                if len(form_data[index]['2'].strip('. ')) > 18:
                    error_matrix[index][2] += 'X - Product Number exceeds 18 characters.\n'
                # end if
            # end if

            if '28' not in form_data or form_data[index]['28'] == 'None':
                form_data[index]['28'] = form_data[index]['2'].strip('.')
            if '1' in form_data[index] and form_data[index]['1'] not in ('None', ''):
                # if form_data[index]['2'].startswith(('.', '..')):
                #     pass # error_matrix[index][1] += 'X - Line number provided when product number indicates relationship.\n'
                if not re.match("^\d+(?:\.\d+){0,2}$|^$", form_data[index]['1'] or '') and form_data[index]['1'] not in ('None', None):
                    error_matrix[index][1] += 'X - Invalid character. Use 0-9 and "." only.\n'
                elif form_data[index]['1'] not in ('None', None):
                    this = form_data[index]['1']
                    if this.count('.') == 0 and this:
                        Parent = int(this)
                        Child = 0
                        Grandchild = 0
                    elif this.count('.') == 1:
                        Parent = int(this[:this.find('.')])
                        Child = int(this[this.rfind('.')+1:])
                        Grandchild = 0
                    elif form_data[index]['1'].count('.') == 2:
                        Parent = int(this[:this.find('.')])
                        Child = int(this[this.find('.')+1: this.rfind('.')])
                        Grandchild = int(this[this.rfind('.')+1:])
                    # end if
                else:
                    form_data[index]['1'] = ''
                # end if

                if '2' in form_data[index] and form_data[index]['2'] not in ('None', ''):
                    if form_data[index]['1'].count('.') == 2 and not form_data[index]['2'].startswith('..'):
                        form_data[index]['2'] = '..' + form_data[index]['2']
                    elif form_data[index]['1'].count('.') == 1 and not form_data[index]['2'].startswith('.'):
                        form_data[index]['2'] = '.' + form_data[index]['2']
                    # end if
                # end if
            else:
                if form_data[index]['2'].startswith('..'):
                    Grandchild += 1
                    form_data[index]['1'] = str(Parent) + "." + str(Child) + "." + str(Grandchild)
                elif form_data[index]['2'].startswith('.'):
                    Child += 1
                    form_data[index]['1'] = str(Parent) + "." + str(Child)
                else:
                    if Parent % 10 == 0:
                        Parent += 10
                    else:
                        Parent += 1
                    Child = 0
                    Grandchild = 0
                    form_data[index]['1'] = str(Parent)
                # end if
            # end if

            # Product Description
            if '3' not in form_data[index] or form_data[index]['3'].strip() in ('None', ''):
                form_data[index]['3'] = ''

            # Order Qty
            if '4' not in form_data[index] or not re.match("^\d+$", form_data[index]['4'] or ''):
                error_matrix[index][4] += 'X - Invalid Order Qty.\n'
                if '4' in form_data[index] and form_data[index]['4'] in ('None',''):
                    form_data[index]['4'] = ''
            # end if

            # Plant
            if '6' in form_data[index] and not re.match("^\d{4}$|^$", form_data[index]['6'] or ''):
                error_matrix[index][6] += 'X - Invalid Plant.\n'
            # end if

            # SLOC
            if '7' in form_data[index] and not re.match("^\d{4}$|^$", form_data[index]['7'] or ''):
                error_matrix[index][7] += 'X - Invalid SLOC.\n'
            # end if

            # Item Cat
            if '8' in form_data[index]:
                if not re.match("^Z[A-Z0-9]{3}$|^$", form_data[index]['8'] or '', re.IGNORECASE):
                    error_matrix[index][8] += 'X - Invalid Item Cat.\n'

                if form_data[index]['8']:
                    form_data[index]['8'] = form_data[index]['8'].upper()
            # end if

            # P-Code
            if '9' in form_data[index]:
                if not re.match("^\d{2,3}$|^\(\d{2,3}-\d{4}\).*$|^[A-Z]\d{2}$|^\([A-Z]\d{2}-\d{4}\).*$|^$", form_data[index]['9'] or '', re.IGNORECASE):
                    error_matrix[index][9] += 'X - Invalid P-Code.\n'

                if form_data[index]['9']:
                    form_data[index]['9'] = form_data[index]['9'].upper()
            # end if

            # HW/SW Ind
            if '10' in form_data[index]:
                if not re.match("^H(ARD)?W(ARE)?$|^S(OFT)?W(ARE)?$|^CS$|^$", form_data[index]['10'] or '', re.IGNORECASE):
                    error_matrix[index][10] += 'X - Invalid HW/SW Ind.\n'

                if form_data[index]['10']:
                    form_data[index]['10'] = form_data[index]['10'].upper()
            # end if

            # Unit Price
            if '17' in form_data[index]:
                if form_data[index]['17'] != 'None':
                    form_data[index]['17'] = form_data[index]['17'].replace('$', '').replace(',','')

                if not re.match("^(?:-)?\d+(?:\.\d+)?$|^$", form_data[index]['17'] or ''):
                    error_matrix[index][17] += 'X - Invalid Unit Price provided.\n'
            # end if

            # Condition Type & Amount Supplied Together
            NoCondition = '21' not in form_data[index] or form_data[index]['21'] in ('', 'None')
            NoAmount = '22' not in form_data[index] or form_data[index]['22'] in ('', 'None')
            if (NoCondition and not NoAmount) or (NoAmount and not NoCondition):
                if NoAmount and not NoCondition:
                    error_matrix[index][22] += 'X - Condition Type provided without Amount.\n'
                else:
                    error_matrix[index][21] += 'X - Amount provided without Condition Type.\n'
                # end if
            # end if

            # Amount
            if '22' in form_data[index]:
                if form_data[index]['22'] != 'None':
                    form_data[index]['22'] = form_data[index]['22'].replace('$', '').replace(',','')

                if not re.match("^(?:-)?\d+(?:\.\d+)?$|^$", form_data[index]['22'] or ''):
                    error_matrix[index][22] += 'X - Invalid Amount provided.\n'
            # end if

            # Traceability Req
            if '23' in form_data[index]:
                form_data[index]['23'] = form_data[index]['23'].strip() if form_data[index]['23'] else form_data[index]['23']
                if not re.match("^Y(?:es)?$|^N(?:o)?$|^$", form_data[index]['23'] or '', re.I):
                    error_matrix[index][23] += "X - Invalid Traceability Req.\n"

                if form_data[index]['23'] != 'None':
                    form_data[index]['23'] = form_data[index]['23'].upper()
            # end if

            # Customer Asset
            if '24' in form_data[index]:
                form_data[index]['24'] = form_data[index]['24'].strip() if form_data[index]['24'] else form_data[index]['24']
                if not re.match("^Y(?:es)?$|^N(?:o)?$|^$", form_data[index]['24'] or '', re.I):
                    error_matrix[index][24] += "X - Invalid Customer Asset.\n"

                if form_data[index]['24'] != 'None':
                    form_data[index]['24'] = form_data[index]['24'].upper()

                if '23' not in form_data[index] or form_data[index]['23'] in ('None', ''):
                    pass  # error_matrix[index][24] += "X - Customer assest set without Traceability Req.\n"
                if '23' in form_data[index] and form_data[index]['23'] in ('N', 'NO') and form_data[index]['24'] in ('Y', 'YES'):
                    pass  # error_matrix[index][24] += "X - Customer asset cannot be 'Y' when Traceability is 'N'.\n"
            # end if

            # Customer Asset tagging Req
            if '25' in form_data[index]:
                form_data[index]['25'] = form_data[index]['25'].strip() if form_data[index]['25'] else form_data[index]['25']
                if not re.match("^Y(?:es)?$|^N(?:o)?$|^$", form_data[index]['25'] or '', re.I):
                    error_matrix[index][25] += "X - Invalid Customer Asset tagging Req.\n"

                if form_data[index]['25']:
                    form_data[index]['25'] = form_data[index]['25'].upper()

                if '24' not in form_data[index] or form_data[index]['24'] in ('None', ''):
                    pass  # error_matrix[index][25] += 'X - Cannot provide Customer Asset Tagging without Customer Asset.\n'
                if '24' in form_data[index] and form_data[index]['24'] in ('N', 'NO') and form_data[index]['25'] in ('Y', 'YES'):
                    error_matrix[index][25] += 'X - Cannot mark Customer Asset Tagging when part is not Customer Asset.\n'
            # end if

            # Collect data from database(s)
            oCursor = connections['BCAMDB'].cursor()

            # Populate Read-only fields
            P_Code = form_data[index]['9'] if '9' in form_data[index] and re.match("^\d{2,3}$", form_data[index]['9'], re.IGNORECASE) else None
            tPCode = None
            oCursor.execute('SELECT DISTINCT [Material Description],[MU-Flag],[X-Plant Status],[Base Unit of Measure],'+
                            '[P Code] FROM dbo.BI_MM_ALL_DATA WHERE [Material]=%s',[form_data[index]['2'].strip('.') if form_data[index]['2'] else None])
            tPartData = oCursor.fetchall()
            if tPartData:
                if '3' not in form_data[index] or form_data[index]['3'] in ('None', None, '') and Header.objects.get(pk=oRequest.session['existing']).configuration_status == 'In Process':
                    form_data[index]['3'] = tPartData[0][0] if tPartData[0][0] not in (None, 'NONE', 'None') else ''
                if '14' not in form_data[index] or form_data[index]['14'] in ('None', None, '') and Header.objects.get(pk=oRequest.session['existing']).configuration_status == 'In Process':
                    form_data[index]['14'] = tPartData[0][1] if tPartData[0][1] not in (None, 'NONE', 'None') else ''
                if '15' not in form_data[index] or form_data[index]['15'] in ('None', None, '') and Header.objects.get(pk=oRequest.session['existing']).configuration_status == 'In Process':
                    form_data[index]['15'] = tPartData[0][2] if tPartData[0][2] not in (None, 'NONE', 'None') else ''
                if '5' not in form_data[index] or form_data[index]['5'] in ('None', None, '') and Header.objects.get(pk=oRequest.session['existing']).configuration_status == 'In Process':
                    form_data[index]['5'] = tPartData[0][3] if tPartData[0][3] not in (None, 'NONE', 'None') else ''

                P_Code = form_data[index]['9'] if '9' in form_data[index] and re.match("^\d{2,3}$", form_data[index]['9'], re.IGNORECASE) else tPartData[0][4]
            else:
                error_matrix[index][2] += '! - Product Number not found.'
                # if Header.objects.get(pk=oRequest.session['existing']).pick_list or index != 0:
                #     # form_data[index]['5'] = ''
                #     # form_data[index]['9'] = ''
                #     # form_data[index]['10'] = ''
                #     # form_data[index]['14'] = ''
                #     # form_data[index]['15'] = ''
                #     error_matrix[index][2] += '! - Product Number not found.'
                # else:
                #     form_data[index]['5'] = 'PC'
                # # end if
            # end def

            if P_Code:
                oCursor.execute('SELECT [PCODE],[FireCODE],[Description],[Commodity] FROM dbo.REF_PCODE_FCODE WHERE [PCODE]=%s', [P_Code])
                tPCode = oCursor.fetchall()
            # end if

            if tPCode:
                # form_data[index]['9'] = '(' + tPCode[0][0] + " - " + tPCode[0][1] + "), " + tPCode[0][2]
                if '9' not in form_data[index] or form_data[index]['9'] in (None, 'None', '') or re.match("^\d{2,3}$", form_data[index]['9'], re.IGNORECASE):
                    form_data[index]['9'] = tPCode[0][2] if tPCode[0][2] else ''
                if '10' not in form_data[index] or form_data[index]['10'] == 'None':
                    form_data[index]['10'] = tPCode[0][3] if tPCode[0][3] else ''
                # end if
            # end if

            if '6' in form_data[index] and form_data[index]['6'] not in ('None', ''):
                oCursor.execute('SELECT [Plant] FROM dbo.REF_PLANTS')
                tPlants = [element[0] for element in oCursor.fetchall()]
                if form_data[index]['6'] not in tPlants:
                    error_matrix[index][6] += "! - Plant not found in database.\n"
            # end if

            if '7' in form_data[index] and form_data[index]['7'] not in ('None', ''):
                oCursor.execute('SELECT DISTINCT [SLOC] FROM dbo.REF_PLANT_SLOC')
                tSLOC = [element[0] for element in oCursor.fetchall()]
                if form_data[index]['7'] not in tSLOC:
                    error_matrix[index][7] += "! - SLOC not found in database.\n"
            # end if

            if '6' in form_data[index] and form_data[index]['6'] not in ('None', '')\
                    and '7' in form_data[index] and form_data[index]['7'] not in ('None', ''):
                oCursor.execute('SELECT [Plant],[SLOC] FROM dbo.REF_PLANT_SLOC WHERE [Plant]=%s AND [SLOC]=%s', [form_data[index]['6'], form_data[index]['7']])
                tResults = oCursor.fetchall()
                if (form_data[index]['6'], form_data[index]['7']) not in tResults:
                    error_matrix[index][6] += '! - Plant/SLOC combination not found.\n'
                    error_matrix[index][7] += '! - Plant/SLOC combination not found.\n'
                # end if
            # end if

            # Validate entries

            # Accumulate running total
            if '4' in form_data[index] and form_data[index]['4'] not in ('None', ''):
                if '21' in form_data[index] and form_data[index]['21'] == 'ZPR1':
                    needs_zpru = True
                    zpru_total += (float(form_data[index]['22'] if '22' in form_data[index] and form_data[index]['22'] else 0) * float(form_data[index]['4']))
                elif '21' in form_data[index] and form_data[index]['21'] == 'ZPRU':
                    needs_zpru = True
                elif '17' in form_data[index] and form_data[index]['17'] not in ('None', None, ''):
                    oHead = Header.objects.get(pk=oRequest.session['existing'])
                    if index == 0 and not oHead.pick_list:
                        # base_total = oHead.configuration.override_net_value or oHead.configuration.net_value
                        base_total = float(form_data[index]['17']) if form_data[index]['17'] not in (None, '', 'None') else 0

                        if '21' in form_data[index] and form_data[index]['21'] == 'ZUST' and  '22' in form_data[index] and form_data[index]['22']:
                            base_total += float(form_data[index]['22'].replace('$','').replace(',',''))
                        # end if
                    # end if

                    if str(form_data[index]['17']).startswith('!'):
                        override_total += float(form_data[index]['17'][1:].replace('$','').replace(',','')) +\
                                          float(form_data[index]['22'].replace('$','').replace(',','')\
                                                    if '22' in form_data[index] and form_data[index]['22'] else 0)
                        error_matrix[index][17] = '! - CPM override in effect.\n'
                    else:
                        net_total += (float(form_data[index]['4']) * (float(form_data[index]['17'].replace('$','')
                                                                            .replace(',','')) if not (index == 0 and not oHead.pick_list) else 0)) +\
                                     float(form_data[index]['22'].replace('$','').replace(',','') if '22' in form_data[index] and form_data[index]['22'] else 0)
                    # end if
                # end if
            # end if

            aLineNumbers.append(form_data[index]['1'])

            # Update line status
            if any("X - " in error for error in error_matrix[index]):
                form_data[index]['0'] = '<span class="glyphicon glyphicon-remove-sign" style="color:#DD0000;"></span>'
                status = 'ERROR'
            elif any("! - " in error for error in error_matrix[index]):
                form_data[index]['0'] = '<span class="glyphicon glyphicon-exclamation-sign" style="color:#FF8800;"></span>'
                if status == 'GOOD':
                    status = 'WARNING'
                # end if
            else:
                form_data[index]['0'] = '<span class="glyphicon glyphicon-ok-sign" style="color:#00AA00;"></span>'
            # end if
        # end for

        # Check for duplicate line numbers
        dIndices = {}
        aDuplicates = []
        for index in range(len(form_data)):
            try:
                dIndices[form_data[index]['1']].append(index)
            except KeyError:
                dIndices[form_data[index]['1']] = [index]
            # end try
        # end for

        for aIndices in dIndices.values():
            if len(aIndices) > 1:
                aDuplicates.extend(aIndices[1:])
            # end if
        # end for

        for index in aDuplicates:
            form_data[index]['0'] = '<span class="glyphicon glyphicon-remove-sign" style="color:#DD0000;"></span>'
            error_matrix[index][1] = 'X - Duplicate line number.\n'
            status = 'ERROR'
        # end for

        if override_total and not base_total:
            form_data[0]['17'] = '!' + str(override_total)
        # end if

        dReturned = {'data': form_data, 'nettotal': round(base_total or override_total or net_total, 2), 'zprutotal': round(zpru_total, 2), 'errors': error_matrix,
                     'zpru': str(needs_zpru), 'status': status}
        return HttpResponse(json.dumps(dReturned))
    else:
        return redirect(reverse('bomconfig:index'))
    # end if
# end def


def Download(oRequest):
    download = oRequest.GET.get('download', None)
    if download:
        oHeader = download[5:-10]
    else:
        oHeader = oRequest.session.get('existing', None)
    # end if


    if oHeader:
        oHeader = Header.objects.get(pk=oHeader)
        sFileName = oHeader.configuration_designation + ".xlsx"

        oFile = openpyxl.load_workbook(
            str(os.path.join(os.getenv('PYTHONPATH', os.getcwd()).split(';')[0], 'BoMConfig/static/BoMConfig/PSM BoM Upload Template PA5 FORMULAS.xlsx'))
        )

        # Populate Header data
        oFile.active = oFile.sheetnames.index('1) BOM Config Header')
        oFile.active['B1'] = oHeader.person_responsible
        oFile.active['B2'] = oHeader.react_request
        oFile.active['B3'] = oHeader.bom_request_type.name
        oFile.active['B4'] = oHeader.customer_unit.name
        oFile.active['B5'] = oHeader.customer_name
        oFile.active['B6'] = oHeader.sales_office
        oFile.active['B7'] = oHeader.sales_group
        oFile.active['B8'] = oHeader.sold_to_party
        oFile.active['B9'] = oHeader.ship_to_party
        oFile.active['B10'] = oHeader.bill_to_party
        oFile.active['B11'] = oHeader.ericsson_contract
        oFile.active['B12'] = oHeader.payment_terms

        oFile.active['B14'] = oHeader.projected_cutover
        oFile.active['B15'] = oHeader.program.name if oHeader.program else None
        oFile.active['B16'] = oHeader.configuration_designation
        oFile.active['B17'] = oHeader.customer_designation
        oFile.active['B18'] = oHeader.technology.name if oHeader.technology else None
        oFile.active['B19'] = oHeader.product_area1.name if oHeader.product_area1 else None
        oFile.active['B20'] = oHeader.product_area2.name if oHeader.product_area2 else None
        oFile.active['B21'] = oHeader.radio_frequency.name if oHeader.radio_frequency else None
        oFile.active['B22'] = oHeader.radio_band.name if oHeader.radio_band else None
        oFile.active['B23'] = oHeader.optional_free_text1
        oFile.active['B24'] = oHeader.optional_free_text2
        oFile.active['B25'] = oHeader.optional_free_text3
        oFile.active['B26'] = oHeader.inquiry_site_template
        oFile.active['B27'] = oHeader.readiness_complete / 100 if oHeader.readiness_complete else oHeader.readiness_complete
        oFile.active['B28'] = ('X' if oHeader.complete_delivery else None)
        oFile.active['B29'] = ('X' if oHeader.no_zip_routing else None)
        oFile.active['B30'] = oHeader.valid_to_date
        oFile.active['B31'] = oHeader.valid_from_date
        oFile.active['B32'] = oHeader.shipping_condition

        oFile.active['B34'] = oHeader.baseline_impacted
        oFile.active['B35'] = oHeader.model
        oFile.active['B36'] = oHeader.model_description
        oFile.active['B37'] = oHeader.model_replaced
        oFile.active['B38'] = oHeader.initial_revision

        oFile.active['B40'] = oHeader.configuration_status

        oFile.active['B42'] = oHeader.workgroup
        oFile.active['B43'] = oHeader.name

        oFile.active['B45'] = ('X' if oHeader.pick_list else None)

        # Populate Config Entry tab
        oFile.active = oFile.sheetnames.index('2) BOM Config Entry')
        if oHeader.configuration:
            oFile.active['D5'] = ('X' if oHeader.configuration.reassign else None)
            oFile.active['D8'] = ('X' if oHeader.configuration.PSM_on_hold else None)
            oFile.active['N4'] = ('X' if oHeader.configuration.internal_external_linkage else None)
            oFile.active['Q4'] = ('$ ' + str(oHeader.configuration.zpru_total) if oHeader.configuration.zpru_total else '$ xx,xxx.xx')

            iRow = 13
            aConfigLines = ConfigLine.objects.filter(config=oHeader.configuration).order_by('line_number')
            aConfigLines = sorted(aConfigLines, key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')])
            for oConfigLine in aConfigLines:
                oFile.active['A' + str(iRow)] = oConfigLine.line_number
                oFile.active['B' + str(iRow)] = ('..' if oConfigLine.is_grandchild else '.' if oConfigLine.is_child else '') + oConfigLine.part.base.product_number
                oFile.active['C' + str(iRow)] = oConfigLine.part.product_description
                oFile.active['D' + str(iRow)] = oConfigLine.order_qty
                oFile.active['E' + str(iRow)] = oConfigLine.part.base.unit_of_measure
                oFile.active['F' + str(iRow)] = oConfigLine.plant
                oFile.active['G' + str(iRow)] = oConfigLine.sloc
                oFile.active['H' + str(iRow)] = oConfigLine.item_category
                oFile.active['I' + str(iRow)] = oConfigLine.pcode
                oFile.active['J' + str(iRow)] = oConfigLine.commodity_type
                oFile.active['K' + str(iRow)] = oConfigLine.package_type
                oFile.active['L' + str(iRow)] = oConfigLine.spud
                oFile.active['M' + str(iRow)] = oConfigLine.REcode
                oFile.active['N' + str(iRow)] = oConfigLine.mu_flag
                oFile.active['O' + str(iRow)] = oConfigLine.x_plant
                oFile.active['P' + str(iRow)] = oConfigLine.internal_notes
                oFile.active['Q' + str(iRow)] = oConfigLine.linepricing.unit_price
                oFile.active['R' + str(iRow)] = oConfigLine.higher_level_item
                oFile.active['S' + str(iRow)] = oConfigLine.material_group_5
                oFile.active['T' + str(iRow)] = oConfigLine.purchase_order_item_num
                oFile.active['U' + str(iRow)] = oConfigLine.condition_type
                oFile.active['V' + str(iRow)] = oConfigLine.amount
                oFile.active['W' + str(iRow)] = oConfigLine.traceability_req
                oFile.active['X' + str(iRow)] = oConfigLine.customer_asset
                oFile.active['Y' + str(iRow)] = oConfigLine.customer_asset_tagging
                oFile.active['Z' + str(iRow)] = oConfigLine.customer_number
                oFile.active['AA' + str(iRow)] = oConfigLine.sec_customer_number
                oFile.active['AB' + str(iRow)] = oConfigLine.vendor_article_number
                oFile.active['AC' + str(iRow)] = oConfigLine.comments
                oFile.active['AD' + str(iRow)] = oConfigLine.additional_ref
                iRow += 1
            # end for
        # end if

        # Populate Config ToC tab
        oFile.active = oFile.sheetnames.index('2a) BoM Config ToC')
        if oHeader.configuration:
            oFile.active['D5'] = ('X' if oHeader.configuration.reassign else None)
            oFile.active['D8'] = ('X' if oHeader.configuration.PSM_on_hold else None)
            oFile.active['A13'] = oHeader.configuration_designation
            oFile.active['B13'] = oHeader.customer_designation or None
            oFile.active['C13'] = oHeader.technology.name if oHeader.technology else None
            oFile.active['D13'] = oHeader.product_area1.name if oHeader.product_area1 else None
            oFile.active['E13'] = oHeader.product_area2.name if oHeader.product_area2 else None
            oFile.active['F13'] = oHeader.model or None
            oFile.active['G13'] = oHeader.model_description or None
            oFile.active['H13'] = oHeader.model_replaced or None
            oFile.active['J13'] = oHeader.bom_request_type.name
            oFile.active['K13'] = oHeader.configuration_status or None
            oFile.active['L13'] = oHeader.inquiry_site_template if str(oHeader.inquiry_site_template).startswith('1') else None
            oFile.active['M13'] = oHeader.inquiry_site_template if str(oHeader.inquiry_site_template).startswith('4') else None
            oFile.active['N13'] = oHeader.internal_notes
            oFile.active['O13'] = oHeader.external_notes
            oFile.active['I13'] = '=HYPERLINK("' + oRequest.build_absolute_uri(reverse('bomconfig:search')) +\
                                  '?config=' + searchscramble(oHeader.pk) + '","' +\
                                  oRequest.build_absolute_uri(reverse('bomconfig:search')) + '?config=' +\
                                  searchscramble(oHeader.pk) + '")'
        # end if

        # Populate Config Revision tab
        oFile.active = oFile.sheetnames.index('2b) BoM Config Revision')
        if oHeader.configuration:
            iRow = 13
            aRevData = BuildDataArray(oHeader, revision=True)
            for oRev in aRevData:
                oFile.active['A' + str(iRow)] = oRev['0']
                oFile.active['B' + str(iRow)] = oRev['1']
                oFile.active['C' + str(iRow)] = oRev['2']
                oFile.active['D' + str(iRow)] = oRev['3']
                oFile.active['E' + str(iRow)] = oRev['4']
                oFile.active['F' + str(iRow)] = oRev['5']
                oFile.active['G' + str(iRow)] = oRev['6']
            # end for
        # end if

        # Populate SAP inquiry tab
        oFile.active = oFile.sheetnames.index('3a) SAP Inquiry Creation')
        if oHeader.configuration:
            oFile.active['I4'] = ('X' if oHeader.configuration.internal_external_linkage else None)
            oFile.active['M4'] = ('$ ' + str(oHeader.configuration.zpru_total) if oHeader.configuration.zpru_total else '$ xx,xxx.xx')

            iRow = 13
            aConfigLines = ConfigLine.objects.filter(config=oHeader.configuration).order_by('line_number')
            aConfigLines = sorted(aConfigLines, key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')])
            for oConfigLine in aConfigLines:
                if '.' in oConfigLine.line_number:
                    continue
                oFile.active['A' + str(iRow)] = oConfigLine.line_number
                oFile.active['B' + str(iRow)] = oConfigLine.part.base.product_number
                oFile.active['C' + str(iRow)] = oConfigLine.part.product_description
                oFile.active['D' + str(iRow)] = oConfigLine.order_qty
                oFile.active['E' + str(iRow)] = oConfigLine.plant
                oFile.active['F' + str(iRow)] = oConfigLine.sloc
                oFile.active['G' + str(iRow)] = oConfigLine.item_category
                oFile.active['H' + str(iRow)] = oConfigLine.linepricing.unit_price
                oFile.active['J' + str(iRow)] = oConfigLine.material_group_5
                oFile.active['K' + str(iRow)] = oConfigLine.purchase_order_item_num
                oFile.active['L' + str(iRow)] = oConfigLine.condition_type
                oFile.active['M' + str(iRow)] = oConfigLine.amount
                iRow += 1
            # end for
        # end if

        # Populate SAP Site Template tab
        oFile.active = oFile.sheetnames.index('3b) SAP ST Creation')
        if oHeader.configuration:
            iRow = 13
            aConfigLines = ConfigLine.objects.filter(config=oHeader.configuration).order_by('line_number')
            aConfigLines = sorted(aConfigLines, key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')])
            for oConfigLine in aConfigLines:
                if '.' in oConfigLine.line_number:
                    continue

                oFile.active['A' + str(iRow)] = oConfigLine.line_number
                oFile.active['B' + str(iRow)] = oConfigLine.part.base.product_number
                oFile.active['C' + str(iRow)] = oConfigLine.part.product_description
                oFile.active['D' + str(iRow)] = oConfigLine.order_qty
                oFile.active['E' + str(iRow)] = oConfigLine.plant
                oFile.active['F' + str(iRow)] = oConfigLine.sloc
                oFile.active['G' + str(iRow)] = oConfigLine.item_category
                oFile.active['H' + str(iRow)] = oConfigLine.higher_level_item
                iRow += 1
            # end for
        # end if

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(sFileName)
        oFile.save(response)

        return response
    else:
        return HttpResponse()
# end def


def ReactSearch(oRequest):
    if oRequest.method != 'POST' or not oRequest.POST:
        return Http404()
    # end if
    dReturnData = {}
    sPSMReq = oRequest.POST['psm_req']
    sQuery = "SELECT [req_id],[assigned_to],[customer_name],[sales_o],[sales_g],[sold_to_code],[ship_to_code],[bill_to_code],"+\
             "[pay_terms],[workgroup],[cu],[mus_cnt_num] FROM dbo.ps_requests WHERE [req_id] = %s AND [Modified_by] IS NULL"
    oCursor = connections['REACT'].cursor()
    oCursor.execute(sQuery, [sPSMReq])
    oResults = oCursor.fetchall()
    if oResults:
        dReturnData.update({
            'req_id': oResults[0][0],
            'person_resp': oResults[0][1][:oResults[0][1].find('@')].replace('.', ' '),
            'cust_name': oResults[0][2],
            'sales_office': oResults[0][3],
            'sales_group': oResults[0][4],
            'sold_to': oResults[0][5],
            'ship_to': oResults[0][6],
            'bill_to': oResults[0][7],
            'terms': oResults[0][8],
            'workgroup': oResults[0][9],
            'cust': oResults[0][10],
            'contract': oResults[0][11]
        })
    # end if

    return JsonResponse(data=dReturnData)
# end def


def BaselineMgmt(oRequest):
    if 'existing' in oRequest.session:
        try:
            Unlock(oRequest, oRequest.session['existing'])
        except:
            pass
        # end try

        del oRequest.session['existing']
    # end if

    if 'status' in oRequest.session:
        del oRequest.session['status']
    # end if

    oTable = None
    bDownloadable = False
    dTableData = None
    aTable = []

    if oRequest.method == 'POST' and oRequest.POST:
        form = SubmitForm(oRequest.POST)
        if form.is_valid():
            oBaseline = Baseline.objects.get(title__iexact=form.cleaned_data['baseline_title'])

            # First get list of revisions going back to 'A'
            aRevisions = sorted(list(set([oBaseRev.version for oBaseRev in oBaseline.baseline_revision_set.order_by('version')])),
                                key=cmp_to_key(lambda x,y:(-1 if len(x.strip('1234567890')) < len(y.strip('1234567890'))
                                                                 or list(x.strip('1234567890')) < (['']*(len(x.strip('1234567890'))-len(y.strip('1234567890'))) +
                                                                                                   list(y.strip('1234567890')))
                                                                 or (x.strip('1234567890') == y.strip('1234567890') and list(x) < list(y)) else 0 if x == y else 1)),
                                reverse=True)

            dTableData = {'baseline': oBaseline.title, 'revisions':[]}

            for sRev in aRevisions:
                aHeads = Header.objects.filter(baseline__baseline=oBaseline).filter(baseline_version=sRev).order_by('configuration_status')
                if aHeads:
                    dTableData['revisions'].append({'revision': sRev, 'configs': aHeads})
                    if not bDownloadable:
                        bDownloadable = True
                # end if
            # end for
        # end if

        if dTableData:
            aTable = [dTableData]
        # end if
    else:
        form = SubmitForm()
        aBaselines = Baseline.objects.all()
        for oBaseline in aBaselines:
            aRevisions = sorted(list(set([oBaseRev.version for oBaseRev in oBaseline.baseline_revision_set.order_by('version') if oBaseRev.version == oBaseRev.baseline.current_active_version])),
                                key=cmp_to_key(lambda x,y:(-1 if len(x.strip('1234567890')) < len(y.strip('1234567890'))
                                                                 or list(x.strip('1234567890')) < (['']*(len(x.strip('1234567890'))-len(y.strip('1234567890'))) +
                                                                                                   list(y.strip('1234567890')))
                                                                 or (x.strip('1234567890') == y.strip('1234567890') and list(x) < list(y)) else 0 if x == y else 1)),
                                reverse=True)

            dTableData = {'baseline': oBaseline.title, 'revisions':[]}

            for sRev in aRevisions[:1]:
                aHeads = Header.objects.filter(baseline__baseline=oBaseline).filter(baseline_version=sRev).order_by('configuration_status')
                if aHeads:
                    dTableData['revisions'].append({'revision': sRev, 'configs': aHeads})
                # end if
            # end for

            aTable.append(dTableData)
        # end for

        aTable.append(
            {
                'baseline':'No Associated Baseline',
                'revisions':[
                    {
                        'revision':'',
                        'configs': Header.objects.filter(baseline__isnull=True).order_by('configuration_designation')
                    }
                ]
            }
        )
    # end if

    aTitles = ['','','Configuration','Status','Program','Customer Number','Model','Model Description','Customer Price',
               'Inquiry/Site Template Number','Model Replacing','Comments','ZUST','P-Code','Plant']

    return Default(oRequest, sTemplate='BoMConfig/baselinemgmt.html', dContext={'form': form, 'tables': aTable, 'downloadable': bDownloadable, 'column_titles': aTitles})
# end def


def DownloadBaseline(oRequest):
    aColumnTitles = ['Line #', 'Product Number', 'Product Description', 'Order Qty', 'UoM', 'HW/SW Ind', 'Net Price 2015',
                     '', 'Traceability Req. (Serialization)', 'Customer Asset?', 'Customer Asset Tagging Requirement',
                     'Customer number', 'Second Customer Number', 'Vendor Article Number','Comments',
                     'Additional Reference\n(if required)']

    if oRequest.method != 'POST' or not oRequest.POST:
        return Http404()
    # end if

    oBaseline = oRequest.POST['baseline']
    sCookie = oRequest.POST['file-cookie']

    if oBaseline:
        oBaseline = Baseline.objects.get(title__iexact=oBaseline)
        sFileName = str(oBaseline.latest_revision) + ".xlsx"
        oFile = openpyxl.Workbook()

        # Fill Revision tab
        oSheet = oFile.active
        oSheet.title = 'Revisions'

        oSheet['A1'] = 'Revision'
        oSheet['C1'] = 'Date'

        GenerateRevisionSummary(oBaseline,oBaseline.current_active_version, oBaseline.current_inprocess_version)
        aRevisions = [oRev.revision for oRev in oBaseline.revisionhistory_set.all()]
        aRevisions = sorted(list(set(aRevisions)), key=cmp_to_key(lambda x,y:(-1 if len(x.strip('1234567890')) < len(y.strip('1234567890'))
                                                                                    or list(x.strip('1234567890')) < (['']*(len(x.strip('1234567890'))-len(y.strip('1234567890'))) +
                                                                                                   list(y.strip('1234567890')))
                                                                                    or (x.strip('1234567890') == y.strip('1234567890') and list(x) < list(y)) else 0 if x == y else 1)))

        iRow = 2
        for sRev in aRevisions:
            sHistory = oBaseline.revisionhistory_set.get(revision=sRev).history
            if not sHistory or not sHistory.strip():
                continue
            oSheet['A' + str(iRow)] = 'Revision: ' + sRev + '\n' + oBaseline.revisionhistory_set.get(revision=sRev).history
            if Baseline_Revision.objects.get(baseline=oBaseline, version=sRev).completed_date:
                oSheet['C' + str(iRow)] = Baseline_Revision.objects.get(baseline=oBaseline, version=sRev).completed_date.strftime('%m/%d/%Y')

            iRow += 2
        # end for

        aHeaders = oBaseline.latest_revision.header_set.exclude(configuration_status__in=('Discontinued', 'Obsolete')).order_by('pick_list', 'configuration_status','configuration_designation')
        # aHeaders = oBaseline.latest_revision.header_set.all().order_by('pick_list', 'configuration_status','configuration_designation')
        aHeaders = list(chain(Baseline_Revision.objects.get(
            baseline=oBaseline, version=oBaseline.current_inprocess_version
        ).header_set.all().order_by('pick_list', 'configuration_status','configuration_designation'), aHeaders))

        for oHeader in aHeaders:
            if oHeader.configuration_status in ('In Process', 'In Process/Pending') and 'In Process' not in oFile.get_sheet_names():
                oFile.create_sheet(title="In Process")
            elif oHeader.configuration_status == 'Active' and 'Active' not in oFile.get_sheet_names():
                oFile.create_sheet(title="Active")
            elif oHeader.configuration_status == 'Discontinued' and 'Discontinued' not in oFile.get_sheet_names():
                oFile.create_sheet(title="Discontinued")
            elif oHeader.configuration_status == 'Obsolete':
                continue

            iCurrentRow = 2
            sTitle = str(oHeader.configuration_designation) + (str("__New") if oHeader.configuration_status in ('In Process', 'In Process/Pending') else '')
            if len(sTitle) > 31:
                sTitle = sTitle.replace('Optional', 'Opt').replace('Hardware','HW').replace('Pick List', 'PL').replace(' - ','-')
            oSheet = oFile.create_sheet(title=sTitle)
            # Build Header row
            for iIndex in range(len(aColumnTitles)):
                oSheet[str(utils.get_column_letter(iIndex + 1)) + '1'] = aColumnTitles[iIndex]
            # end for

            # Add line items (ordered by line number)
            aLineItems = oHeader.configuration.configline_set.all().order_by('line_number')
            aLineItems = sorted(aLineItems, key=lambda x: [int(y) for y in getattr(x, 'line_number').split('.')])
            for oLineItem in aLineItems:
                oSheet['A' + str(iCurrentRow)] = oLineItem.line_number
                oSheet['B' + str(iCurrentRow)] = ('..' if oLineItem.is_grandchild else '.' if oLineItem.is_child else '') + oLineItem.part.base.product_number
                oSheet['C' + str(iCurrentRow)] = oLineItem.part.product_description
                oSheet['D' + str(iCurrentRow)] = oLineItem.order_qty
                oSheet['E' + str(iCurrentRow)] = oLineItem.part.base.unit_of_measure
                oSheet['F' + str(iCurrentRow)] = oLineItem.commodity_type
                oSheet['G' + str(iCurrentRow)] = oLineItem.linepricing.unit_price if oHeader.pick_list else ''
                oSheet['H' + str(iCurrentRow)] = oLineItem.amount if oLineItem.condition_type == 'ZUST' else ''
                oSheet['I' + str(iCurrentRow)] = oLineItem.traceability_req
                oSheet['J' + str(iCurrentRow)] = oLineItem.customer_asset
                oSheet['K' + str(iCurrentRow)] = oLineItem.customer_asset_tagging
                oSheet['L' + str(iCurrentRow)] = oLineItem.customer_number
                oSheet['M' + str(iCurrentRow)] = oLineItem.sec_customer_number
                oSheet['N' + str(iCurrentRow)] = oLineItem.vendor_article_number
                oSheet['O' + str(iCurrentRow)] = oLineItem.comments
                oSheet['P' + str(iCurrentRow)] = oLineItem.additional_ref
                iCurrentRow += 1
            # end for
            if not oHeader.pick_list:
                oSheet['G2'] = oHeader.configuration.get_first_line().linepricing.override_price or oHeader.configuration.get_first_line().linepricing.unit_price or ''
            # end if

            if oHeader.customer_unit.name == 'AT&T':
                oSheet['H1'] = 'Transportation Fee'
                oTransport = oHeader.configuration.configline_set.filter(condition_type__iexact='ZUST')
                if oTransport:
                    oSheet['H2'] = oTransport[0].amount
                # end if
        # end for

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment;filename="{0}"'.format(sFileName)
        response.set_cookie('fileMark', sCookie, max_age=60)
        oFile.save(response)

        return response
    else:
        return HttpResponse()
# end def


def Pricing(oRequest, mgmt):

    bCanReadPricing = bool(SecurityPermission.objects.filter(title='Detailed_Price_Read').filter(user__in=oRequest.user.groups.all()))
    bCanWritePricing = bool(SecurityPermission.objects.filter(title='Detailed_Price_Write').filter(user__in=oRequest.user.groups.all()))

    if 'existing' in oRequest.session:
        try:
            Unlock(oRequest, oRequest.session['existing'])
        except:
            pass
        # end try

        del oRequest.session['existing']
    # end if

    if 'status' in oRequest.session:
        del oRequest.session['status']
    # end if

    status_message = None
    sConfig = oRequest.POST['config'] if 'config' in oRequest.POST else None

    if mgmt:
        sTemplate = 'BoMConfig/subpricing.html'
        aConfigLines = []
        dContext = {'configlines':aConfigLines}
        if oRequest.method == 'POST' and oRequest.POST:
            if oRequest.POST['action'] == 'save':
                if oRequest.POST['config'] == oRequest.POST['initial']:
                    for dLine in json.loads(oRequest.POST['data_form']):
                        oLineToEdit = ConfigLine.objects.filter(config__header__configuration_status='Active')\
                            .filter(config__header__configuration_designation__iexact=sConfig)\
                            .filter(line_number=dLine['0'])[0]
                        oLinePrice = LinePricing.objects.get(config_line=oLineToEdit)
                        oLinePrice.unit_price = dLine['5'] or None
                        oLinePrice.override_price = dLine['7'] or None
                        oLinePrice.save()
                    # end for

                    oConfig = Configuration.objects.get(header__configuration_designation__iexact=sConfig)
                    oConfig.override_net_value = float(oRequest.POST['net_value'])
                    oConfig.save()

                    status_message = 'Data saved.'
                else:
                    status_message = 'Cannot change configuration during save.'
                    sConfig = oRequest.POST['initial']
                # end if
            # end if

            aLine = ConfigLine.objects.filter(config__header__configuration_designation__iexact=sConfig)
            aLine = sorted(aLine, key=lambda x: ([int(y) for y in x.line_number.split('.')]))

            aConfigLines = [{
                '0': oLine.line_number,
                '1': ('..' if oLine.is_grandchild else '.' if oLine.is_child else '') + oLine.part.base.product_number,
                '2': str(oLine.part.base.product_number) + str('_'+ oLine.spud if oLine.spud else ''),
                '3': oLine.part.product_description,
                '4': oLine.order_qty if oLine.order_qty else 0,
                '5': oLine.linepricing.unit_price or 0,
                '6': str(float(oLine.order_qty or 0) * float(oLine.linepricing.unit_price or 0)),
                '7': oLine.linepricing.override_price or '',
                '8': oLine.higher_level_item or '',
                '9': oLine.material_group_5 or '',
                '10': oLine.commodity_type or '',
                '11': oLine.comments or '',
                '12': oLine.additional_ref or ''
                            } for oLine in aLine]
            dContext['configlines'] = aConfigLines
            dContext.update({'config': sConfig, 'is_not_pick_list': not aLine[0].config.header.pick_list if aLine else False})
    else:
        global aHeaderList
        sTemplate='BoMConfig/pricing.html'

        aHeaderList = Header.objects.all().order_by('baseline','pick_list')
        # aLine = ConfigLine.objects.filter(config__header__configuration_status='Active')
        if aHeaderList:
            aLine = aHeaderList[0].configuration.configline_set.all()
            aLine = sorted(aLine, key=lambda x: ([int(y) for y in x.line_number.split('.')]))
            aLine = sorted(aLine, key=lambda x: (x.config.header.baseline.title if x.config.header.baseline else '',
                                                 x.config.header.configuration_designation))
        else:
            aLine = []
        # end if

        aConfigLines = [{
            '0': oLine.config.header.baseline.title if oLine.config.header.baseline else oLine.config.header.configuration_designation,
            '1': oLine.line_number,
            '2': oLine.config.header.configuration_designation,
            '3': ('..' if oLine.is_grandchild else '.' if oLine.is_child else '') + oLine.part.base.product_number,
            '4': oLine.part.base.product_number,
            '5': oLine.order_qty if oLine.order_qty else 0,
            '6': oLine.part.product_description,
            '7': oLine.linepricing.unit_price or 0 if oLine.config.header.pick_list or str(oLine.line_number) != '10' else '',
            '8': str(float(oLine.order_qty or 0) * float(oLine.linepricing.unit_price or 0)),
            '9': oLine.linepricing.override_price or ''
                        } for oLine in aLine]

        dContext = {
            'configlines': aConfigLines
        }
    # end if

    dContext.update({
        'status_message': status_message,
        'pricing_read_authorized': bCanReadPricing,
        'pricing_write_authorized': bCanWritePricing,
    })

    return Default(oRequest, sTemplate, dContext)
# end def


def GetConfigLines(oRequest):
    if oRequest.method == 'GET':
        raise Http404('Access Denied')
    else:
        global aHeaderList
        iMultiplier = int(oRequest.POST['multiplier'])

        if iMultiplier < len(aHeaderList):
            aLine = aHeaderList[iMultiplier].configuration.configline_set.all()
        else:
            aLine = []

        aLine = sorted(aLine, key=lambda x: [int(y) for y in x.line_number.split('.')])
        aLine = sorted(aLine, key=lambda x: (x.config.header.baseline.title if x.config.header.baseline else '',
                                             x.config.header.configuration_designation))

        aConfigLines = [{
            '0': oLine.config.header.baseline.title if oLine.config.header.baseline else oLine.config.header.configuration_designation,
            '1': oLine.line_number,
            '2': oLine.config.header.configuration_designation,
            '3': ('..' if oLine.is_grandchild else '.' if oLine.is_child else '') + oLine.part.base.product_number,
            '4': oLine.part.base.product_number,
            '5': oLine.order_qty if oLine.order_qty else 0,
            '6': oLine.part.product_description,
            '7': oLine.linepricing.unit_price or 0 if oLine.config.header.pick_list or str(oLine.line_number) != '10' else '',
            '8': str(float(oLine.order_qty or 0) * float(oLine.linepricing.unit_price or 0)),
            '9': oLine.linepricing.override_price or ''
                        } for oLine in aLine]

        return JsonResponse(aConfigLines, safe=False)
    # end if
#end def


def Approval(oRequest):
    if 'existing' in oRequest.session:
        try:
            Unlock(oRequest, oRequest.session['existing'])
        except:
            pass
        # end try

        del oRequest.session['existing']
    # end if
    SecurityPermission.objects.filter(title__iregex='^.*Approval.*$').filter(user__in=oRequest.user.groups.all())
    dContext = {
        'approval_wait': Header.objects.filter(configuration_status='In Process/Pending'),
        'customer_list': ['All'] + [obj.name for obj in REF_CUSTOMER.objects.all()],
        'approval_seq': HeaderTimeTracker.approvals(),
        'deaddate': timezone.datetime(1900,1,1),
        'namelist':['PSM Configuration Mgr.', 'SCM #1', 'SCM #2', 'CSR','Comm. Price Mgmt.','ACR','PSM Baseline Mgmt.',
                    'Customer #1','Customer #2','Customer Warehouse','Ericsson VAR','Baseline Release & Dist.'],
        'viewauthorized': SecurityPermission.objects.filter(title__iregex='^.*Approval.*$').filter(user__in=oRequest.user.groups.all())
    }
    return Default(oRequest, sTemplate='BoMConfig/approvals.html', dContext=dContext)
# end def


def Action(oRequest):
    if 'existing' in oRequest.session:
        try:
            Unlock(oRequest, oRequest.session['existing'])
        except:
            pass
        # end try

        del oRequest.session['existing']
    # end if

    dContext = {
        'in_process': Header.objects.filter(configuration_status='In Process'),
        'active': Header.objects.filter(configuration_status='Active'),
        'on_hold': Header.objects.filter(configuration_status='On Hold'),
        'customer_list': ['All'] + [obj.name for obj in REF_CUSTOMER.objects.all()],
        'viewauthorized': bool(oRequest.user.groups.filter(name__in=['BOM_BPMA_Architect','BOM_PSM_Product_Supply_Manager', 'BOM_PSM_Baseline_Manager']))
    }
    return Default(oRequest, sTemplate='BoMConfig/actions.html', dContext=dContext)
# end def


def ListFill(oRequest):
    if oRequest.method == 'POST' and oRequest.POST:
        iParentID = int(oRequest.POST['id'])
        cParentClass = Header._meta.get_field(oRequest.POST['parent']).rel.to
        oParent = cParentClass.objects.get(pk=iParentID)
        if oRequest.POST['child'] != 'baseline_impacted':
            cChildClass = Header._meta.get_field(oRequest.POST['child']).rel.to
            result = {obj.id:obj.name for obj in cChildClass.objects.filter(parent=oParent)}
        else:
            cChildClass = Baseline
            result = {obj.title:obj.title for obj in cChildClass.objects.filter(customer=oParent)}

        return JsonResponse(result)
    else:
        raise Http404
    # end if
#end def


def ListREACTFill(oRequest):
    from itertools import chain
    if oRequest.method == 'POST' and oRequest.POST:
        iParentID = int(oRequest.POST['id'])
        cParentClass = Header._meta.get_field(oRequest.POST['parent']).rel.to

        oParent = cParentClass.objects.get(pk=iParentID)

        oCursor = connections['REACT'].cursor()
        oCursor.execute('SELECT DISTINCT [Customer] FROM ps_fas_contracts WHERE [CustomerUnit]=%s',[oParent.name])
        tResults = oCursor.fetchall()
        result = {obj:obj for obj in chain.from_iterable(tResults)}
        return JsonResponse(result)
    else:
        raise Http404
    # end if
#end def


def AjaxApprove(oRequest):
    if oRequest.method == 'POST' and oRequest.POST:
        if oRequest.POST.get('action', None) not in ('approve','disapprove','skip','clone','delete','send_to_approve', 'hold', 'unhold', 'cancel'):
            raise Http404
        sAction = oRequest.POST.get('action')
        aRecords = [int(record) for record in json.loads(oRequest.POST.get('data'))]
        if 'comments' in oRequest.POST:
            aComments = [record for record in json.loads(oRequest.POST.get('comments', None))]
        if 'destinations' in oRequest.POST:
            aDestinations = [record for record in json.loads(oRequest.POST.get('destinations', None))]

        if sAction == 'send_to_approve':
            for iRecord in aRecords:
                oHeader = Header.objects.get(pk=iRecord)
                oCreated = oHeader.headertimetracker_set.first().created_on
                if oHeader.headertimetracker_set.filter(submitted_for_approval=None):
                    oLatestTracker = oHeader.headertimetracker_set.filter(submitted_for_approval=None)[0]
                    oLatestTracker.submitted_for_approval = timezone.now()
                    oLatestTracker.psm_config_approver = oRequest.user.username
                    oLatestTracker.created_on = oCreated
                    oLatestTracker.save()
                else:
                    HeaderTimeTracker.objects.create(**{
                        'header': oHeader,
                        'submitted_for_approval': timezone.now(),
                        'psm_config_approver': oRequest.user.username,
                        'created_on': oCreated
                    })

                oHeader.configuration_status = 'In Process/Pending'
                oHeader.save()
        elif sAction in ('approve', 'disapprove', 'skip'):
            aChain = HeaderTimeTracker.approvals()
            aBaselinesCompleted = []
            for index in range(len(aRecords)):
                oHeader = Header.objects.get(pk=aRecords[index])
                oLatestTracker = oHeader.headertimetracker_set.order_by('-submitted_for_approval')[0]

                sNeededLevel = None
                # Find earliest empty approval slot, and see if user has approval permission for that slot
                for level in aChain:
                    if getattr(oLatestTracker, level+'_denied_approval', None) is not None:
                        break
                    # end if

                    if getattr(oLatestTracker, level+'_approver','no attr') in (None, ''):
                        sNeededLevel = level
                        break
                    # end if
                # end for

                aNames = HeaderTimeTracker.permission_entry(sNeededLevel)
                bCanApprove = bool(SecurityPermission.objects.filter(title__in=aNames).filter(user__in=oRequest.user.groups.all()))
                # end if

                # If so, approve, disapprove, or skip as requested
                if bCanApprove:
                    if sAction == 'approve':
                        setattr(oLatestTracker, sNeededLevel+'_approver', oRequest.user.username)
                        setattr(oLatestTracker, sNeededLevel+'_approved_on', timezone.now())
                        setattr(oLatestTracker, sNeededLevel+'_comments', aComments[index])
                    elif sAction == 'disapprove':
                        setattr(oLatestTracker, sNeededLevel+'_approver', oRequest.user.username)
                        setattr(oLatestTracker, sNeededLevel+'_denied_approval', timezone.now())
                        setattr(oLatestTracker, sNeededLevel+'_comments', aComments[index])
                        oLatestTracker.disapproved_on = timezone.now()

                        # Create a new time tracker, if not disapproved back to PSM
                        if aDestinations[index] != 'psm_config':
                            oNewTracker = HeaderTimeTracker.objects.create(**{'header': oHeader,
                                                                              'created_on': oLatestTracker.created_on,
                                                                              'submitted_for_approval': timezone.now()
                                                                              })
                            # Copy over approval data for each level before destination level
                            for level in aChain:
                                if level == aDestinations[index]:
                                    break
                                # end if

                                setattr(oNewTracker, level+'_approver', getattr(oLatestTracker, level + '_approver', None))

                                if level == 'psm_config':
                                    continue
                                # end if

                                setattr(oNewTracker, level+'_denied_approval', getattr(oLatestTracker, level + '_denied_approval', None))
                                setattr(oNewTracker, level+'_approved_on', getattr(oLatestTracker, level + '_approved_on', None))
                                setattr(oNewTracker, level+'_comments', getattr(oLatestTracker, level + '_comments', None))
                            # end for

                            oNewTracker.save()
                        # Else, send header back to 'In Process' status
                        else:
                            oHeader.configuration_status = 'In Process'
                            oHeader.save()
                        # end if
                    elif sAction == 'skip':
                        setattr(oLatestTracker, sNeededLevel+'_approver', oRequest.user.username)
                        setattr(oLatestTracker, sNeededLevel+'_approved_on', timezone.datetime(1900,1,1))
                    oLatestTracker.save()

                    if sNeededLevel == 'brd' and sAction == 'approve':
                        oLatestTracker.completed_on = timezone.now()
                        oLatestTracker.save()

                        # Alter configuration status
                        if oHeader.bom_request_type.name in ('Discontinue','Legacy'):
                            oHeader.configuration_status = 'Discontinued'
                        elif oHeader.bom_request_type.name in ('New','Update', 'Replacement'):
                            oHeader.configuration_status = 'Active'
                        elif oHeader.bom_request_type.name == 'Preliminary':
                            oHeader.configuration_status = 'In Process'
                            oHeader.bom_request_type = REF_REQUEST.objects.get(name='New')
                        oHeader.save()

                        if oHeader.configuration_status in ('Discontinued', 'Active') and oHeader.baseline_impacted:
                            aBaselinesCompleted.append(oHeader.baseline_impacted)
                        # end if
                    # end if
                # end if
            # end for

            aBaselinesCompleted = list(set(aBaselinesCompleted))
            aBaselinesToComplete = Baseline.objects.filter(title__in=aBaselinesCompleted)
            for oBaseline in aBaselinesToComplete:
                UpRev(oBaseline)
            # end for
        elif sAction == 'clone':
            oOldHeader = Header.objects.get(pk=aRecords[0])
            oNewHeader = copy.deepcopy(oOldHeader)
            oNewHeader.pk = None
            oNewHeader.configuration_designation = oOldHeader.configuration_designation + '_______CLONE_______'
            oNewHeader.configuration_status = 'In Process'
            if oNewHeader.react_request is None:
                oNewHeader.react_request = ''
            # end if

            if oNewHeader.baseline_impacted:
                oNewHeader.baseline = Baseline_Revision.objects.get(baseline=Baseline.objects.get(title=oNewHeader.baseline_impacted),
                                                                    version=Baseline.objects.get(title=oNewHeader.baseline_impacted).current_inprocess_version)
                oNewHeader.baseline_version = oNewHeader.baseline.version
            # end if

            oNewHeader.save()

            oNewConfig = copy.deepcopy(oOldHeader.configuration)
            oNewConfig.pk = None
            oNewConfig.header = oNewHeader
            oNewConfig.save()

            for oConfigLine in oOldHeader.configuration.configline_set.all():
                oNewLine = copy.deepcopy(oConfigLine)
                oNewLine.pk = None
                oNewLine.config = oNewConfig
                oNewLine.save()

                if hasattr(oConfigLine,'linepricing'):
                    oNewPrice = copy.deepcopy(oConfigLine.linepricing)
                    oNewPrice.pk = None
                    oNewPrice.config_line = oNewLine
                    oNewPrice.save()
                # end if
            # end for

            oRequest.session['existing'] = oNewHeader.pk
            return HttpResponse(reverse('bomconfig:configheader'))
        elif sAction == 'delete':
            Header.objects.filter(pk__in=aRecords).delete()
        elif sAction == 'cancel':
            Header.objects.filter(pk__in=aRecords).update(configuration_status='Cancelled')
        elif sAction in ('hold', 'unhold'):
            for iRecord in aRecords:
                oHeader = Header.objects.get(pk=iRecord)
                oHeader.configuration.PSM_on_hold = not oHeader.configuration.PSM_on_hold
                oHeader.configuration.save()
            #end for
        # end if

        return HttpResponse()
    else:
        raise Http404()
# end def


def Report(oRequest):
    return Default(oRequest, 'BoMConfig/report.html')
# end def
